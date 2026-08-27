from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.utils import timezone
from django.db.models.functions import ExtractYear, Coalesce
from django.db.models import (
    Case,
    CharField,
    DecimalField,
    F,
    Q,
    Sum,
    Value,
    When,
)
from decimal import Decimal
from .models import Trip, TripPayment
from .forms import TripForm, TripPaymentForm
from core.audit import log_action

from master_data.models import Material, VehicleType
from vehicles.models import Vehicle
from labour.models import Labour
from customers.models import Customer
from core.utils import get_safe_next

def _filtered_trips(request):
    """
    Shared queryset builder — trip_list aur trip_export dono SAME filters use karte hain.
    Returns (filtered+annotated queryset, filters dict for context/filename).
    """
    trips = Trip.objects.select_related(
        'customer',
        'vehicle',
        'material',
    ).prefetch_related(
        'drivers',
        'payments',
    ).annotate(
        calculated_received=Coalesce(
            Sum('payments__amount'),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
    ).annotate(
        calculated_status=Case(
            When(
                Q(calculated_received__isnull=True) | Q(calculated_received__lte=0),
                then=Value('UNPAID'),
            ),
            When(
                calculated_received__lt=F('total_amount'),
                then=Value('PARTIAL'),
            ),
            default=Value('PAID'),
            output_field=CharField(),
        ),
    )

    category = request.GET.get('category', '').strip().lower()

    if category == 'tractor':
        trips = trips.filter(
            Q(vehicle__vehicle_type__code='TRACTOR') |
            Q(vehicle__vehicle_type__name__iexact='Tractor') |
            Q(vehicle__registration_number__icontains='Tractor')
        )
    elif category == 'halfton':
        trips = trips.filter(
            Q(vehicle__vehicle_type__code='HALFTON') |
            Q(vehicle__vehicle_type__name__iexact='Halfton') |
            Q(vehicle__registration_number__icontains='Halfton')
        )
    elif category == 'hyva':
        trips = trips.filter(
            Q(vehicle__vehicle_type__code='HYVA') |
            Q(vehicle__vehicle_type__name__iexact='Hyva') |
            Q(vehicle__registration_number__icontains='Hyva')
        )
    elif category == 'jcb':
        trips = trips.filter(
            Q(vehicle__vehicle_type__code='JCB') |
            Q(vehicle__vehicle_type__name__iexact='Jcb') |
            Q(vehicle__registration_number__icontains='Jcb')
        )

    search = request.GET.get('search', '').strip()

    if search:
        trips = trips.filter(
            Q(trip_code__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(vehicle__registration_number__icontains=search) |
            Q(destination__icontains=search) |
            Q(drivers__name__icontains=search)
        ).distinct()

    trip_status = request.GET.get('trip_status', '')

    if trip_status:
        trips = trips.filter(
            trip_status=trip_status
        )

    payment_status = request.GET.get(
        'payment_status',
        ''
    )

    if payment_status:
        statuses = [s.strip() for s in payment_status.split(',') if s.strip()]
        if statuses:
            trips = trips.filter(
                calculated_status__in=statuses
            )

    material_id = request.GET.get('material', '').strip()
    if material_id:
        trips = trips.filter(material_id=material_id)

    vehicle_type_code = request.GET.get('vehicle', '').strip().upper()
    if vehicle_type_code:
        trips = trips.filter(
            Q(vehicle__vehicle_type__code=vehicle_type_code) |
            Q(vehicle__vehicle_type__name__iexact=vehicle_type_code) |
            Q(vehicle__registration_number__icontains=vehicle_type_code)
        )

    destination = request.GET.get('destination', '').strip()
    if destination:
        trips = trips.filter(destination__iexact=destination)

    driver_id = request.GET.get('driver', '').strip()
    if driver_id:
        trips = trips.filter(drivers__id=driver_id).distinct()

    transaction_type = request.GET.get('transaction_type', 'ALL').strip()

    if transaction_type == 'CUSTOMER_DELIVERY':
        trips = trips.filter(transaction_type='CUSTOMER_DELIVERY')
    elif transaction_type == 'INTERNAL_STOCK':
        trips = trips.filter(transaction_type='INTERNAL_STOCK')
    elif transaction_type == 'ALL':
        pass
    else:
        transaction_type = 'CUSTOMER_DELIVERY'
        trips = trips.filter(transaction_type='CUSTOMER_DELIVERY')

    year = request.GET.get('year', '').strip()
    month = request.GET.get('month', '').strip()
    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()

    if year:
        trips = trips.filter(trip_date__year=int(year))
    if month:
        trips = trips.filter(trip_date__month=int(month))
    if from_date:
        trips = trips.filter(trip_date__gte=from_date)
    if to_date:
        trips = trips.filter(trip_date__lte=to_date)

    # DATE ORDER: newest (default, latest upar) / oldest (purani trips upar)
    sort = request.GET.get('sort', '').strip().lower()
    if sort not in ('newest', 'oldest'):
        sort = 'newest'
    if sort == 'oldest':
        trips = trips.order_by('trip_date', 'id')
    else:
        trips = trips.order_by('-trip_date', '-id')

    filters = {
        'search': search,
        'selected_trip_status': trip_status,
        'selected_payment_status': payment_status,
        'selected_transaction_type': transaction_type,
        'selected_material': material_id,
        'selected_vehicle': vehicle_type_code,
        'selected_destination': destination,
        'selected_driver': driver_id,
        'selected_year': year,
        'selected_month': month,
        'from_date': from_date,
        'to_date': to_date,
        'category': category,
        'selected_sort': sort,
    }
    return trips, filters


@login_required(login_url='/login/')
def trip_export(request):
    """
    Filtered trips ki Excel (.xlsx) report + Liquid Glass Preview.
    Hyva / Tractor / JCB / Halfton — sab category pages aur saare filters
    (month/year, date range, search, material, driver, payment status etc.)
    ke saath kaam karta hai. Monthly report ke liye Month+Year filter laga kar Export dabao.
    ?preview=1 par HTML preview fragment return hota hai (modal me dikhta hai).
    """
    import calendar

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    trips, filters = _filtered_trips(request)
    # Export/preview me bhi wahi date-order jo list par chuna hai
    trips_list = list(trips)

    status_labels = dict(Trip.TRIP_STATUS_CHOICES)
    pay_labels = {'UNPAID': 'Unpaid', 'PARTIAL': 'Partial', 'PAID': 'Paid'}

    total_qty = Decimal('0')
    total_rev = Decimal('0')
    total_rec = Decimal('0')
    total_out = Decimal('0')
    table_rows = []

    for t in trips_list:
        qty = t.quantity or Decimal('0')
        received = t.calculated_received or Decimal('0')
        outstanding = t.outstanding_amount or Decimal('0')
        total_qty += qty
        total_rev += t.total_amount or Decimal('0')
        total_rec += received
        total_out += outstanding

        driver_names = [d.name for d in t.drivers.all()]

        table_rows.append({
            'code': t.trip_code or '',
            'date': t.trip_date.strftime('%d-%m-%Y') if t.trip_date else '',
            'customer': t.customer.name if t.customer else (
                'Internal Stock' if t.transaction_type == 'INTERNAL_STOCK' else ''
            ),
            'destination': t.destination or '',
            'vehicle': str(t.vehicle.registration_number) if t.vehicle else '',
            'drivers': ', '.join(driver_names) if driver_names else '-',
            'material': t.material.name if t.material else '',
            'quantity': qty,
            'rate': t.rate or Decimal('0'),
            'revenue': t.total_amount or Decimal('0'),
            'received': received,
            'outstanding': outstanding,
            'status': status_labels.get(t.trip_status, t.trip_status),
            'payment': pay_labels.get(t.calculated_payment_status, t.calculated_payment_status),
        })

    # Filename: category + period aware (e.g. hyva-trips-August-2026-20260825.xlsx)
    cat = filters.get('category') or 'all'
    period = ''
    if filters.get('from_date') or filters.get('to_date'):
        period = "{0}_to_{1}".format(
            filters.get('from_date') or 'start',
            filters.get('to_date') or 'today',
        )
    elif filters.get('selected_month') and filters.get('selected_year'):
        period = "{0}-{1}".format(
            calendar.month_name[int(filters['selected_month'])],
            filters['selected_year'],
        )
    elif filters.get('selected_year'):
        period = str(filters['selected_year'])

    stamp = timezone.localdate().strftime('%Y%m%d')
    filename = "{0}-trips{1}-{2}.xlsx".format(
        cat,
        "-{0}".format(period) if period else '',
        stamp,
    )

    # ---- PREVIEW MODE: HTML fragment for liquid-glass modal ----
    if request.GET.get('preview') == '1':
        context = {
            'rows': table_rows,
            'count': len(table_rows),
            'totals': {
                'qty': total_qty,
                'revenue': total_rev,
                'received': total_rec,
                'outstanding': total_out,
            },
            'filename': filename,
            'querystring': request.GET.urlencode(),
        }
        return render(request, 'trips/_export_preview.html', context)

    # ---- CSV DOWNLOAD (?format=csv) ----
    if request.GET.get('format') == 'csv':
        import csv as csv_module

        csv_filename = "{0}.csv".format(filename[:-len('.xlsx')] if filename.endswith('.xlsx') else filename)
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="{0}"'.format(csv_filename)
        writer = csv_module.writer(response)
        writer.writerow(headers := [
            'Trip Code', 'Date', 'Customer', 'Destination', 'Vehicle No.',
            'Drivers / Labour', 'Material', 'Quantity', 'Rate',
            'Revenue', 'Received', 'Outstanding',
            'Trip Status', 'Payment Status',
        ])
        for r in table_rows:
            writer.writerow([
                r['code'], r['date'], r['customer'], r['destination'], r['vehicle'],
                r['drivers'], r['material'],
                r['quantity'], r['rate'],
                r['revenue'], r['received'], r['outstanding'],
                r['status'], r['payment'],
            ])
        writer.writerow([
            'TOTAL', '', '', '', '', '', '',
            total_qty, '', total_rev, total_rec, total_out, '', ''
        ])
        return response

    # ---- EXCEL (.xlsx) DOWNLOAD ----
    wb = Workbook()
    ws = wb.active
    ws.title = 'Trips Report'

    headers = [
        'Trip Code', 'Date', 'Customer', 'Destination', 'Vehicle No.',
        'Drivers / Labour', 'Material', 'Quantity', 'Rate',
        'Revenue', 'Received', 'Outstanding',
        'Trip Status', 'Payment Status',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in table_rows:
        ws.append([
            r['code'], r['date'], r['customer'], r['destination'], r['vehicle'],
            r['drivers'], r['material'],
            float(r['quantity']), float(r['rate']),
            float(r['revenue']), float(r['received']), float(r['outstanding']),
            r['status'], r['payment'],
        ])

    # TOTALS row
    totals_row = len(table_rows) + 2
    ws.append([
        'TOTAL', '', '', '', '', '', '',
        float(total_qty), '', float(total_rev), float(total_rec), float(total_out), '', ''
    ])
    for cell in ws[totals_row]:
        cell.font = Font(bold=True)

    # Column widths
    widths = [15, 12, 24, 18, 14, 26, 18, 10, 10, 13, 13, 14, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="{0}"'.format(filename)
    wb.save(response)
    return response


@login_required(login_url='/login/')
def trip_list(request):
    trips, filters = _filtered_trips(request)

    materials_list = Material.objects.filter(is_active=True).order_by('name')

    vehicles_list = VehicleType.objects.filter(is_active=True).order_by('name')

    destinations_list = list(
        Trip.objects.exclude(destination__isnull=True)
        .exclude(destination='')
        .values_list('destination', flat=True)
        .distinct()
        .order_by('destination')
    )

    drivers_list = Labour.objects.filter(is_active=True).order_by('name')

    available_years = list(
        Trip.objects.annotate(y=ExtractYear('trip_date'))
        .values_list('y', flat=True)
        .distinct()
        .order_by('-y')
    )
    current_year = timezone.localdate().year
    if current_year not in available_years:
        available_years.insert(0, current_year)

    month_choices = [
        ('1', 'January'),
        ('2', 'February'),
        ('3', 'March'),
        ('4', 'April'),
        ('5', 'May'),
        ('6', 'June'),
        ('7', 'July'),
        ('8', 'August'),
        ('9', 'September'),
        ('10', 'October'),
        ('11', 'November'),
        ('12', 'December'),
    ]

    trips_list = list(trips)
    summary = {
        'total_count': len(trips_list),
        'total_qty': sum((t.quantity or Decimal('0')) for t in trips_list),
        'total_revenue': sum(t.total_amount for t in trips_list) if trips_list else Decimal('0'),
        'total_received': sum((t.calculated_received or Decimal('0')) for t in trips_list) if trips_list else Decimal('0'),
        'total_pending': sum((t.outstanding_amount or Decimal('0')) for t in trips_list) if trips_list else Decimal('0'),
        'paid_count': len([t for t in trips_list if t.calculated_status == 'PAID']),
        'pending_count': len([t for t in trips_list if t.calculated_status in ['UNPAID', 'PARTIAL']]),
    }

    context = {
        'trips': trips_list,
        'summary': summary,
        'materials_list': materials_list,
        'vehicles_list': vehicles_list,
        'destinations_list': destinations_list,
        'drivers_list': drivers_list,
        'available_years': available_years,
        'month_choices': month_choices,
        'trip_status_choices': Trip.TRIP_STATUS_CHOICES,
        'payment_status_choices': Trip.PAYMENT_STATUS_CHOICES,
        **filters,
    }

    return render(
        request,
        'trips/trip_list.html',
        context
    )
from django.contrib import messages

@login_required(login_url='/login/')
def trip_create(request):
    if request.method == 'POST':
        form = TripForm(request.POST)
        if form.is_valid():
            trip = form.save()

            # Save driver trip counts breakdown
            driver_counts = {}
            for d in trip.drivers.all():
                param_name = f'driver_count_{d.id}'
                if param_name in request.POST:
                    val = request.POST.get(param_name, '').strip()
                    if val:
                        try:
                            driver_counts[str(d.id)] = float(val) if '.' in val else int(val)
                        except ValueError:
                            pass
            trip.driver_trip_counts = driver_counts

            # Server-side quantity auto-sum: driver allocations ka total hi quantity hai
            counts_sum = sum(Decimal(str(v)) for v in driver_counts.values())
            update_fields = ['driver_trip_counts']
            if driver_counts and counts_sum > 0 and trip.quantity != counts_sum:
                trip.quantity = counts_sum.quantize(Decimal('0.01'))
                update_fields.append('quantity')
                bhatta = trip.driver_bhatta or Decimal('0.00')
                trip.total_amount = (trip.quantity * trip.rate) + bhatta
                update_fields.append('total_amount')
            trip.save(update_fields=update_fields)

            messages.success(request, f"Trip {trip.trip_code} created successfully!")

            # Inward (VENDOR_SUPPLY) → redirect to link customer page
            if trip.transaction_type == 'VENDOR_SUPPLY':
                return redirect('trips:link_customer', trip_id=trip.id)

            if request.POST.get('save_and_view_statement') == '1' and trip.customer:
                return redirect('ledger:customer_statement', customer_id=trip.customer.id)

            return redirect(
                'trips:detail',
                trip_id=trip.id
            )
        else:
            err_list = [f"{field}: {', '.join(errs)}" for field, errs in form.errors.items()]
            messages.error(request, f"Failed to create trip: {'; '.join(err_list)}")
    else:
        form = TripForm()

    context = {
        'form': form,
        'page_title': 'Add Trip',
    }

    return render(
        request,
        'trips/trip_create.html',
        context
    )

@login_required(login_url='/login/')
def trip_edit(request, trip_id):
    trip = get_object_or_404(
        Trip,
        pk=trip_id
    )

    if request.method == 'POST':
        form = TripForm(
            request.POST,
            instance=trip
        )
        if form.is_valid():
            trip = form.save()

            # Save driver trip counts breakdown
            driver_counts = {}
            for d in trip.drivers.all():
                param_name = f'driver_count_{d.id}'
                if param_name in request.POST:
                    val = request.POST.get(param_name, '').strip()
                    if val:
                        try:
                            driver_counts[str(d.id)] = float(val) if '.' in val else int(val)
                        except ValueError:
                            pass
            trip.driver_trip_counts = driver_counts

            # Server-side quantity auto-sum: driver allocations ka total hi quantity hai
            # (JS fail ho ya user manually badal de — data hamesha consistent rahega)
            counts_sum = sum(Decimal(str(v)) for v in driver_counts.values())
            update_fields = ['driver_trip_counts']
            if driver_counts and counts_sum > 0 and trip.quantity != counts_sum:
                trip.quantity = counts_sum.quantize(Decimal('0.01'))
                update_fields.append('quantity')
                # Quantity change se total_amount bhi update hona chahiye
                bhatta = trip.driver_bhatta or Decimal('0.00')
                trip.total_amount = (trip.quantity * trip.rate) + bhatta
                update_fields.append('total_amount')
            trip.save(update_fields=update_fields)

            messages.success(request, f"Trip {trip.trip_code} updated successfully!")

            if request.POST.get('save_and_view_statement') == '1' and trip.customer:
                return redirect('ledger:customer_statement', customer_id=trip.customer.id)

            # Jahan se aaya tha (statement/detail) wahin wapas — ?next= chain
            nxt = request.POST.get('next') or request.GET.get('next')
            if nxt and nxt.startswith('/') and not nxt.startswith('//'):
                return redirect(nxt)

            # Default: trip detail page
            return redirect('trips:detail', trip_id=trip.id)
        else:
            # Invalid save — page wapas edit par dikhega (errors ke saath).
            # Log me exact reason capture karo taaki diagnose easy ho.
            import logging
            logging.getLogger('trips').warning(
                "Trip edit FAILED | trip_id=%s | errors=%s",
                trip_id, dict(form.errors),
            )
            err_list = [f"{field}: {', '.join(errs)}" for field, errs in form.errors.items()]
            messages.error(request, f"Failed to update trip: {'; '.join(err_list)}")
    else:
        form = TripForm(
            instance=trip
        )

    context = {
        'form': form,
        'trip': trip,
        'page_title': 'Edit Trip',
    }

    return render(
        request,
        'trips/trip_edit.html',
        context
    )

@login_required(login_url='/login/')
def trip_delete(request, trip_id):

    trip = get_object_or_404(
        Trip,
        pk=trip_id
    )

    from core.utils import get_safe_next, get_safe_next_or_referer
    from django.urls import reverse

    if request.method == 'POST':
        try:
            code = trip.trip_code
            deleted_amount = trip.total_amount
            trip.delete()
            log_action(
                request,
                'TRIP_DELETE',
                model_name='Trip',
                object_repr=code,
                details=f"Deleted trip of \u20B9{deleted_amount}",
            )
            messages.success(request, f"Trip {code} deleted successfully!")
            # Return to where the user came from (e.g. customer statement),
            # passed through the form's hidden 'next' field.
            target = get_safe_next(request, reverse('trips:list'))
            # Never return to the pages of the trip we just deleted
            if f"/trips/{trip_id}/" in target or f"/trips/{trip_id}/" in target + "/":
                target = reverse('trips:list')
            return redirect(target)
        except ProtectedError:
            context = {
                'trip': trip,
                'delete_error': (
                    'This trip cannot be deleted because '
                    'expenses or other records are linked to it.'
                ),
            }
            return render(
                request,
                'trips/trip_delete.html',
                context
            )

    # Where the user came from (statement / trip detail), for Cancel + post-delete return
    back_url = get_safe_next_or_referer(request, reverse('trips:list'))

    context = {
        'trip': trip,
        'back_url': back_url,
    }

    return render(
        request,
        'trips/trip_delete.html',
        context
    )


@login_required(login_url='/login/')
def trip_payment_create(request, trip_id):
    trip = get_object_or_404(
        Trip,
        pk=trip_id
    )
    next_url = get_safe_next(request, f'/trips/{trip.id}/')

    if request.method == 'POST':
        payment_instance = TripPayment(
            trip=trip
        )
        form = TripPaymentForm(
            request.POST,
            instance=payment_instance,
            trip=trip
        )
        if form.is_valid():
            form.save()
            log_action(
                request,
                'PAYMENT_CREATE',
                obj=form.instance,
                details=f"Trip {trip.trip_code} | Amount \u20B9{form.instance.amount} | Date {form.instance.payment_date}",
            )
            messages.success(request, f"Payment of ₹{payment_instance.amount} added to trip {trip.trip_code}!")
            if next_url:
                return redirect(next_url)
            return redirect(
                'trips:detail',
                trip_id=trip.id
            )
    else:
        payment_instance = TripPayment(
            trip=trip
        )
        form = TripPaymentForm(
            instance=payment_instance,
            trip=trip
        )

    context = {
        'form': form,
        'trip': trip,
        'next_url': next_url,
    }
    return render(
        request,
        'trips/trip_payment_create.html',
        context
    )

@login_required(login_url='/login/')
def trip_payment_edit(request, payment_id):
    payment = get_object_or_404(
        TripPayment.objects.select_related('trip', 'customer'),
        pk=payment_id
    )
    trip = payment.trip
    next_url = get_safe_next(
        request,
        f'/trips/{trip.id}/' if trip else '/trips/',
    )

    if request.method == 'POST':
        form = TripPaymentForm(
            request.POST,
            instance=payment,
            trip=trip
        )
        if form.is_valid():
            form.save()
            log_action(
                request,
                'PAYMENT_EDIT',
                obj=form.instance,
                details=f"Trip {trip.trip_code if trip else '-'} | Amount \u20B9{form.instance.amount} | Date {form.instance.payment_date}",
            )
            messages.success(request, f"Payment updated successfully!")
            if next_url:
                return redirect(next_url)
            if trip:
                return redirect('trips:detail', trip_id=trip.id)
            elif payment.customer:
                return redirect('ledger:customer_statement', customer_id=payment.customer.id)
            return redirect('trips:list')
    else:
        form = TripPaymentForm(
            instance=payment,
            trip=trip
        )

    context = {
        'form': form,
        'trip': trip,
        'payment': payment,
        'page_title': 'Edit Payment',
        'next_url': next_url,
    }
    return render(
        request,
        'trips/trip_payment_edit.html',
        context
    )

@login_required(login_url='/login/')
def trip_payment_delete(request, payment_id):
    payment = get_object_or_404(
        TripPayment.objects.select_related('trip', 'customer'),
        pk=payment_id
    )
    trip = payment.trip
    next_url = get_safe_next(
        request,
        f'/trips/{trip.id}/' if trip else '/trips/',
    )

    if request.method == 'POST':
        customer = payment.customer
        deleted_amount = payment.amount
        payment.delete()
        log_action(
            request,
            'PAYMENT_DELETE',
            model_name='TripPayment',
            object_repr=f"{payment.payment_code or ''} {deleted_amount}".strip(),
            details=f"Deleted payment of \u20B9{deleted_amount} (trip {trip.trip_code if trip else '-'})",
        )
        messages.success(request, f"Payment removed successfully!")
        if next_url:
            return redirect(next_url)
        if trip:
            return redirect('trips:detail', trip_id=trip.id)
        elif customer:
            return redirect('ledger:customer_statement', customer_id=customer.id)
        return redirect('trips:list')

    context = {
        'payment': payment,
        'trip': trip,
        'next_url': next_url,
    }
    return render(
        request,
        'trips/trip_payment_delete.html',
        context
    )

@login_required(login_url='/login/')
def trip_detail(request, trip_id):

    trip = get_object_or_404(
        Trip.objects.select_related(
            'customer',
            'vehicle',
            'material',
        ).prefetch_related(
            'drivers',
            'payments__payment_method',
        ),
        pk=trip_id,
    )

    from core.utils import get_safe_next_or_referer
    from django.urls import reverse

    # Smart back: return to where the user came from (e.g. customer statement),
    # falling back to the trips list only when there is no internal referrer.
    back_url = get_safe_next_or_referer(
        request,
        reverse('trips:list'),
    )

    context = {
        'trip': trip,
        'payments': trip.payments.all(),
        'back_url': back_url,
    }

    return render(
        request,
        'trips/trip_detail.html',
        context
    )

@login_required(login_url='/login/')
def trip_link_customer(request, trip_id):
    """
    Inward (VENDOR_SUPPLY) trip se Outward (Customer Delivery) trip create karta hai.
    Sirf VENDOR_SUPPLY trips ke liye available.
    """
    from decimal import Decimal, InvalidOperation

    trip = get_object_or_404(Trip, pk=trip_id)

    if trip.transaction_type != 'VENDOR_SUPPLY':
        messages.error(request, "Sirf Inward (Vendor Supply) trips customer se link ho sakti hain.")
        return redirect('trips:detail', trip_id=trip.id)

    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        rate_str = request.POST.get('rate', '').strip()

        if not customer_id or not rate_str:
            messages.error(request, "Customer aur Rate dono required hain.")
            return redirect('trips:link_customer', trip_id=trip.id)

        try:
            rate_val = Decimal(rate_str)
            if rate_val <= 0:
                raise ValueError
        except (ValueError, InvalidOperation):
            messages.error(request, "Rate valid positive number hona chahiye.")
            return redirect('trips:link_customer', trip_id=trip.id)

        customer = get_object_or_404(Customer, pk=customer_id, is_active=True)

        # Outward trip create
        outward = Trip(
            transaction_type='CUSTOMER_DELIVERY',
            customer=customer,
            material=trip.material,
            quantity=trip.quantity,
            rate=rate_val,
            vehicle=trip.vehicle,
            destination=trip.destination,
            trip_date=trip.trip_date,
            driver_bhatta=Decimal('0.00'),  # sale amount pure
            notes=f"Outward sale of inward {trip.trip_code}",
            linked_inward_trip=trip,
        )
        outward.save()
        # Copy drivers and their trip counts
        outward.drivers.set(trip.drivers.all())
        outward.driver_trip_counts = trip.driver_trip_counts
        outward.save(update_fields=['driver_trip_counts'])

        messages.success(request, f"Outward trip {outward.trip_code} created for {customer.name}!")
        return redirect('ledger:customer_statement', customer_id=customer.id)

    # GET: form
    customers = Customer.objects.filter(is_active=True).order_by('name')
    default_rate = round((trip.total_amount or 0) / (trip.quantity or 1), 2)

    context = {
        'inward': trip,
        'customers': customers,
        'default_rate': default_rate,
    }
    return render(request, 'trips/trip_link_customer.html', context)


@login_required(login_url='/login/')
def trip_quickfill(request):
    from django.http import JsonResponse

    customer_id = request.GET.get('customer', '').strip()
    material_id = request.GET.get('material', '').strip()

    if not customer_id.isdigit():
        return JsonResponse({'found': False})

    qs = Trip.objects.filter(customer_id=int(customer_id))

    if material_id.isdigit():
        match = (
            qs.filter(material_id=int(material_id))
            .order_by('-trip_date', '-id')
            .values('rate', 'quantity', 'vehicle_id', 'destination')
            .first()
        )
        if match:
            return JsonResponse({'found': True, 'matched_on': 'material', **match})

    last_any = (
        qs.order_by('-trip_date', '-id')
        .values('rate', 'quantity', 'vehicle_id', 'destination', 'material_id')
        .first()
    )
    if last_any:
        return JsonResponse({'found': True, 'matched_on': 'customer', **last_any})

    return JsonResponse({'found': False})
