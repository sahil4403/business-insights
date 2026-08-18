from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.db import transaction, models

from customers.models import Customer
from labour.models import Labour
from master_data.models import Material, PaymentMethod
from trips.models import Trip, TripPayment


TRIP_FILES = {
    "HYVA_WHITE_SAND": (
        "historical_data/Feb_to_Dec_Hyva_White_Sand_CLEAN.xlsx",
        "CUSTOMER_DELIVERY",
    ),
    "HYVA_FLY_ASH": (
        "historical_data/Feb_to_Dec_Hyva_Fly_Ash_CLEAN.xlsx",
        "CUSTOMER_DELIVERY",
    ),
    "HALFTON_WHITE_SAND": (
        "historical_data/Feb_to_Dec_Halfton_White_Sand_CLEAN.xlsx",
        "CUSTOMER_DELIVERY",
    ),
    "TRACTOR_WHITE_SAND": (
        "historical_data/Feb_to_Dec_Tractor_White_Sand_CLEAN.xlsx",
        None,
    ),
}


PAYMENT_FILES = {
    "HYVA_WHITE_SAND": (
        "historical_data/Feb_to_Dec_Hyva_White_Sand_Payment_CLEAN.xlsx",
    ),
    "HYVA_FLY_ASH": (
        "historical_data/Feb_to_Dec_Hyva_Fly_Ash_Payment_CLEAN.xlsx",
    ),
    "HALFTON_WHITE_SAND": (
        "historical_data/Feb_to_Dec_Halfton_White_Sand_Payment_CLEAN.xlsx",
    ),
    "TRACTOR_WHITE_SAND": (
        "historical_data/Feb_to_Dec_Tractor_White_Sand_Payment_CLEAN.xlsx",
    ),
}


def clean(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    return str(value).strip()


def decimal_value(value):
    if value in (None, ""):
        return None

    if isinstance(value, Decimal):
        return value

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return None


def normalize_name(value):
    return " ".join(clean(value).lower().split())


def as_date(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return value


class Command(BaseCommand):

    help = "Import February to December historical trips and payments"

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate February-December import without changing database.",
        )

    def handle(self, *args, **options):

        dry_run = options["dry_run"]

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("FEBRUARY - DECEMBER HISTORICAL IMPORT")
        self.stdout.write("=" * 70)

        if dry_run:
            self.stdout.write(
                self.style.WARNING(
                    "DRY RUN - NO DATABASE CHANGES WILL BE MADE"
                )
            )
        else:
            self.stdout.write(
                self.style.ERROR(
                    "LIVE IMPORT - DATABASE WILL BE CHANGED"
                )
            )

        # ---------------------------------------------------------
        # MASTER DATA
        # ---------------------------------------------------------

        customers = {
            clean(obj.customer_code): obj
            for obj in Customer.objects.all()
            if obj.customer_code
        }

        materials = {
            normalize_name(obj.name): obj
            for obj in Material.objects.all()
        }

        labour_by_name = {
            normalize_name(obj.name): obj
            for obj in Labour.objects.all()
        }

        unknown_payment_method = PaymentMethod.objects.filter(
            code="UNKNOWN"
        ).first()

        if not unknown_payment_method:
            raise RuntimeError(
                "PaymentMethod with code='UNKNOWN' does not exist."
            )

        if "CUST-033" not in customers:
            raise RuntimeError(
                "CUST-033 Customer Unknown is missing."
            )

        self.stdout.write(
            f"Customers loaded : {len(customers)}"
        )
        self.stdout.write(
            f"Materials loaded : {len(materials)}"
        )
        self.stdout.write(
            f"Labour loaded    : {len(labour_by_name)}"
        )

        # ---------------------------------------------------------
        # EXISTING DATABASE CODES
        # ---------------------------------------------------------

        existing_trip_codes = set(
            Trip.objects.values_list(
                "trip_code",
                flat=True,
            )
        )

        existing_payment_codes = set(
            TripPayment.objects.exclude(
                payment_code__isnull=True
            ).exclude(
                payment_code=""
            ).values_list(
                "payment_code",
                flat=True,
            )
        )

        # ---------------------------------------------------------
        # TRIPS
        # ---------------------------------------------------------

        trip_rows = []
        trip_codes = set()
        trip_received = defaultdict(Decimal)

        total_trip_amount = Decimal("0")
        total_received = Decimal("0")

        errors = []

        self.stdout.write("")
        self.stdout.write("TRIP VALIDATION")
        self.stdout.write("-" * 70)

        for category, file_info in TRIP_FILES.items():

            file_path, default_transaction_type = file_info

            self.stdout.write(
                f"\n[{category}]"
            )

            wb = openpyxl.load_workbook(
                file_path,
                data_only=True,
            )

            ws = wb.active

            headers = {
                clean(ws.cell(1, col).value): col
                for col in range(1, ws.max_column + 1)
            }

            required_headers = [
                "Trip Code",
                "Trip Date",
                "Customer Code",
                "Customer Name",
                "Destination",
                "Quantity",
                "Rate",
                "Total Amount",
                "Received",
                "Balance",
                "Payment Status",
                "Vechile Type",
            ]

            for header in required_headers:

                missing_headers = []

                for header in required_headers:

                    if header not in headers:
                        missing_headers.append(header)

                if missing_headers:
                    errors.append(
                        f"{category}: Missing columns: "
                        f"{', '.join(missing_headers)}"
                    )

                    continue

            for row in range(2, ws.max_row + 1):

                trip_code = clean(
                    ws.cell(
                        row,
                        headers["Trip Code"]
                    ).value
                )

                if not trip_code:
                    continue

                if trip_code in trip_codes:

                    errors.append(
                        f"Duplicate Trip Code: {trip_code}"
                    )

                if trip_code in existing_trip_codes:

                    errors.append(
                        f"{trip_code}: Trip Code already exists in DB."
                    )

                trip_codes.add(trip_code)

                trip_date = as_date(
                    ws.cell(
                        row,
                        headers["Trip Date"]
                    ).value
                )

                customer_code = clean(
                    ws.cell(
                        row,
                        headers["Customer Code"]
                    ).value
                )

                if customer_code == "0":
                    customer_code = "CUST-033"

                customer_name = clean(
                    ws.cell(
                        row,
                        headers["Customer Name"]
                    ).value
                )

                destination = clean(
                    ws.cell(
                        row,
                        headers["Destination"]
                    ).value
                )

                quantity = decimal_value(
                    ws.cell(
                        row,
                        headers["Quantity"]
                    ).value
                )

                rate = decimal_value(
                    ws.cell(
                        row,
                        headers["Rate"]
                    ).value
                )

                total_amount = decimal_value(
                    ws.cell(
                        row,
                        headers["Total Amount"]
                    ).value
                )

                received = decimal_value(
                    ws.cell(
                        row,
                        headers["Received"]
                    ).value
                )

                balance = decimal_value(
                    ws.cell(
                        row,
                        headers["Balance"]
                    ).value
                )

                vehicle_type = clean(
                    ws.cell(
                        row,
                        headers["Vechile Type"]
                    ).value
                )

                material_type = clean(
                    ws.cell(
                        row,
                        headers.get(
                            "Material Type",
                            0
                        )
                    ).value
                ) if "Material Type" in headers else ""

                if not material_type:

                    if category == "HYVA_FLY_ASH":
                        material_type = "Fly Ash"

                    else:
                        material_type = "White Sand"

                if default_transaction_type:

                    transaction_type = (
                        default_transaction_type
                    )

                else:

                    raw_type = clean(
                        ws.cell(
                            row,
                            headers["Transaction Type"]
                        ).value
                    )

                    if raw_type.upper() == "INTERNAL STOCK":

                        transaction_type = "INTERNAL_STOCK"

                    else:

                        transaction_type = "CUSTOMER_DELIVERY"

                # -------------------------------------------------
                # INTERNAL STOCK RATE NORMALIZATION
                # -------------------------------------------------

                if transaction_type == "INTERNAL_STOCK" and rate is None:
                    rate = Decimal("0.00")

                if transaction_type == "INTERNAL_STOCK" and total_amount is None:
                    total_amount = Decimal("0.00")

                if received is None:
                    received = Decimal("0.00")

                if balance is None and total_amount is not None:
                    balance = (
                            total_amount - received
                    ).quantize(
                        Decimal("0.01")
                    )

                # -------------------------------------------------
                # CUSTOMER
                # -------------------------------------------------

                customer = None

                if transaction_type == "INTERNAL_STOCK":

                    customer = None

                else:

                    customer = customers.get(
                        customer_code
                    )

                    if not customer:

                        errors.append(
                            f"{trip_code}: Customer "
                            f"{customer_code} not found in DB."
                        )

                # -------------------------------------------------
                # MATERIAL
                # -------------------------------------------------

                material = materials.get(
                    normalize_name(material_type)
                )

                if not material:

                    errors.append(
                        f"{trip_code}: Material "
                        f"'{material_type}' not found in DB."
                    )

                # -------------------------------------------------
                # DRIVER
                # -------------------------------------------------

                driver_names = []

                for driver_column in (
                    "Driver 1",
                    "Driver 2",
                    "Driver 3",
                ):

                    if driver_column not in headers:
                        continue

                    driver_name = clean(
                        ws.cell(
                            row,
                            headers[driver_column]
                        ).value
                    )

                    if not driver_name:
                        continue

                    driver = labour_by_name.get(
                        normalize_name(driver_name)
                    )

                    if not driver:

                        errors.append(
                            f"{trip_code}: Driver/Labour "
                            f"'{driver_name}' not found in DB."
                        )

                    else:

                        driver_names.append(
                            driver
                        )

                # -------------------------------------------------
                # DATE
                # -------------------------------------------------

                if not isinstance(trip_date, date):

                    errors.append(
                        f"{trip_code}: Invalid Trip Date."
                    )

                # -------------------------------------------------
                # TOTAL
                # -------------------------------------------------

                if (
                    quantity is not None
                    and rate is not None
                    and total_amount is not None
                ):

                    calculated_total = (
                        quantity * rate
                    ).quantize(
                        Decimal("0.01")
                    )

                    if calculated_total != total_amount:

                        errors.append(
                            f"{trip_code}: Total mismatch. "
                            f"Excel={total_amount}, "
                            f"Calculated={calculated_total}"
                        )

                # -------------------------------------------------
                # BALANCE
                # -------------------------------------------------

                if (
                    total_amount is not None
                    and received is not None
                    and balance is not None
                ):

                    calculated_balance = (
                        total_amount - received
                    ).quantize(
                        Decimal("0.01")
                    )

                    if calculated_balance != balance:

                        errors.append(
                            f"{trip_code}: Balance mismatch. "
                            f"Excel={balance}, "
                            f"Calculated={calculated_balance}"
                        )

                # -------------------------------------------------
                # RECEIVED
                # -------------------------------------------------

                if received is not None:

                    trip_received[
                        trip_code
                    ] += received

                    total_received += received

                if total_amount is not None:

                    total_trip_amount += total_amount

                trip_rows.append({
                    "category": category,
                    "trip_code": trip_code,
                    "trip_date": trip_date,
                    "customer": customer,
                    "destination": destination,
                    "vehicle_type": vehicle_type,
                    "material": material,
                    "transaction_type": transaction_type,
                    "driver_names": driver_names,
                    "quantity": quantity,
                    "rate": rate,
                    "total_amount": total_amount,
                    "received": received,
                    "balance": balance,
                })

            self.stdout.write(
                f"Rows checked: {ws.max_row - 1}"
            )

        # ---------------------------------------------------------
        # PAYMENTS
        # ---------------------------------------------------------

        payment_rows = []
        payment_codes = set()
        payment_trip_amount = defaultdict(Decimal)

        total_payment_amount = Decimal("0")

        self.stdout.write("")
        self.stdout.write("PAYMENT VALIDATION")
        self.stdout.write("-" * 70)

        for category, file_info in PAYMENT_FILES.items():

            file_path = file_info[0]

            self.stdout.write(
                f"\n[{category}]"
            )

            wb = openpyxl.load_workbook(
                file_path,
                data_only=True,
            )

            ws = wb.active

            headers = {
                clean(ws.cell(1, col).value): col
                for col in range(1, ws.max_column + 1)
            }

            for row in range(2, ws.max_row + 1):

                payment_code = clean(
                    ws.cell(
                        row,
                        headers["Payment Code"]
                    ).value
                )

                if not payment_code:
                    continue

                if payment_code in payment_codes:

                    errors.append(
                        f"Duplicate Payment Code: "
                        f"{payment_code}"
                    )

                if payment_code in existing_payment_codes:

                    errors.append(
                        f"{payment_code}: Payment Code "
                        f"already exists in DB."
                    )

                payment_codes.add(payment_code)

                trip_code = clean(
                    ws.cell(
                        row,
                        headers["Trip Code"]
                    ).value
                )

                amount = decimal_value(
                    ws.cell(
                        row,
                        headers["Amount"]
                    ).value
                )

                payment_date = as_date(
                    ws.cell(
                        row,
                        headers["Payment Date"]
                    ).value
                )

                payment_method_name = clean(
                    ws.cell(
                        row,
                        headers["Payment Method"]
                    ).value
                )

                reference_number = clean(
                    ws.cell(
                        row,
                        headers["Reference Number"]
                    ).value
                )

                notes = clean(
                    ws.cell(
                        row,
                        headers["Notes"]
                    ).value
                )

                if trip_code not in trip_codes:

                    errors.append(
                        f"{payment_code}: Trip Code "
                        f"{trip_code} not found in Feb-Dec trips."
                    )

                trip_row = next(
                    (
                        item
                        for item in trip_rows
                        if item["trip_code"] == trip_code
                    ),
                    None
                )

                if trip_row:

                    if (
                        trip_row["transaction_type"]
                        == "INTERNAL_STOCK"
                    ):

                        errors.append(
                            f"{payment_code}: Payment linked "
                            f"to Internal Stock trip {trip_code}."
                        )

                if amount is None or amount <= 0:

                    errors.append(
                        f"{payment_code}: Invalid payment amount."
                    )

                    continue

                if payment_method_name:
                    payment_method = PaymentMethod.objects.filter(
                        name__iexact=payment_method_name
                    ).first()

                if payment_method_name:

                    payment_method = PaymentMethod.objects.filter(
                        name__iexact=payment_method_name
                    ).first()

                    if not payment_method:

                        errors.append(
                            f"{payment_code}: Payment Method "
                            f"'{payment_method_name}' not found in DB."
                        )

                else:

                    payment_method = unknown_payment_method

                payment_trip_amount[
                    trip_code
                ] += amount

                total_payment_amount += amount

                payment_rows.append({
                    "category": category,
                    "payment_code": payment_code,
                    "trip_code": trip_code,
                    "payment_date": payment_date,
                    "amount": amount,
                    "payment_method": payment_method,
                    "reference_number": reference_number,
                    "notes": notes,
                })

        # ---------------------------------------------------------
        # RECONCILIATION
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("PAYMENT RECONCILIATION")
        self.stdout.write("-" * 70)

        for trip_code in sorted(
            set(trip_received.keys())
            | set(payment_trip_amount.keys())
        ):

            received = trip_received.get(
                trip_code,
                Decimal("0")
            )

            payments = payment_trip_amount.get(
                trip_code,
                Decimal("0")
            )

            if received != payments:

                errors.append(
                    f"{trip_code}: Received/Payment mismatch. "
                    f"Trip Received={received}, "
                    f"Payments={payments}"
                )

        # ---------------------------------------------------------
        # SUMMARY
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("IMPORT SUMMARY")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Trips to process    : {len(trip_rows)}"
        )

        self.stdout.write(
            f"Payments to process : {len(payment_rows)}"
        )

        self.stdout.write(
            f"Trip amount         : ₹{total_trip_amount:,.2f}"
        )

        self.stdout.write(
            f"Received            : ₹{total_received:,.2f}"
        )

        self.stdout.write(
            f"Payment amount      : ₹{total_payment_amount:,.2f}"
        )

        self.stdout.write(
            f"Errors              : {len(errors)}"
        )

        if errors:

            self.stdout.write("")
            self.stdout.write(
                self.style.ERROR(
                    "IMPORT VALIDATION FAILED"
                )
            )

            for error in errors[:100]:

                self.stdout.write(
                    self.style.ERROR(
                        f"- {error}"
                    )
                )

            if len(errors) > 100:

                self.stdout.write(
                    self.style.ERROR(
                        f"... and {len(errors) - 100} more errors."
                    )
                )

            raise RuntimeError(
                "Import stopped because validation errors were found."
            )

        # ---------------------------------------------------------
        # DRY RUN
        # ---------------------------------------------------------

        if dry_run:

            self.stdout.write("")
            self.stdout.write(
                self.style.SUCCESS(
                    "DRY RUN PASSED - NO DATABASE CHANGES WERE MADE."
                )
            )

            return

        # ---------------------------------------------------------
        # LIVE IMPORT
        # ---------------------------------------------------------

        created_trips = 0
        created_payments = 0

        with transaction.atomic():

            trip_objects = {}

            for row in trip_rows:

                trip = Trip.objects.create(
                    trip_code=row["trip_code"],
                    trip_date=row["trip_date"],
                    customer=row["customer"],
                    destination=row["destination"],
                    material=row["material"],
                    transaction_type=row["transaction_type"],
                    quantity=row["quantity"],
                    rate=row["rate"],
                    total_amount=row["total_amount"] or Decimal("0"),
                    trip_status="COMPLETED",
                    payment_status="UNPAID",
                )

                for driver in row["driver_names"]:

                    trip.drivers.add(driver)

                trip_objects[
                    row["trip_code"]
                ] = trip

                created_trips += 1

            for row in payment_rows:

                trip = trip_objects[
                    row["trip_code"]
                ]

                TripPayment.objects.create(
                    trip=trip,
                    payment_code=row["payment_code"],
                    payment_date=row["payment_date"],
                    amount=row["amount"],
                    payment_method=row["payment_method"],
                    reference_number=row["reference_number"],
                    notes=row["notes"],
                )

                created_payments += 1

            # -----------------------------------------------------
            # FINAL DB VERIFICATION
            # -----------------------------------------------------

            db_trip_amount = (
                Trip.objects.filter(
                    trip_code__in=trip_codes
                ).aggregate(
                    total=models.Sum("total_amount")
                )["total"]
                or Decimal("0")
            )

            db_payment_amount = (
                TripPayment.objects.filter(
                    payment_code__in=payment_codes
                ).aggregate(
                    total=models.Sum("amount")
                )["total"]
                or Decimal("0")
            )

            if db_trip_amount != total_trip_amount:

                raise RuntimeError(
                    "DB trip amount verification failed."
                )

            if db_payment_amount != total_payment_amount:

                raise RuntimeError(
                    "DB payment amount verification failed."
                )

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("IMPORT COMPLETE")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Trips created       : {created_trips}"
        )

        self.stdout.write(
            f"Payments created    : {created_payments}"
        )

        self.stdout.write(
            f"DB trip amount      : ₹{db_trip_amount:,.2f}"
        )

        self.stdout.write(
            f"DB payment amount   : ₹{db_payment_amount:,.2f}"
        )

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                "POST-IMPORT VERIFICATION PASSED."
            )
        )