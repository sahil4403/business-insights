from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from customers.models import Customer
from labour.models import Labour
from master_data.models import Material, PaymentMethod
from trips.models import Trip, TripPayment


TRIPS_FILE = "historical_data/All Hyva Trips Clean Data.xlsx"
PAYMENTS_FILE = "historical_data/January_Payments_Clean_v2.xlsx"

TRIP_SHEETS = {
    "Hyva Trips  White Sand": "CUSTOMER_DELIVERY",
    "Halfton trips": "CUSTOMER_DELIVERY",
    "Hyva Trip Fly Ash": "CUSTOMER_DELIVERY",
    "Tractor": None,
}

PAYMENT_SHEETS = {
    "Hyva White Sand Payment",
    "Halfton White Sand",
    "Tractor",
    "Hyva Fly Ash Payment",
}


def clean(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    return str(value).strip()


def decimal_value(value):
    if value in (None, ""):
        return None

    if isinstance(value, Decimal):
        return value

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return None


def normalize_name(value):
    return " ".join(clean(value).lower().split())


def as_date(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return value


class Command(BaseCommand):
    help = "Import January historical trips and payments"

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate the complete import without changing the database.",
        )

    def handle(self, *args, **options):
        dry_run = options["dry_run"]

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("JANUARY HISTORICAL IMPORT")
        self.stdout.write("=" * 70)

        if dry_run:
            self.stdout.write(
                self.style.WARNING(
                    "DRY RUN - NO DATABASE CHANGES WILL BE MADE"
                )
            )
        else:
            self.stdout.write(
                self.style.WARNING(
                    "LIVE IMPORT - DATABASE WILL BE CHANGED"
                )
            )

        # ---------------------------------------------------------
        # LOAD MASTER DATA
        # ---------------------------------------------------------

        customers = {
            clean(obj.customer_code): obj
            for obj in Customer.objects.all()
            if obj.customer_code
        }

        materials = {
            normalize_name(obj.name): obj
            for obj in Material.objects.all()
        }

        labour_by_name = {
            normalize_name(obj.name): obj
            for obj in Labour.objects.all()
        }

        unknown_payment_method = PaymentMethod.objects.filter(
            code="UNKNOWN"
        ).first()

        if not unknown_payment_method:
            raise RuntimeError(
                "PaymentMethod with code='UNKNOWN' does not exist."
            )

        if "CUST-033" not in customers:
            raise RuntimeError(
                "CUST-033 Customer Unknown is missing."
            )

        self.stdout.write(
            f"Customers loaded : {len(customers)}"
        )
        self.stdout.write(
            f"Materials loaded : {len(materials)}"
        )
        self.stdout.write(
            f"Labour loaded    : {len(labour_by_name)}"
        )

        # ---------------------------------------------------------
        # LOAD WORKBOOKS
        # ---------------------------------------------------------

        trips_wb = openpyxl.load_workbook(
            TRIPS_FILE,
            data_only=True,
        )

        payments_wb = openpyxl.load_workbook(
            PAYMENTS_FILE,
            data_only=True,
        )

        # ---------------------------------------------------------
        # READ + VALIDATE TRIPS
        # ---------------------------------------------------------

        trip_rows = []
        trip_codes = set()
        trip_received = defaultdict(Decimal)

        total_trip_amount = Decimal("0")
        total_received = Decimal("0")
        errors = []

        for sheet_name, default_transaction_type in TRIP_SHEETS.items():

            ws = trips_wb[sheet_name]

            headers = [
                clean(ws.cell(1, col).value)
                for col in range(1, ws.max_column + 1)
            ]

            header_map = {
                name: index + 1
                for index, name in enumerate(headers)
            }

            for row in range(2, ws.max_row + 1):

                trip_code = clean(
                    ws.cell(
                        row,
                        header_map["Trip Code"]
                    ).value
                )

                if not trip_code:
                    continue

                if trip_code in trip_codes:
                    errors.append(
                        f"Duplicate Trip Code: {trip_code}"
                    )

                trip_codes.add(trip_code)

                trip_date = as_date(
                    ws.cell(
                        row,
                        header_map["Trip Date"]
                    ).value
                )

                customer_code = clean(
                    ws.cell(
                        row,
                        header_map["Customer Code"]
                    ).value
                )

                # Historical workbook uses 0 for unknown customer.
                if customer_code == "0":
                    customer_code = "CUST-033"

                customer_name = clean(
                    ws.cell(
                        row,
                        header_map["Customer Name"]
                    ).value
                )

                destination = clean(
                    ws.cell(
                        row,
                        header_map["Destination"]
                    ).value
                )

                quantity = decimal_value(
                    ws.cell(
                        row,
                        header_map["Quantity"]
                    ).value
                )

                rate = decimal_value(
                    ws.cell(
                        row,
                        header_map["Rate"]
                    ).value
                )

                excel_total = decimal_value(
                    ws.cell(
                        row,
                        header_map["Total Amount"]
                    ).value
                )

                received = decimal_value(
                    ws.cell(
                        row,
                        header_map["Received"]
                    ).value
                )

                balance = decimal_value(
                    ws.cell(
                        row,
                        header_map["Balance"]
                    ).value
                )

                vehicle_type = clean(
                    ws.cell(
                        row,
                        header_map["Vechile Type"]
                    ).value
                )

                material_type = ""

                if "Material Type" in header_map:
                    material_type = clean(
                        ws.cell(
                            row,
                            header_map["Material Type"]
                        ).value
                    )

                if not material_type:
                    if sheet_name == "Hyva Trip Fly Ash":
                        material_type = "Fly Ash"

                if default_transaction_type:
                    transaction_type = default_transaction_type
                else:
                    transaction_type = clean(
                        ws.cell(
                            row,
                            header_map["Transaction Type"]
                        ).value
                    )

                    if transaction_type.upper() == "INTERNAL STOCK":
                        transaction_type = "INTERNAL_STOCK"

                    elif transaction_type.upper() == "CUSTOMER DELIVERY":
                        transaction_type = "CUSTOMER_DELIVERY"

                notes = ""

                if "Notes" in header_map:
                    notes = clean(
                        ws.cell(
                            row,
                            header_map["Notes"]
                        ).value
                    )

                # -------------------------------------------------
                # REQUIRED FIELD VALIDATION
                # -------------------------------------------------

                if not trip_date:
                    errors.append(
                        f"{trip_code}: Trip Date is blank."
                    )

                if not quantity or quantity <= 0:
                    errors.append(
                        f"{trip_code}: Invalid quantity."
                    )

                # TRIP-T-009 is historical Internal Stock with
                # blank rate/amount. The Trip model requires a
                # numeric rate, so preserve it as zero.
                if rate is None:
                    rate = Decimal("0.00")

                if excel_total is None:
                    excel_total = Decimal("0.00")

                if received is None:
                    received = Decimal("0.00")

                if balance is None:
                    balance = excel_total - received

                if not material_type:
                    errors.append(
                        f"{trip_code}: Material Type blank."
                    )

                material = materials.get(
                    normalize_name(material_type)
                )

                if not material:
                    errors.append(
                        f"{trip_code}: Material "
                        f"'{material_type}' not found in DB."
                    )

                # Customer is optional for internal stock.
                customer = None

                if transaction_type != "INTERNAL_STOCK":

                    customer = customers.get(customer_code)

                    if not customer:
                        errors.append(
                            f"{trip_code}: Customer "
                            f"{customer_code} not found in DB."
                        )

                # -------------------------------------------------
                # DRIVER READ + VALIDATION
                # -------------------------------------------------

                driver_names = []

                for driver_column in (
                    "Driver 1",
                    "Driver 2",
                    "Driver 3",
                ):

                    driver_name = clean(
                        ws.cell(
                            row,
                            header_map[driver_column]
                        ).value
                    )

                    if not driver_name:
                        continue

                    driver_names.append(driver_name)

                    if normalize_name(driver_name) not in labour_by_name:
                        errors.append(
                            f"{trip_code}: Driver/Labour "
                            f"'{driver_name}' not found in DB."
                        )

                # -------------------------------------------------
                # TOTAL VALIDATION
                # -------------------------------------------------

                calculated_total = (
                    quantity * rate
                ).quantize(
                    Decimal("0.01")
                )

                # Only validate calculated total when source
                # actually contains a total.
                if excel_total != Decimal("0.00"):

                    if calculated_total != excel_total:
                        errors.append(
                            f"{trip_code}: Total mismatch. "
                            f"Excel={excel_total}, "
                            f"Calculated={calculated_total}"
                        )

                calculated_balance = (
                    excel_total - received
                ).quantize(
                    Decimal("0.01")
                )

                if calculated_balance != balance:
                    errors.append(
                        f"{trip_code}: Balance mismatch. "
                        f"Excel={balance}, "
                        f"Calculated={calculated_balance}"
                    )

                trip_received[trip_code] += received
                total_received += received
                total_trip_amount += excel_total

                trip_rows.append({
                    "trip_code": trip_code,
                    "trip_date": trip_date,
                    "customer": customer,
                    "customer_code": customer_code,
                    "destination": destination,
                    "vehicle_type": vehicle_type,
                    "material": material,
                    "transaction_type": transaction_type,
                    "quantity": quantity,
                    "rate": rate,
                    "total_amount": excel_total,
                    "received": received,
                    "notes": notes,
                    "driver_names": driver_names,
                })

        # ---------------------------------------------------------
        # READ + VALIDATE PAYMENTS
        # ---------------------------------------------------------

        payment_rows = []
        payment_codes = set()
        payment_trip_amount = defaultdict(Decimal)

        total_payment_amount = Decimal("0.00")

        for sheet_name in PAYMENT_SHEETS:

            ws = payments_wb[sheet_name]

            headers = [
                clean(ws.cell(1, col).value)
                for col in range(1, ws.max_column + 1)
            ]

            header_map = {
                name: index + 1
                for index, name in enumerate(headers)
            }

            for row in range(2, ws.max_row + 1):

                payment_code = clean(
                    ws.cell(
                        row,
                        header_map["Payment Code"]
                    ).value
                )

                if not payment_code:
                    continue

                if payment_code in payment_codes:
                    errors.append(
                        f"Duplicate Payment Code: {payment_code}"
                    )

                payment_codes.add(payment_code)

                payment_date = as_date(
                    ws.cell(
                        row,
                        header_map["Payment Date"]
                    ).value
                )

                trip_code = clean(
                    ws.cell(
                        row,
                        header_map["Trip Code"]
                    ).value
                )

                amount = decimal_value(
                    ws.cell(
                        row,
                        header_map["Amount"]
                    ).value
                )

                payment_method_name = clean(
                    ws.cell(
                        row,
                        header_map["Payment Method"]
                    ).value
                )

                reference_number = clean(
                    ws.cell(
                        row,
                        header_map["Reference Number"]
                    ).value
                )

                notes = clean(
                    ws.cell(
                        row,
                        header_map["Notes"]
                    ).value
                )

                if trip_code not in trip_codes:
                    errors.append(
                        f"{payment_code}: Trip Code "
                        f"{trip_code} not found."
                    )

                if amount is None or amount <= 0:
                    errors.append(
                        f"{payment_code}: Invalid amount."
                    )
                    continue

                payment_trip_amount[trip_code] += amount
                total_payment_amount += amount

                payment_rows.append({
                    "payment_code": payment_code,
                    "payment_date": payment_date,
                    "trip_code": trip_code,
                    "amount": amount,
                    "payment_method_name": payment_method_name,
                    "reference_number": reference_number,
                    "notes": notes,
                })

        # ---------------------------------------------------------
        # PAYMENT RECONCILIATION
        # ---------------------------------------------------------

        for trip_code in sorted(
            set(payment_trip_amount) | set(trip_received)
        ):

            received = trip_received.get(
                trip_code,
                Decimal("0.00")
            )

            payments = payment_trip_amount.get(
                trip_code,
                Decimal("0.00")
            )

            if received != payments:
                errors.append(
                    f"{trip_code}: Received/Payment mismatch. "
                    f"Received={received}, "
                    f"Payments={payments}"
                )

        # ---------------------------------------------------------
        # SUMMARY
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("IMPORT SUMMARY")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Trips to process    : {len(trip_rows)}"
        )

        self.stdout.write(
            f"Payments to process : {len(payment_rows)}"
        )

        self.stdout.write(
            f"Trip amount         : ₹{total_trip_amount:,.2f}"
        )

        self.stdout.write(
            f"Received            : ₹{total_received:,.2f}"
        )

        self.stdout.write(
            f"Payment amount      : ₹{total_payment_amount:,.2f}"
        )

        self.stdout.write(
            f"Errors              : {len(errors)}"
        )

        if errors:

            self.stdout.write("")
            self.stdout.write(
                self.style.ERROR("IMPORT VALIDATION FAILED")
            )

            for error in errors:
                self.stdout.write(
                    self.style.ERROR(f"- {error}")
                )

            raise RuntimeError(
                "Import stopped because validation errors were found."
            )

        if dry_run:

            self.stdout.write("")
            self.stdout.write(
                self.style.SUCCESS(
                    "DRY RUN PASSED - NO DATABASE CHANGES WERE MADE."
                )
            )

            return

        # ---------------------------------------------------------
        # LIVE IMPORT
        # ---------------------------------------------------------

        created_trips = 0
        updated_trips = 0
        created_payments = 0
        updated_payments = 0

        with transaction.atomic():

            trip_objects = {}

            for row in trip_rows:

                trip = Trip.objects.filter(
                    trip_code=row["trip_code"]
                ).first()

                if trip:

                    updated_trips += 1

                else:

                    trip = Trip(
                        trip_code=row["trip_code"]
                    )

                    created_trips += 1

                trip.trip_date = row["trip_date"]
                trip.customer = row["customer"]
                trip.destination = row["destination"]
                trip.vehicle = None
                trip.material = row["material"]
                trip.transaction_type = row["transaction_type"]
                trip.quantity = row["quantity"]
                trip.rate = row["rate"]
                trip.total_amount = row["total_amount"]
                trip.notes = row["notes"]

                received = row["received"]

                if received <= 0:
                    trip.payment_status = "UNPAID"
                elif received >= row["total_amount"]:
                    trip.payment_status = "PAID"
                else:
                    trip.payment_status = "PARTIAL"

                trip.trip_status = "COMPLETED"

                trip.save()

                # Preserve exact historical trip code.
                if trip.trip_code != row["trip_code"]:
                    Trip.objects.filter(
                        pk=trip.pk
                    ).update(
                        trip_code=row["trip_code"]
                    )

                    trip.trip_code = row["trip_code"]

                trip.drivers.set([
                    labour_by_name[normalize_name(name)]
                    for name in row["driver_names"]
                ])

                trip_objects[row["trip_code"]] = trip

            for row in payment_rows:

                trip = trip_objects[row["trip_code"]]

                payment_method = unknown_payment_method

                if row["payment_method_name"]:

                    payment_method = (
                        PaymentMethod.objects.filter(
                            name__iexact=row["payment_method_name"]
                        ).first()
                        or PaymentMethod.objects.filter(
                            code__iexact=row["payment_method_name"]
                        ).first()
                        or unknown_payment_method
                    )

                payment = TripPayment.objects.filter(
                    payment_code=row["payment_code"]
                ).first()

                if payment:

                    updated_payments += 1

                else:

                    payment = TripPayment(
                        payment_code=row["payment_code"]
                    )

                    created_payments += 1

                payment.trip = trip
                payment.payment_date = row["payment_date"]
                payment.amount = row["amount"]
                payment.payment_method = payment_method
                payment.reference_number = row["reference_number"]
                payment.notes = row["notes"]

                payment.save()

        # ---------------------------------------------------------
        # POST-IMPORT VERIFICATION
        # ---------------------------------------------------------

        imported_trip_count = Trip.objects.filter(
            trip_code__in=trip_codes
        ).count()

        imported_payment_count = TripPayment.objects.filter(
            payment_code__in=payment_codes
        ).count()

        db_trip_amount = sum(
            (
                Trip.objects.filter(
                    trip_code__in=trip_codes
                ).values_list(
                    "total_amount",
                    flat=True
                )
            ),
            Decimal("0.00")
        )

        db_payment_amount = sum(
            (
                TripPayment.objects.filter(
                    payment_code__in=payment_codes
                ).values_list(
                    "amount",
                    flat=True
                )
            ),
            Decimal("0.00")
        )

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("IMPORT COMPLETE")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Trips created       : {created_trips}"
        )

        self.stdout.write(
            f"Trips updated       : {updated_trips}"
        )

        self.stdout.write(
            f"Payments created    : {created_payments}"
        )

        self.stdout.write(
            f"Payments updated    : {updated_payments}"
        )

        self.stdout.write(
            f"Trips verified      : {imported_trip_count}"
        )

        self.stdout.write(
            f"Payments verified   : {imported_payment_count}"
        )

        self.stdout.write(
            f"DB trip amount      : ₹{db_trip_amount:,.2f}"
        )

        self.stdout.write(
            f"DB payment amount   : ₹{db_payment_amount:,.2f}"
        )

        if (
            imported_trip_count != len(trip_codes)
            or imported_payment_count != len(payment_codes)
            or db_trip_amount != total_trip_amount
            or db_payment_amount != total_payment_amount
        ):
            raise RuntimeError(
                "POST-IMPORT VERIFICATION FAILED."
            )

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                "POST-IMPORT VERIFICATION PASSED."
            )
        )
