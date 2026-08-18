import os
from collections import defaultdict
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand
from django.db.models import Sum

from customers.models import Customer
from labour.models import Labour
from master_data.models import Material
from trips.models import Trip


TRIPS_FILE = "historical_data/All Hyva Trips Clean Data.xlsx"
PAYMENTS_FILE = "historical_data/January_Payments_Clean_v2.xlsx"


TRIP_SHEETS = {
    "Hyva Trips  White Sand": "CUSTOMER_DELIVERY",
    "Halfton trips": "CUSTOMER_DELIVERY",
    "Hyva Trip Fly Ash": "CUSTOMER_DELIVERY",
    "Tractor": None,  # Read Transaction Type from Excel
}

PAYMENT_SHEETS = {
    "Hyva White Sand Payment",
    "Halfton White Sand",
    "Tractor",
    "Hyva Fly Ash Payment",
}


def clean(value):
    if value is None:
        return ""

    return str(value).strip()


def decimal_value(value):
    if value is None or value == "":
        return Decimal("0")

    if isinstance(value, Decimal):
        return value

    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def normalize_name(value):
    return " ".join(clean(value).lower().split())


class Command(BaseCommand):
    help = "Preview January historical trip/payment import without writing to DB."

    def handle(self, *args, **options):

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("JANUARY HISTORICAL IMPORT PREVIEW")
        self.stdout.write("=" * 70)

        # ---------------------------------------------------------
        # FILE CHECK
        # ---------------------------------------------------------

        if not os.path.exists(TRIPS_FILE):
            self.stdout.write(
                self.style.ERROR(
                    f"Trips file not found: {TRIPS_FILE}"
                )
            )
            return

        if not os.path.exists(PAYMENTS_FILE):
            self.stdout.write(
                self.style.ERROR(
                    f"Payments file not found: {PAYMENTS_FILE}"
                )
            )
            return

        trips_wb = openpyxl.load_workbook(
            TRIPS_FILE,
            data_only=True
        )

        payments_wb = openpyxl.load_workbook(
            PAYMENTS_FILE,
            data_only=True
        )

        # ---------------------------------------------------------
        # LOAD EXISTING MASTERS
        # ---------------------------------------------------------

        customers = {
            clean(obj.customer_code): obj
            for obj in Customer.objects.all()
        }

        labour_by_name = {
            normalize_name(obj.name): obj
            for obj in Labour.objects.all()
        }

        materials = {
            normalize_name(obj.name): obj
            for obj in Material.objects.all()
        }

        self.stdout.write("")
        self.stdout.write("DATABASE MASTER COUNTS")
        self.stdout.write("-" * 70)
        self.stdout.write(
            f"Customers : {len(customers)}"
        )
        self.stdout.write(
            f"Labour    : {len(labour_by_name)}"
        )
        self.stdout.write(
            f"Materials : {len(materials)}"
        )

        # ---------------------------------------------------------
        # TRIP PREVIEW
        # ---------------------------------------------------------

        trip_codes = set()
        trip_rows = []
        trip_received = defaultdict(Decimal)

        errors = []
        warnings = []

        total_trip_rows = 0
        total_trip_amount = Decimal("0")
        total_received = Decimal("0")

        self.stdout.write("")
        self.stdout.write("TRIP VALIDATION")
        self.stdout.write("-" * 70)

        for sheet_name, default_transaction_type in TRIP_SHEETS.items():

            ws = trips_wb[sheet_name]

            headers = [
                clean(ws.cell(1, col).value)
                for col in range(1, ws.max_column + 1)
            ]

            header_map = {
                name: index + 1
                for index, name in enumerate(headers)
            }

            self.stdout.write(
                f"\n[{sheet_name}]"
            )

            for row in range(2, ws.max_row + 1):

                trip_code = clean(
                    ws.cell(
                        row,
                        header_map["Trip Code"]
                    ).value
                )

                if not trip_code:
                    continue

                total_trip_rows += 1

                if trip_code in trip_codes:
                    errors.append(
                        f"Duplicate Trip Code: {trip_code}"
                    )

                trip_codes.add(trip_code)

                trip_date = ws.cell(
                    row,
                    header_map["Trip Date"]
                ).value

                customer_code = clean(
                    ws.cell(
                        row,
                        header_map["Customer Code"]
                    ).value
                )
                if customer_code == "0":
                    customer_code = "CUST-033"

                customer_name = clean(
                    ws.cell(
                        row,
                        header_map["Customer Name"]
                    ).value
                )

                quantity = decimal_value(
                    ws.cell(
                        row,
                        header_map["Quantity"]
                    ).value
                )

                rate = decimal_value(
                    ws.cell(
                        row,
                        header_map["Rate"]
                    ).value
                )

                excel_total = decimal_value(
                    ws.cell(
                        row,
                        header_map["Total Amount"]
                    ).value
                )

                received = decimal_value(
                    ws.cell(
                        row,
                        header_map["Received"]
                    ).value
                )

                balance = decimal_value(
                    ws.cell(
                        row,
                        header_map["Balance"]
                    ).value
                )

                vehicle_type = clean(
                    ws.cell(
                        row,
                        header_map["Vechile Type"]
                    ).value
                )

                material_type = ""

                if "Material Type" in header_map:
                    material_type = clean(
                        ws.cell(
                            row,
                            header_map["Material Type"]
                        ).value
                    )

                if not material_type:
                    if sheet_name == "Hyva Trip Fly Ash":
                        material_type = "Fly Ash"

                if default_transaction_type:
                    transaction_type = default_transaction_type
                else:
                    transaction_type = clean(
                        ws.cell(
                            row,
                            header_map["Transaction Type"]
                        ).value
                    )

                # -------------------------------------------------
                # CUSTOMER CHECK
                # -------------------------------------------------

                if transaction_type == "INTERNAL_STOCK":

                    if customer_code or customer_name:
                        warnings.append(
                            f"{trip_code}: Internal Stock has "
                            f"customer information."
                        )

                else:

                    if customer_code:

                        if customer_code == "NOT FOUND":
                            warnings.append(
                                f"{trip_code}: Customer code is "
                                f"NOT FOUND."
                            )

                        elif customer_code not in customers:
                            errors.append(
                                f"{trip_code}: Customer code "
                                f"{customer_code} not found in DB."
                            )

                    else:
                        warnings.append(
                            f"{trip_code}: Customer code blank."
                        )

                # -------------------------------------------------
                # MATERIAL CHECK
                # -------------------------------------------------

                if not material_type:

                    errors.append(
                        f"{trip_code}: Material Type blank."
                    )

                else:

                    material = materials.get(
                        normalize_name(material_type)
                    )

                    if not material:
                        errors.append(
                            f"{trip_code}: Material "
                            f"'{material_type}' not found in DB."
                        )

                # -------------------------------------------------
                # DRIVER CHECK
                # -------------------------------------------------

                driver_columns = [
                    "Driver 1",
                    "Driver 2",
                    "Driver 3",
                ]

                for driver_column in driver_columns:

                    driver_name = clean(
                        ws.cell(
                            row,
                            header_map[driver_column]
                        ).value
                    )

                    if not driver_name:
                        continue

                    if normalize_name(driver_name) not in labour_by_name:
                        errors.append(
                            f"{trip_code}: Driver/Labour "
                            f"'{driver_name}' not found in DB."
                        )

                # -------------------------------------------------
                # TOTAL CHECK
                # -------------------------------------------------

                if quantity is not None and rate is not None:

                    calculated_total = (
                        quantity * rate
                    ).quantize(
                        Decimal("0.01")
                    )

                    if excel_total is not None:

                        if calculated_total != excel_total:
                            errors.append(
                                f"{trip_code}: Total mismatch. "
                                f"Excel={excel_total}, "
                                f"Calculated={calculated_total}"
                            )

                # -------------------------------------------------
                # BALANCE CHECK
                # -------------------------------------------------

                if (
                    excel_total is not None
                    and received is not None
                    and balance is not None
                ):

                    calculated_balance = (
                        excel_total - received
                    ).quantize(
                        Decimal("0.01")
                    )

                    if calculated_balance != balance:
                        errors.append(
                            f"{trip_code}: Balance mismatch. "
                            f"Excel={balance}, "
                            f"Calculated={calculated_balance}"
                        )

                # -------------------------------------------------
                # RECEIVED TOTAL
                # -------------------------------------------------

                if received is not None:

                    trip_received[trip_code] += received
                    total_received += received

                if excel_total is not None:
                    total_trip_amount += excel_total

                trip_rows.append({
                    "trip_code": trip_code,
                    "trip_date": trip_date,
                    "customer_code": customer_code,
                    "customer_name": customer_name,
                    "vehicle_type": vehicle_type,
                    "material_type": material_type,
                    "transaction_type": transaction_type,
                    "quantity": quantity,
                    "rate": rate,
                    "total_amount": excel_total,
                    "received": received,

                    "drivers": [
                        clean(
                            ws.cell(
                                row,
                                header_map[driver_column]
                            ).value
                        )
                        for driver_column in [
                            "Driver 1",
                            "Driver 2",
                            "Driver 3",
                        ]
                        if clean(
                            ws.cell(
                                row,
                                header_map[driver_column]
                            ).value
                        )
                    ],
                })

            self.stdout.write(
                f"Rows checked: {ws.max_row - 1}"
            )

        # ---------------------------------------------------------
        # PAYMENT PREVIEW
        # ---------------------------------------------------------

        payment_codes = set()
        payment_trip_amount = defaultdict(Decimal)

        total_payment_rows = 0
        total_payment_amount = Decimal("0")

        self.stdout.write("")
        self.stdout.write("PAYMENT VALIDATION")
        self.stdout.write("-" * 70)

        for sheet_name in PAYMENT_SHEETS:

            ws = payments_wb[sheet_name]

            headers = [
                clean(ws.cell(1, col).value)
                for col in range(1, ws.max_column + 1)
            ]

            header_map = {
                name: index + 1
                for index, name in enumerate(headers)
            }

            self.stdout.write(
                f"\n[{sheet_name}]"
            )

            for row in range(2, ws.max_row + 1):

                payment_code = clean(
                    ws.cell(
                        row,
                        header_map["Payment Code"]
                    ).value
                )

                if not payment_code:
                    continue

                total_payment_rows += 1

                if payment_code in payment_codes:
                    errors.append(
                        f"Duplicate Payment Code: "
                        f"{payment_code}"
                    )

                payment_codes.add(payment_code)

                trip_code = clean(
                    ws.cell(
                        row,
                        header_map["Trip Code"]
                    ).value
                )

                amount = decimal_value(
                    ws.cell(
                        row,
                        header_map["Amount"]
                    ).value
                )

                payment_method = clean(
                    ws.cell(
                        row,
                        header_map["Payment Method"]
                    ).value
                )

                payment_date = ws.cell(
                    row,
                    header_map["Payment Date"]
                ).value

                # -------------------------------------------------
                # TRIP LINK CHECK
                # -------------------------------------------------

                if trip_code not in trip_codes:
                    errors.append(
                        f"{payment_code}: Trip Code "
                        f"{trip_code} not found in trip workbook."
                    )

                # -------------------------------------------------
                # PAYMENT METHOD
                # -------------------------------------------------

                if not payment_method:
                    warnings.append(
                        f"{payment_code}: Payment Method "
                        f"is blank (historical unknown)."
                    )

                # -------------------------------------------------
                # PAYMENT DATE
                # -------------------------------------------------

                if not payment_date:
                    warnings.append(
                        f"{payment_code}: Payment Date "
                        f"is blank (historical unknown)."
                    )

                # -------------------------------------------------
                # AMOUNT
                # -------------------------------------------------

                if amount is None or amount <= 0:
                    errors.append(
                        f"{payment_code}: Invalid payment amount."
                    )
                    continue

                payment_trip_amount[trip_code] += amount
                total_payment_amount += amount

            self.stdout.write(
                f"Rows checked: {ws.max_row - 1}"
            )

        # ---------------------------------------------------------
        # PAYMENT VS TRIP RECEIVED
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("PAYMENT RECONCILIATION")
        self.stdout.write("-" * 70)

        all_payment_trip_codes = set(
            payment_trip_amount.keys()
        )

        all_received_trip_codes = set(
            trip_received.keys()
        )

        for trip_code in sorted(
            all_payment_trip_codes | all_received_trip_codes
        ):

            received = trip_received.get(
                trip_code,
                Decimal("0")
            )

            payments = payment_trip_amount.get(
                trip_code,
                Decimal("0")
            )

            if received != payments:

                errors.append(
                    f"{trip_code}: Received/Payment mismatch. "
                    f"Trip Received={received}, "
                    f"Payments={payments}"
                )

        # ---------------------------------------------------------
        # VEHICLE INFORMATION
        # ---------------------------------------------------------

        vehicle_warning_count = 0

        for trip in trip_rows:

            if not trip["vehicle_type"]:

                errors.append(
                    f"{trip['trip_code']}: Vehicle Type blank."
                )

            # Historical January workbook has vehicle type,
            # but not the exact registration number/date mapping.
            vehicle_warning_count += 1

        warnings.append(
            f"{vehicle_warning_count} trips have Vehicle Type "
            f"but no reliable historical registration-number mapping. "
            f"Vehicle FK will therefore remain blank during historical import."
        )

        # ---------------------------------------------------------
        # FINAL SUMMARY
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("PREVIEW SUMMARY")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Trip rows       : {total_trip_rows}"
        )

        self.stdout.write(
            f"Trip amount     : ₹{total_trip_amount:,.2f}"
        )

        self.stdout.write(
            f"Trip received   : ₹{total_received:,.2f}"
        )

        self.stdout.write(
            f"Payment rows    : {total_payment_rows}"
        )

        self.stdout.write(
            f"Payment amount  : ₹{total_payment_amount:,.2f}"
        )

        self.stdout.write(
            f"Unique trips    : {len(trip_codes)}"
        )

        self.stdout.write(
            f"Unique payments : {len(payment_codes)}"
        )

        self.stdout.write("")
        self.stdout.write(
            f"ERRORS   : {len(errors)}"
        )

        self.stdout.write(
            f"WARNINGS : {len(warnings)}"
        )

        # ---------------------------------------------------------
        # ERRORS
        # ---------------------------------------------------------

        if errors:

            self.stdout.write("")
            self.stdout.write(
                self.style.ERROR(
                    "ERROR DETAILS"
                )
            )

            for error in errors:
                self.stdout.write(
                    self.style.ERROR(
                        f"  ❌ {error}"
                    )
                )

        # ---------------------------------------------------------
        # WARNINGS
        # ---------------------------------------------------------

        if warnings:

            self.stdout.write("")
            self.stdout.write(
                self.style.WARNING(
                    "WARNING DETAILS"
                )
            )

            for warning in warnings:
                self.stdout.write(
                    self.style.WARNING(
                        f"  ⚠ {warning}"
                    )
                )

        # ---------------------------------------------------------
        # FINAL DECISION
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("=" * 70)

        if errors:

            self.stdout.write(
                self.style.ERROR(
                    "PREVIEW FAILED — DO NOT IMPORT."
                )
            )

        else:

            self.stdout.write(
                self.style.SUCCESS(
                    "PREVIEW PASSED — NO DATABASE CHANGES WERE MADE."
                )
            )

        self.stdout.write("=" * 70)