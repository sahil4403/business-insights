import csv
import os
from decimal import Decimal
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def _register_georgia_font():
    font_path = os.path.join(settings.BASE_DIR, 'ledger', 'fonts', 'NotoSans-Regular.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('GeorgiaUnicode', font_path))
    elif os.path.exists('/System/Library/Fonts/Supplemental/Georgia.ttf'):
        pdfmetrics.registerFont(TTFont('GeorgiaUnicode', '/System/Library/Fonts/Supplemental/Georgia.ttf'))

from django.contrib.auth.decorators import login_required
from datetime import date
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages
from django.urls import reverse
from django.utils.http import urlencode


from django.contrib.auth import authenticate, login as auth_login
from django.contrib.admin.forms import AdminAuthenticationForm
from django.http import HttpResponse
from core.utils import get_safe_next
from core.rate_limit import (
    login_rate_limit_check,
    record_login_failure,
    clear_login_rate_limit,
)

def custom_admin_login(request):
    """
    Renders Django Admin Login form and verifies credentials strictly.
    """
    next_url = get_safe_next(request, '/management-portal-x99/')
    error_message = None

    if request.method == 'POST':
        allowed, retry_after = login_rate_limit_check(request, 'admin_login')
        if not allowed:
            return HttpResponse(
                f'Too many login attempts. Please try again in {max(1, retry_after // 60)} minute(s).',
                status=429,
            )

        form = AdminAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user and user.is_staff:
                clear_login_rate_limit(request, 'admin_login')
                auth_login(request, user)
                request.session['admin_verified'] = True
                return redirect(next_url)
            else:
                record_login_failure(request, 'admin_login')
                error_message = "Please enter a valid staff/admin account."
        else:
            record_login_failure(request, 'admin_login')
            error_message = "Invalid username or password. Please try again."
    else:
        form = AdminAuthenticationForm(request)

    context = {
        'form': form,
        'next': next_url,
        'error_message': error_message,
        'site_header': 'Django Administration Security Login',
    }
    return render(request, 'admin/login.html', context)


@login_required(login_url='/login/')
def admin_reauth(request):
    """
    Directly opens Django Administration for logged in staff/admin users.
    """
    if request.user.is_authenticated and request.user.is_staff:
        request.session['admin_verified'] = True
        return redirect('/management-portal-x99/')
    return redirect('/management-portal-x99/login/?next=/management-portal-x99/')



from expenses.models import Expense
from customers.models import Customer
from trips.models import (
    Trip,
    TripPayment,
)

from master_data.models import PaymentMethod, VehicleType
from vehicles.models import Vehicle
from labour.models import Labour
from django.db.models.functions import Coalesce
from django.db.models import (
    Case,
    CharField,
    Count,
    DecimalField,
    F,
    Q,
    Sum,
    Value,
    When,
)


from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

from trips.forms import TripForm

@login_required(login_url='/login/')
def dashboard(request):
    today = timezone.localdate()

    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    # -----------------------------------
    # BASE QUERYSETS
    # -----------------------------------

    trips = Trip.objects.all()
    expenses = Expense.objects.all()

    # -----------------------------------
    # DATE FILTER
    # -----------------------------------

    if from_date:
        trips = trips.filter(
            trip_date__gte=from_date
        )

        expenses = expenses.filter(
            expense_date__gte=from_date
        )

    if to_date:
        trips = trips.filter(
            trip_date__lte=to_date
        )

        expenses = expenses.filter(
            expense_date__lte=to_date
        )

    # -----------------------------------
    # REVENUE
    # -----------------------------------

    total_revenue = (
        trips.aggregate(
            total=Sum('total_amount')
        )['total'] or 0
    )

    # -----------------------------------
    # EXPENSE
    # -----------------------------------

    total_expense = (
        expenses.aggregate(
            total=Sum('amount')
        )['total'] or 0
    )

    # -----------------------------------
    # PROFIT
    # -----------------------------------

    net_profit = (
        total_revenue - total_expense
    )

    # -----------------------------------
    # TRIP COUNT
    # -----------------------------------

    total_trips = trips.count()

    # -----------------------------------
    # PRODUCTION
    # -----------------------------------

    total_quantity = (
        trips.aggregate(
            total=Sum('quantity')
        )['total'] or 0
    )

    # -----------------------------------
    # PAYMENT RECEIVED
    # -----------------------------------

    total_received = (
        trips.aggregate(
            total=Sum('payments__amount')
        )['total'] or 0
    )

    # -----------------------------------
    # OUTSTANDING (includes opening balances)
    # -----------------------------------

    total_opening_balance = (
        Customer.objects.aggregate(
            total=Sum('opening_balance')
        )['total'] or 0
    )

    total_outstanding = max(
        total_opening_balance + total_revenue - total_received,
        0
    )

    # -----------------------------------
    # PROFIT MARGIN
    # -----------------------------------

    if total_revenue:
        profit_margin = round(
            (
                float(net_profit)
                / float(total_revenue)
            ) * 100,
            1
        )
    else:
        profit_margin = 0

    # -----------------------------------
    # PAYMENT COLLECTION ANALYTICS
    # -----------------------------------

    if total_revenue:
        collection_percentage = round(
            (
                float(total_received)
                / float(total_revenue)
            ) * 100,
            1
        )
    else:
        collection_percentage = 0

    outstanding_percentage = round(
        100 - collection_percentage,
        1
    )

    # -----------------------------------
    # TRIP STATUS ANALYTICS
    # -----------------------------------

    trip_status_counts = []

    for status_code, status_label in Trip.TRIP_STATUS_CHOICES:

        status_trips = trips.filter(
            trip_status=status_code
        )

        trip_status_counts.append({
            'code': status_code,
            'label': status_label,
            'count': status_trips.count(),
        })

    # -----------------------------------
    # PAYMENT STATUS ANALYTICS
    # -----------------------------------

    payment_status_counts = []

    for status_code, status_label in Trip.PAYMENT_STATUS_CHOICES:
        status_trips = trips.filter(
            payment_status=status_code
        )

        payment_status_counts.append({
            'code': status_code,
            'label': status_label,
            'count': status_trips.count(),
        })

    # -----------------------------------
    # CUSTOMER-WISE REVENUE ANALYTICS
    # -----------------------------------

    customer_revenue = (
        trips
        .values(
            'customer__name'
        )
        .annotate(
            revenue=Sum('total_amount')
        )
        .order_by(
            '-revenue'
        )[:10]
    )

    # -----------------------------------
    # VEHICLE-WISE ANALYTICS
    # -----------------------------------

    vehicle_analytics = (
        trips
        .values(
            'vehicle__registration_number'
        )
        .annotate(
            trips_count=Count('id', distinct=True),
            revenue=Sum('total_amount')
        )
        .order_by(
            '-revenue'
        )
    )


    # -----------------------------------
    # MONTHLY BUSINESS PERFORMANCE
    # -----------------------------------

    monthly_business = []

    for i in range(5, -1, -1):

        month = today.month - i
        year = today.year

        while month <= 0:
            month += 12
            year -= 1

        month_trips = Trip.objects.filter(
            trip_date__year=year,
            trip_date__month=month,
        )

        month_expenses = Expense.objects.filter(
            expense_date__year=year,
            expense_date__month=month,
        )

        if from_date:
            month_trips = month_trips.filter(
                trip_date__gte=from_date
            )

            month_expenses = month_expenses.filter(
                expense_date__gte=from_date
            )

        if to_date:
            month_trips = month_trips.filter(
                trip_date__lte=to_date
            )

            month_expenses = month_expenses.filter(
                expense_date__lte=to_date
            )

        month_revenue = (
            month_trips.aggregate(
                total=Sum('total_amount')
            )['total'] or 0
        )

        month_expense = (
            month_expenses.aggregate(
                total=Sum('amount')
            )['total'] or 0
        )

        month_profit = (
            month_revenue - month_expense
        )

        if month_revenue:

            month_margin = round(
                (
                    float(month_profit)
                    / float(month_revenue)
                ) * 100,
                1
            )

        else:

            month_margin = 0

        monthly_business.append({
            'month': date(year, month, 1).strftime('%b %Y'),
            'revenue': month_revenue,
            'expense': month_expense,
            'profit': month_profit,
            'margin': month_margin,
        })

    # -----------------------------------
    # CONTEXT
    # -----------------------------------

    context = {

        'today': today,

        'from_date': from_date,
        'to_date': to_date,

        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_profit': net_profit,

        'total_trips': total_trips,
        'total_quantity': total_quantity,

        'total_received': total_received,
        'total_outstanding': total_outstanding,
        'collection_percentage': collection_percentage,
        'outstanding_percentage': outstanding_percentage,

        'trip_status_counts': trip_status_counts,
        'payment_status_counts': payment_status_counts,
        'customer_revenue': customer_revenue,
        'vehicle_analytics': vehicle_analytics,

        'profit_margin': profit_margin,
        'monthly_business': monthly_business,
        'trips': trips.select_related('customer', 'vehicle', 'material').prefetch_related('drivers').order_by('-trip_date', '-id'),
        'is_admin_user': request.user.is_authenticated and request.user.is_superuser,
    }

    return render(
        request,
        'core/dashboard.html',
        context
    )

# -----------------------------------
# CUSTOMER REPORT
# -----------------------------------
@login_required(login_url='/login/')
def customer_report(request):

    from_date = request.GET.get(
        'from_date',
        ''
    ).strip()

    to_date = request.GET.get(
        'to_date',
        ''
    ).strip()

    search = request.GET.get(
        'search',
        ''
    ).strip()

    min_amount_str = request.GET.get('min_amount', '').strip()
    max_amount_str = request.GET.get('max_amount', '').strip()

    min_amount = None
    if min_amount_str:
        try:
            min_amount = Decimal(min_amount_str)
        except Exception:
            min_amount = None

    max_amount = None
    if max_amount_str:
        try:
            max_amount = Decimal(max_amount_str)
        except Exception:
            max_amount = None


    # -----------------------------------
    # BASE QUERYSET (All Customers)
    # -----------------------------------

    customers = Customer.objects.filter(is_active=True)

    if search:
        customers = customers.filter(
            Q(name__icontains=search) | Q(customer_code__icontains=search)
        )

    def _parse_date(value):
        try:
            return date.fromisoformat(value)
        except Exception:
            return None

    from_date_parsed = _parse_date(from_date)
    to_date_parsed = _parse_date(to_date)

    window_end = to_date_parsed or timezone.localdate()

    # -----------------------------------
    # CUSTOMER REPORT DATA
    # Computed with the SAME sign rules as the ledger statement so that
    # the Outstanding column reconciles 1:1 with the customer statement:
    #   • VENDOR_SUPPLY trips are inward supply -> credit (reduces receivable)
    #   • PAID-type payments are money paid out -> debit (increases receivable)
    #   • RECEIVED / CONTRA payments reduce receivable
    #   • payment effective date = payment_date, falling back to trip_date
    # -----------------------------------

    customers_list = list(
        customers.order_by('name').values('id', 'name', 'opening_balance')
    )

    customer_ids = [c['id'] for c in customers_list]

    trips_by_customer = {cid: [] for cid in customer_ids}
    payments_by_customer = {cid: [] for cid in customer_ids}

    all_trips = (
        Trip.objects.filter(customer_id__in=customer_ids)
        .select_related('vehicle', 'vehicle__vehicle_type')
    )

    all_payments = (
        TripPayment.objects.filter(
            Q(customer_id__in=customer_ids)
            | Q(trip__customer_id__in=customer_ids)
        ).select_related('trip')
    )

    for trip in all_trips:
        trips_by_customer[trip.customer_id].append(trip)

    for payment in all_payments:
        owner_id = (
            payment.customer_id
            or (payment.trip.customer_id if payment.trip else None)
        )
        if owner_id in payments_by_customer:
            payments_by_customer[owner_id].append(payment)

    ZERO = Decimal('0')

    customer_rows = []

    for info in customers_list:

        trips = trips_by_customer[info['id']]
        payments = payments_by_customer[info['id']]

        period_sales = ZERO
        previous_sales = ZERO

        total_trips_count = 0
        jcb_trips_count = 0
        jcb_hours_total = ZERO

        for trip in trips:
            signed_amount = (
                -trip.total_amount
                if trip.transaction_type == 'VENDOR_SUPPLY'
                else trip.total_amount
            )

            if from_date_parsed and trip.trip_date < from_date_parsed:
                previous_sales += signed_amount
                continue

            if trip.trip_date > window_end:
                continue

            # Trips column counts ALL entries for the customer,
            # including JCB work entries.
            total_trips_count += 1

            vehicle_type_code = (
                trip.vehicle.vehicle_type.code
                if trip.vehicle and trip.vehicle.vehicle_type
                else None
            )
            if vehicle_type_code == 'JCB':
                jcb_trips_count += 1
                jcb_hours_total += trip.quantity or ZERO

            period_sales += signed_amount

        period_received = ZERO
        previous_received = ZERO

        for payment in payments:
            effective_date = (
                payment.payment_date
                or (payment.trip.trip_date if payment.trip else None)
            )

            if effective_date is None:
                continue

            signed_amount = (
                -payment.amount
                if payment.payment_type == 'PAID'
                else payment.amount
            )

            if from_date_parsed and effective_date < from_date_parsed:
                previous_received += signed_amount
                continue

            if effective_date > window_end:
                continue

            period_received += signed_amount

        opening_bal = info['opening_balance'] or ZERO
        effective_opening = opening_bal + previous_sales - previous_received

        outstanding = (
            effective_opening + period_sales - period_received
        )

        customer_rows.append({
            'customer_id':
                info['id'],

            'customer_name':
                info['name'],

            'opening_balance':
                effective_opening,

            'total_trips':
                total_trips_count,

            'jcb_trips':
                jcb_trips_count,

            'jcb_hours':
                jcb_hours_total,

            'total_revenue':
                period_sales,

            'total_received':
                period_received,

            'total_outstanding':
                outstanding,
        })

    # -----------------------------------
    # BLANK EXPORT GUARD (no data -> block CSV/Excel/PDF)
    # -----------------------------------

    if request.GET.get('export'):

        def _passes_amount_filters(row):
            outstanding = row['total_outstanding']
            if min_amount is not None and outstanding < min_amount:
                return False
            if max_amount is not None and outstanding > max_amount:
                return False
            return True

        exportable_count = sum(
            1 for row in customer_rows if _passes_amount_filters(row)
        )

        if exportable_count == 0:

            messages.error(
                request,
                '❌ Export blocked: no data found for the selected filters.'
            )

            clean_params = {
                k: v for k, v in request.GET.items() if k != 'export'
            }

            return redirect(
                f"{reverse('core:customer_report')}?"
                f"{urlencode(clean_params)}"
            )

    # -----------------------------------
    # PDF EXPORT
    # -----------------------------------

    if request.GET.get('export') in ['pdf', 'pdf_preview']:

        response = HttpResponse(
            content_type='application/pdf'
        )

        from core.pdf_utils import (
            get_registered_font,
            build_pdf_header_elements,
            get_indian_current_time_str,
            apply_data_table_style,
            finish_document,
        )
        font_name = get_registered_font()
        _register_georgia_font()

        disposition_type = 'inline' if request.GET.get('preview') == 'true' or request.GET.get('export') == 'pdf_preview' else 'attachment'
        response[
            'Content-Disposition'
        ] = f'{disposition_type}; filename="customer_report.pdf"'

        document = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=25,
            leftMargin=25,
            topMargin=25,
            bottomMargin=25,
        )

        styles = getSampleStyleSheet()

        meta_info = f"Generated on: {get_indian_current_time_str()}"
        if from_date or to_date:
            meta_info += f" | Filter Period: {from_date or 'All'} to {to_date or 'Present'}"

        elements = build_pdf_header_elements(
            font_name=font_name,
            report_title="Customer Wise Financial Report",
            report_subtitle="Customer-wise revenue, payments and outstanding balance summary",
            extra_meta=meta_info
        )

        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=9,
            leading=11,
            textColor=colors.white,
            alignment=0,
        )

        header_right_style = ParagraphStyle(
            'HeaderRightStyle',
            parent=header_style,
            alignment=2,
        )

        header_center_style = ParagraphStyle(
            'HeaderCenterStyle',
            parent=header_style,
            alignment=1,
        )

        body_style = ParagraphStyle(
            'BodyStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor('#0F172A'),
        )

        right_body_style = ParagraphStyle(
            'RightBodyStyle',
            parent=body_style,
            alignment=2,
        )

        center_body_style = ParagraphStyle(
            'CenterBodyStyle',
            parent=body_style,
            alignment=1,
        )

        data = [
            [
                Paragraph('<b>Customer</b>', header_style),
                Paragraph('<b>Opening Bal.</b>', header_right_style),
                Paragraph('<b>Trips</b>', header_center_style),
                Paragraph('<b>JCB Work</b>', header_center_style),
                Paragraph('<b>Revenue</b>', header_right_style),
                Paragraph('<b>Received</b>', header_right_style),
                Paragraph('<b>Outstanding</b>', header_right_style),
            ]
        ]

        total_open = Decimal('0')
        total_trips_count = 0
        total_jcb_trips_count = 0
        total_rev = Decimal('0')
        total_rec = Decimal('0')
        total_out = Decimal('0')

        for customer in customer_rows:
            opening_bal = customer['opening_balance']
            revenue = customer['total_revenue']
            received = customer['total_received']
            outstanding = customer['total_outstanding']

            if min_amount is not None and outstanding < min_amount:
                continue
            if max_amount is not None and outstanding > max_amount:
                continue

            trips_cnt = customer['total_trips'] or 0
            jcb_cnt = customer.get('jcb_trips') or 0
            jcb_hrs = customer.get('jcb_hours') or Decimal('0')
            total_open += opening_bal
            total_trips_count += trips_cnt
            total_jcb_trips_count += jcb_cnt
            total_rev += revenue
            total_rec += received
            total_out += outstanding

            jcb_label = f'{float(jcb_hrs):.1f}h' if jcb_cnt else '—'

            data.append([
                Paragraph(str(customer['customer_name'] or '—'), body_style),
                Paragraph(f'₹{opening_bal:,.2f}', right_body_style),
                Paragraph(str(trips_cnt), center_body_style),
                Paragraph(jcb_label, center_body_style),
                Paragraph(f'₹{revenue:,.2f}', right_body_style),
                Paragraph(f'₹{received:,.2f}', right_body_style),
                Paragraph(f'₹{outstanding:,.2f}', right_body_style),
            ])

        # Add Summary Row
        data.append([
            Paragraph('<b>TOTAL</b>', body_style),
            Paragraph(f'<b>₹{total_open:,.2f}</b>', right_body_style),
            Paragraph(f'<b>{total_trips_count}</b>', center_body_style),
            Paragraph(f'<b>{total_jcb_trips_count}</b>', center_body_style),
            Paragraph(f'<b>₹{total_rev:,.2f}</b>', right_body_style),
            Paragraph(f'<b>₹{total_rec:,.2f}</b>', right_body_style),
            Paragraph(f'<b>₹{total_out:,.2f}</b>', right_body_style),
        ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[160, 95, 65, 90, 105, 105, 105],
        )

        apply_data_table_style(table, total_row=True)

        elements.append(table)
        finish_document(document, elements, font_name=font_name)

        return response

    # -----------------------------------
    # EXCEL EXPORT
    # -----------------------------------

    if request.GET.get('export') == 'excel':

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = 'Customer Report'


        headers = [
            'Customer',
            'Opening Balance',
            'Trips',
            'JCB Work (trips · hrs)',
            'Revenue',
            'Received',
            'Outstanding',
        ]

        worksheet.append(headers)


        for cell in worksheet[1]:

            cell.font = Font(
                bold=True
            )


        for customer in customer_rows:

            opening_bal = (
                customer['opening_balance']
                or 0
            )

            revenue = (
                customer['total_revenue']
                or 0
            )

            received = customer['total_received']

            outstanding = customer['total_outstanding']

            if min_amount is not None and outstanding < min_amount:
                continue
            if max_amount is not None and outstanding > max_amount:
                continue

            _jcb_cnt = customer.get('jcb_trips') or 0
            _jcb_hrs = float(customer.get('jcb_hours') or 0)
            jcb_label = f'{_jcb_hrs:.1f}h' if _jcb_cnt else ''

            worksheet.append([
                customer['customer_name'],
                opening_bal,
                customer['total_trips'],
                jcb_label,
                revenue,
                received,
                outstanding,
            ])


        # -----------------------------------
        # CURRENCY FORMAT
        # -----------------------------------

        for row in worksheet.iter_rows(
            min_row=2,
            min_col=2,
            max_col=6
        ):

            for cell in row:

                if cell.column not in (3, 4):  # Skip 'Trips' and 'JCB Work' columns
                    cell.number_format = (
                        '₹#,##0.00'
                    )


        # -----------------------------------
        # FREEZE HEADER
        # -----------------------------------

        worksheet.freeze_panes = 'A2'


        # -----------------------------------
        # AUTO FILTER
        # -----------------------------------

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


        # -----------------------------------
        # COLUMN WIDTH
        # -----------------------------------

        for column in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                35
            )


        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            )
        )

        response[
            'Content-Disposition'
        ] = (
            'attachment; '
            'filename="customer_report.xlsx"'
        )


        workbook.save(
            response
        )

        return response

    # -----------------------------------
    # CSV EXPORT
    # -----------------------------------

    if request.GET.get('export') == 'csv':

        response = HttpResponse(
            content_type='text/csv'
        )

        response[
            'Content-Disposition'
        ] = 'attachment; filename="customer_report.csv"'

        writer = csv.writer(response)

        writer.writerow([
            'Customer',
            'Opening Balance',
            'Total Trips',
            'Revenue',
            'Received',
            'Outstanding',
        ])

        for customer in customer_rows:

            outstanding = customer['total_outstanding']

            if min_amount is not None and outstanding < min_amount:
                continue
            if max_amount is not None and outstanding > max_amount:
                continue

            writer.writerow([
                customer['customer_name'],
                customer['opening_balance'],
                customer['total_trips'],
                customer['total_revenue'],
                customer['total_received'],
                outstanding,
            ])

        return response


    # -----------------------------------
    # OUTSTANDING FILTERING + SORTING
    # (customer_rows are already computed above with statement-consistent
    #  math; here we only apply the display filters/sorting)
    # -----------------------------------

    is_limited = False
    if not (search or min_amount_str or max_amount_str or from_date or to_date):
        # Show ALL customers having an outstanding balance (no limit)
        customer_rows = [
            row for row in customer_rows
            if row['total_outstanding'] > 0
        ]
        customer_rows.sort(
            key=lambda x: (
                x['total_outstanding'],
                x['opening_balance'],
            ),
            reverse=True
        )

    # Name sort toggle (A→Z / Z→A) — shows only customers with outstanding balance
    sort_order = request.GET.get('sort', '').strip().lower()
    if sort_order not in ('asc', 'desc'):
        sort_order = ''
    else:
        customer_rows = [
            row for row in customer_rows
            if row['total_outstanding'] > 0
        ]
        customer_rows.sort(
            key=lambda row: (row['customer_name'] or '').lower(),
            reverse=(sort_order == 'desc')
        )

    # Sort toggle URLs (preserve other active filters)
    base_params = request.GET.copy()
    sort_asc_params = base_params.copy()
    sort_asc_params['sort'] = 'asc'
    sort_desc_params = base_params.copy()
    sort_desc_params['sort'] = 'desc'
    clear_sort_params = base_params.copy()
    clear_sort_params.pop('sort', None)

    context = {
        'customer_rows': customer_rows,
        'from_date': from_date,
        'to_date': to_date,
        'search': search,
        'selected_sort': sort_order,
        'sort_asc_url': '?' + sort_asc_params.urlencode(),
        'sort_desc_url': '?' + sort_desc_params.urlencode(),
        'clear_sort_url': '?' + clear_sort_params.urlencode(),
        'min_amount': min_amount_str,
        'max_amount': max_amount_str,
        'is_limited': is_limited,
        'showing_count': len(customer_rows),
        'is_admin_user': request.user.is_authenticated and request.user.is_superuser,
    }


    return render(
        request,
        'core/customer_report.html',
        context
    )

# -----------------------------------
# VEHICLE REPORT
# -----------------------------------
@login_required(login_url='/login/')
def vehicle_report(request):

    from_date = request.GET.get(
        'from_date',
        ''
    ).strip()

    to_date = request.GET.get(
        'to_date',
        ''
    ).strip()

    search = request.GET.get(
        'search',
        ''
    ).strip()

    vehicle_type = request.GET.get(
        'vehicle_type',
        ''
    ).strip()

    vehicle_id = request.GET.get(
        'vehicle_id',
        ''
    ).strip()

    driver_id = request.GET.get(
        'driver_id',
        ''
    ).strip()


    # -----------------------------------
    # BASE QUERYSET
    # -----------------------------------

    trips = Trip.objects.all()


    # -----------------------------------
    # DATE FILTER
    # -----------------------------------

    if from_date:

        trips = trips.filter(
            trip_date__gte=from_date
        )


    if to_date:

        trips = trips.filter(
            trip_date__lte=to_date
        )


    # -----------------------------------
    # VEHICLE TYPE FILTER
    # -----------------------------------

    if vehicle_type:

        trips = trips.filter(
            vehicle__vehicle_type_id=vehicle_type
        )


    # -----------------------------------
    # VEHICLE FILTER
    # -----------------------------------

    if vehicle_id:

        trips = trips.filter(
            vehicle_id=vehicle_id
        )


    # -----------------------------------
    # DRIVER FILTER
    # -----------------------------------

    if driver_id:

        trips = trips.filter(
            drivers__id=driver_id
        )


    # -----------------------------------
    # SEARCH FILTER
    # -----------------------------------

    if search:

        trips = trips.filter(
            vehicle__registration_number__icontains=search
        )


    # -----------------------------------
    # VEHICLE REPORT
    # -----------------------------------

    vehicle_report_data = (
        trips
        .values(
            'vehicle_id',
            'vehicle__registration_number'
        )
        .annotate(
            total_trips=Count('id', distinct=True),
            total_revenue=Sum('total_amount'),
            total_received=Sum(
                'payments__amount'
            ),
        )
        .order_by(
            '-total_revenue'
        )
    )

    # -----------------------------------
    # BLANK EXPORT GUARD (no data -> block CSV/Excel/PDF)
    # -----------------------------------

    if request.GET.get('export'):

        if not vehicle_report_data.exists():

            messages.error(
                request,
                '❌ Export blocked: no vehicle data found for the selected filters.'
            )

            clean_params = {
                k: v for k, v in request.GET.items() if k != 'export'
            }

            return redirect(
                f"{reverse('core:vehicle_report')}?"
                f"{urlencode(clean_params)}"
            )

    # -----------------------------------
    # PDF EXPORT
    # -----------------------------------

    if request.GET.get('export') in ['pdf', 'pdf_preview']:

        response = HttpResponse(
            content_type='application/pdf'
        )

        from core.pdf_utils import (
            get_registered_font,
            build_pdf_header_elements,
            get_indian_current_time_str,
            apply_data_table_style,
            finish_document,
        )
        font_name = get_registered_font()
        _register_georgia_font()

        disposition_type = 'inline' if request.GET.get('preview') == 'true' or request.GET.get('export') == 'pdf_preview' else 'attachment'
        response[
            'Content-Disposition'
        ] = f'{disposition_type}; filename="vehicle_report.pdf"'

        document = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=25,
            leftMargin=25,
            topMargin=25,
            bottomMargin=25,
        )

        styles = getSampleStyleSheet()

        meta_info = f"Generated on: {get_indian_current_time_str()}"
        if from_date or to_date:
            meta_info += f" | Filter Period: {from_date or 'All'} to {to_date or 'Present'}"

        elements = build_pdf_header_elements(
            font_name=font_name,
            report_title="Vehicle Wise Financial Summary Report",
            report_subtitle="Vehicle-wise revenue, received payments and outstanding balance",
            extra_meta=meta_info
        )

        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=9,
            leading=11,
            textColor=colors.white,
            alignment=0,
        )

        header_right_style = ParagraphStyle(
            'HeaderRightStyle',
            parent=header_style,
            alignment=2,
        )

        header_center_style = ParagraphStyle(
            'HeaderCenterStyle',
            parent=header_style,
            alignment=1,
        )

        body_style = ParagraphStyle(
            'BodyStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor('#0F172A'),
        )

        right_body_style = ParagraphStyle(
            'RightBodyStyle',
            parent=body_style,
            alignment=2,
        )

        center_body_style = ParagraphStyle(
            'CenterBodyStyle',
            parent=body_style,
            alignment=1,
        )

        data = [
            [
                Paragraph('<b>Vehicle Registration</b>', header_style),
                Paragraph('<b>Total Trips</b>', header_center_style),
                Paragraph('<b>Total Revenue</b>', header_right_style),
                Paragraph('<b>Total Received</b>', header_right_style),
                Paragraph('<b>Outstanding</b>', header_right_style),
            ]
        ]

        tot_trips = 0
        tot_rev = Decimal('0')
        tot_rec = Decimal('0')
        tot_out = Decimal('0')

        for vehicle in vehicle_report_data:
            revenue = Decimal(str(vehicle['total_revenue'] or 0))
            received = Decimal(str(vehicle['total_received'] or 0))
            outstanding = max(revenue - received, Decimal('0'))
            trips_count = vehicle['total_trips'] or 0

            tot_trips += trips_count
            tot_rev += revenue
            tot_rec += received
            tot_out += outstanding

            data.append([
                Paragraph(str(vehicle['vehicle__registration_number'] or '—'), body_style),
                Paragraph(str(trips_count), center_body_style),
                Paragraph(f'₹{revenue:,.2f}', right_body_style),
                Paragraph(f'₹{received:,.2f}', right_body_style),
                Paragraph(f'₹{outstanding:,.2f}', right_body_style),
            ])

        # Summary row
        data.append([
            Paragraph('<b>TOTAL</b>', body_style),
            Paragraph(f'<b>{tot_trips}</b>', center_body_style),
            Paragraph(f'<b>₹{tot_rev:,.2f}</b>', right_body_style),
            Paragraph(f'<b>₹{tot_rec:,.2f}</b>', right_body_style),
            Paragraph(f'<b>₹{tot_out:,.2f}</b>', right_body_style),
        ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[200, 110, 140, 140, 140],
        )

        apply_data_table_style(table, total_row=True)

        elements.append(table)
        finish_document(document, elements, font_name=font_name)

        return response

    # -----------------------------------
    # EXCEL EXPORT
    # -----------------------------------

    if request.GET.get('export') == 'excel':

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = 'Vehicle Report'


        headers = [
            'Vehicle',
            'Total Trips',
            'Revenue',
            'Received',
            'Outstanding',
        ]

        worksheet.append(headers)


        for cell in worksheet[1]:

            cell.font = Font(
                bold=True
            )


        for vehicle in vehicle_report_data:

            revenue = (
                vehicle['total_revenue']
                or 0
            )

            received = (
                vehicle['total_received']
                or 0
            )

            outstanding = max(
                revenue - received,
                0
            )

            worksheet.append([
                vehicle[
                    'vehicle__registration_number'
                ],
                vehicle['total_trips'],
                revenue,
                received,
                outstanding,
            ])


        # -----------------------------------
        # CURRENCY FORMAT
        # -----------------------------------

        for row in worksheet.iter_rows(
            min_row=2,
            min_col=3,
            max_col=5
        ):

            for cell in row:

                cell.number_format = (
                    '₹#,##0.00'
                )


        # -----------------------------------
        # FREEZE HEADER
        # -----------------------------------

        worksheet.freeze_panes = 'A2'


        # -----------------------------------
        # AUTO FILTER
        # -----------------------------------

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


        # -----------------------------------
        # COLUMN WIDTH
        # -----------------------------------

        for column in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                35
            )


        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            )
        )

        response[
            'Content-Disposition'
        ] = (
            'attachment; '
            'filename="vehicle_report.xlsx"'
        )


        workbook.save(
            response
        )

        return response

    # -----------------------------------
    # CSV EXPORT
    # -----------------------------------

    if request.GET.get('export') == 'csv':

        response = HttpResponse(
            content_type='text/csv'
        )

        response[
            'Content-Disposition'
        ] = 'attachment; filename="vehicle_report.csv"'

        writer = csv.writer(response)

        writer.writerow([
            'Vehicle',
            'Total Trips',
            'Revenue',
            'Received',
            'Outstanding',
        ])

        for vehicle in vehicle_report_data:

            revenue = (
                vehicle['total_revenue']
                or 0
            )

            received = (
                vehicle['total_received']
                or 0
            )

            outstanding = max(
                revenue - received,
                0
            )

            writer.writerow([
                vehicle[
                    'vehicle__registration_number'
                ],
                vehicle['total_trips'],
                revenue,
                received,
                outstanding,
            ])

        return response


    # -----------------------------------
    # SEARCH
    # -----------------------------------

    if search:

        vehicle_report_data = (
            vehicle_report_data
            .filter(
                vehicle__registration_number__icontains=search
            )
        )


    # -----------------------------------
    # OUTSTANDING
    # -----------------------------------

    vehicle_rows = []

    for vehicle in vehicle_report_data:

        revenue = (
            vehicle['total_revenue']
            or 0
        )

        received = (
            vehicle['total_received']
            or 0
        )

        outstanding = max(
            revenue - received,
            0
        )

        vehicle_rows.append({

            'vehicle_id':
                vehicle['vehicle_id'],

            'vehicle_number':
                vehicle[
                    'vehicle__registration_number'
                ],

            'total_trips':
                vehicle['total_trips'],

            'total_revenue':
                revenue,

            'total_received':
                received,

            'total_outstanding':
                outstanding,

        })


    context = {

        'vehicle_rows':
            vehicle_rows,

        'from_date':
            from_date,

        'to_date':
            to_date,

        'search':
            search,

        'vehicle_type':
            vehicle_type,

        'vehicle_id':
            vehicle_id,

        'driver_id':
            driver_id,

        'vehicle_types':
            VehicleType.objects
            .filter(is_active=True)
            .order_by('name'),

        'vehicles':
            Vehicle.objects
            .filter(
                is_active=True,
                status='ACTIVE'
            )
            .select_related('vehicle_type')
            .order_by('registration_number'),

        'drivers':
            Labour.objects
            .filter(
                is_active=True,
                status='ACTIVE'
            )
            .order_by('name'),

    }


    return render(
        request,
        'core/vehicle_report.html',
        context
    )

# -----------------------------------
# PAYMENT REPORT
# -----------------------------------
@login_required(login_url='/login/')
def payment_report(request):

    from_date = request.GET.get(
        'from_date',
        ''
    ).strip()

    to_date = request.GET.get(
        'to_date',
        ''
    ).strip()

    search = request.GET.get(
        'search',
        ''
    ).strip()

    payment_method = request.GET.get(
        'payment_method',
        ''
    ).strip()


    # -----------------------------------
    # BASE QUERYSET
    # -----------------------------------

    payments = (
        TripPayment.objects
        .annotate(
            effective_date=Coalesce('payment_date', 'trip__trip_date')
        )
        .select_related(
            'trip',
            'trip__customer',
            'customer',
            'payment_method',
        )
        .all()
    )


    # -----------------------------------
    # DATE FILTER
    # -----------------------------------

    if from_date:

        payments = payments.filter(
            effective_date__gte=from_date
        )


    if to_date:

        payments = payments.filter(
            effective_date__lte=to_date
        )


    # -----------------------------------
    # CUSTOMER SEARCH
    # -----------------------------------

    if search:

        payments = payments.filter(
            Q(trip__customer__name__icontains=search) |
            Q(customer__name__icontains=search) |
            Q(reference_number__icontains=search)
        )


    # -----------------------------------
    # PAYMENT METHOD FILTER
    # -----------------------------------

    if payment_method:

        payments = payments.filter(
            payment_method_id=payment_method
        )



    # -----------------------------------
    # PAYMENT METHODS
    # -----------------------------------

    payment_methods = (
        PaymentMethod.objects
        .filter(
            is_active=True
        )
        .order_by('name')
    )

    # -----------------------------------
    # BLANK EXPORT GUARD (no data -> block CSV/Excel/PDF)
    # -----------------------------------

    if request.GET.get('export'):

        if not payments.exists():

            messages.error(
                request,
                '❌ Export blocked: no payment data found for the selected filters.'
            )

            clean_params = {
                k: v for k, v in request.GET.items() if k != 'export'
            }

            return redirect(
                f"{reverse('core:payment_report')}?"
                f"{urlencode(clean_params)}"
            )

    # -----------------------------------
    # PDF EXPORT
    # -----------------------------------

    if request.GET.get('export') in ['pdf', 'pdf_preview']:

        response = HttpResponse(
            content_type='application/pdf'
        )

        from core.pdf_utils import (
            get_registered_font,
            build_pdf_header_elements,
            get_indian_current_time_str,
            apply_data_table_style,
            finish_document,
        )
        font_name = get_registered_font()
        _register_georgia_font()

        disposition_type = 'inline' if request.GET.get('preview') == 'true' or request.GET.get('export') == 'pdf_preview' else 'attachment'
        response[
            'Content-Disposition'
        ] = f'{disposition_type}; filename="payment_report.pdf"'

        document = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()

        meta_info = f"Generated on: {get_indian_current_time_str()}"
        if from_date or to_date:
            meta_info += f" | Filter Period: {from_date or 'All'} to {to_date or 'Present'}"

        elements = build_pdf_header_elements(
            font_name=font_name,
            report_title="Payment & Collections Transaction Report",
            report_subtitle="Detailed record of payments and collection transactions",
            extra_meta=meta_info
        )

        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8.5,
            leading=10,
            textColor=colors.white,
        )

        header_right_style = ParagraphStyle(
            'HeaderRightStyle',
            parent=header_style,
            alignment=2,
        )

        body_style = ParagraphStyle(
            'BodyStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#0F172A'),
        )

        right_body_style = ParagraphStyle(
            'RightBodyStyle',
            parent=body_style,
            alignment=2,
        )

        data = [
            [
                Paragraph('<b>Date</b>', header_style),
                Paragraph('<b>Customer</b>', header_style),
                Paragraph('<b>Trip / Ref</b>', header_style),
                Paragraph('<b>Amount</b>', header_right_style),
                Paragraph('<b>Payment Method</b>', header_style),
                Paragraph('<b>Reference No.</b>', header_style),
                Paragraph('<b>Notes</b>', header_style),
            ]
        ]

        total_amount = Decimal('0')

        for payment in payments:
            cust_name = payment.customer.name if payment.customer else (payment.trip.customer.name if payment.trip and payment.trip.customer else '—')
            trip_code = payment.trip.trip_code if payment.trip else 'On-Account'
            amt = Decimal(str(payment.amount or 0))
            total_amount += amt
            p_date = payment.payment_date or payment.effective_date

            data.append([
                Paragraph(p_date.strftime('%d-%b-%Y') if p_date else '—', body_style),
                Paragraph(str(cust_name), body_style),
                Paragraph(str(trip_code), body_style),
                Paragraph(f'₹{amt:,.2f}', right_body_style),
                Paragraph(str(payment.payment_method.name if payment.payment_method else '—'), body_style),
                Paragraph(str(payment.reference_number or '—'), body_style),
                Paragraph(str(payment.notes or '—'), body_style),
            ])

        # Summary Row
        data.append([
            Paragraph('<b>TOTAL</b>', body_style),
            Paragraph('', body_style),
            Paragraph('', body_style),
            Paragraph(f'<b>₹{total_amount:,.2f}</b>', right_body_style),
            Paragraph('', body_style),
            Paragraph('', body_style),
            Paragraph('', body_style),
        ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[75, 140, 100, 95, 100, 100, 145],
        )

        apply_data_table_style(table, total_row=True)

        elements.append(table)
        finish_document(document, elements, font_name=font_name)

        return response

    # -----------------------------------
    # EXCEL EXPORT
    # -----------------------------------

    if request.GET.get('export') == 'excel':

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = 'Payment Report'


        headers = [
            'Date',
            'Customer',
            'Trip',
            'Amount',
            'Payment Method',
            'Reference',
            'Notes',
        ]

        worksheet.append(headers)


        for cell in worksheet[1]:

            cell.font = Font(
                bold=True
            )


        for payment in payments:
            cust_name = payment.customer.name if payment.customer else (payment.trip.customer.name if payment.trip and payment.trip.customer else '')
            trip_code = payment.trip.trip_code if payment.trip else 'On-Account'

            worksheet.append([
                payment.payment_date or payment.effective_date,
                cust_name,
                trip_code,
                payment.amount,
                payment.payment_method.name if payment.payment_method else '',
                payment.reference_number or '',
                payment.notes or '',
            ])


        # -----------------------------------
        # CURRENCY FORMAT
        # -----------------------------------

        for cell in worksheet['D'][1:]:

            cell.number_format = (
                '₹#,##0.00'
            )


        # -----------------------------------
        # DATE FORMAT
        # -----------------------------------

        for cell in worksheet['A'][1:]:

            cell.number_format = (
                'DD-MMM-YYYY'
            )


        # -----------------------------------
        # FREEZE HEADER
        # -----------------------------------

        worksheet.freeze_panes = 'A2'


        # -----------------------------------
        # AUTO FILTER
        # -----------------------------------

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


        # -----------------------------------
        # COLUMN WIDTH
        # -----------------------------------

        for column in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                40
            )


        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            )
        )

        response[
            'Content-Disposition'
        ] = (
            'attachment; '
            'filename="payment_report.xlsx"'
        )


        workbook.save(
            response
        )

        return response


    # -----------------------------------
    # CSV EXPORT
    # -----------------------------------

    if request.GET.get('export') == 'csv':

        response = HttpResponse(
            content_type='text/csv'
        )

        response[
            'Content-Disposition'
        ] = 'attachment; filename="payment_report.csv"'

        writer = csv.writer(response)

        writer.writerow([
            'Date',
            'Customer',
            'Trip',
            'Amount',
            'Payment Method',
            'Reference',
            'Notes',
        ])

        for payment in payments:

            writer.writerow([
                payment.payment_date,
                payment.trip.customer.name,
                payment.trip.trip_code,
                payment.amount,
                payment.payment_method.name if payment.payment_method else '',
                payment.reference_number or '',
                payment.notes or '',
            ])

        return response


    context = {

        'payments':
            payments,

        'payment_methods':
            payment_methods,

        'from_date':
            from_date,

        'to_date':
            to_date,

        'search':
            search,

        'selected_payment_method':
            payment_method,

    }


    return render(
        request,
        'core/payment_report.html',
        context
    )


from django.http import JsonResponse
from django.views.decorators.http import require_POST
from customers.models import Customer
from master_data.models import CustomerType

@require_POST
@login_required(login_url='/login/')
def quick_add_customer(request):
    name = request.POST.get('name', '').strip()
    mobile = request.POST.get('mobile', '').strip()
    city = request.POST.get('city', '').strip()

    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required.'}, status=400)

    # Resolve or create CustomerType (code: CUSTOMER)
    customer_type = CustomerType.objects.filter(code='CUSTOMER').first()
    if not customer_type:
        customer_type = CustomerType.objects.first()
        if not customer_type:
            customer_type = CustomerType.objects.create(code='CUSTOMER', name='Customer')

    # Generate customer_code
    prefix = "CUST-"
    next_id = Customer.objects.count() + 1
    while True:
        code = f"{prefix}{next_id:03d}"
        if not Customer.objects.filter(customer_code=code).exists():
            break
        next_id += 1

    try:
        customer = Customer.objects.create(
            customer_code=code,
            name=name,
            customer_type=customer_type,
            mobile=mobile,
            city=city,
            is_active=True
        )
        return JsonResponse({
            'success': True,
            'id': customer.id,
            'name': customer.name
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/login/')
def customer_search_api(request):

    q = request.GET.get(
        'q',
        ''
    ).strip()

    if not q:
        return JsonResponse({'results': []})

    customers = (
        Customer.objects
        .filter(
            is_active=True,
            name__icontains=q
        )
        .order_by('name')[:8]
    )

    return JsonResponse({
        'results': [
            {
                'id': c.id,
                'name': c.name,
                'mobile': c.mobile or '',
                'city': c.city or '',
            }
            for c in customers
        ]
    })


# -----------------------------------
# PAYMENT REPORT - ADD PAYMENT (standalone)
# -----------------------------------
@login_required(login_url='/login/')
def payment_report_add(request):

    from trips.forms import PaymentReportForm

    next_url = get_safe_next(
        request,
        reverse('core:payment_report')
    )

    if request.method == 'POST':

        form = PaymentReportForm(
            request.POST
        )

        if form.is_valid():

            customer = form.cleaned_data['customer']

            payment = TripPayment(
                customer=customer,
                trip=None,
                payment_type='RECEIVED',
                payment_date=form.cleaned_data['payment_date'],
                amount=form.cleaned_data['amount'],
                payment_method=form.cleaned_data.get('payment_method'),
                reference_number=form.cleaned_data.get('reference_number') or '',
                notes=form.cleaned_data.get('notes') or '',
            )

            payment.save()

            messages.success(
                request,
                f'✅ ₹{payment.amount} payment added for {customer.name}!'
            )

            return redirect(next_url)

    else:

        form = PaymentReportForm()

    selected_customer_name = None

    if form['customer'].value():

        selected_customer = Customer.objects.filter(
            pk=form['customer'].value()
        ).first()

        if selected_customer:
            selected_customer_name = selected_customer.name

    context = {
        'form': form,
        'next_url': next_url,
        'selected_customer_name': selected_customer_name,
    }

    return render(
        request,
        'core/payment_report_add.html',
        context
    )

# -----------------------------------
# OVERDUE REMINDERS LIST
# -----------------------------------
@login_required(login_url='/login/')
def overdue_reminders(request):
    from decimal import Decimal as _D
    from trips.models import TripPayment
    from django.db.models import Q as _Q
    from django.db.models.functions import Coalesce as _Coalesce

    today = timezone.localdate()
    customers = Customer.objects.filter(is_active=True).order_by('name')

    rows = []
    total_due = _D('0')

    for customer in customers:
        trips = list(Trip.objects.select_related('vehicle', 'vehicle__vehicle_type').filter(customer=customer))
        payments = (
            TripPayment.objects.filter(
                _Q(customer=customer) | _Q(trip__customer=customer)
            ).annotate(effective_date=_Coalesce('payment_date', 'trip__trip_date'))
        )

        opening = customer.opening_balance or _D('0')
        sales = sum(
            (-t.total_amount if t.transaction_type == 'VENDOR_SUPPLY' else t.total_amount)
            for t in trips
        )
        received = sum(
            (-p.amount if p.payment_type == 'PAID' else p.amount)
            for p in payments if p.effective_date
        )
        outstanding = opening + sales - received

        if outstanding <= 0:
            continue

        last_payment = (
            payments.exclude(payment_type='PAID')
            .filter(effective_date__isnull=False)
            .order_by('-effective_date')
            .first()
        )

        digits = ''.join(ch for ch in (customer.mobile or '') if ch.isdigit())
        if len(digits) == 10:
            wa_number = '91' + digits
        elif len(digits) == 11 and digits.startswith('0'):
            wa_number = '91' + digits[1:]
        elif len(digits) == 12 and digits.startswith('91'):
            wa_number = digits
        else:
            wa_number = ''

        reminder_text = (
            f"Namaste {customer.name} ji 🙏\n"
            f"{today.strftime('%d %b %Y')} tak aapka pending: *₹{outstanding:,.0f}*\n"
            f"Kripya jald payment kar dijiye.\n"
            f"Dhanyavaad! 🙏"
        )

        rows.append({
            'customer': customer,
            'outstanding': outstanding,
            'trips_count': len(trips),
            'last_payment_date': last_payment.effective_date if last_payment else None,
            'days_since_payment': (
                (today - last_payment.effective_date).days
                if last_payment and last_payment.effective_date else None
            ),
            'wa_link': (
                f"https://wa.me/{wa_number}?text={reminder_text.replace(' ', '%20').replace(chr(10), '%0A')}"
                if wa_number else ''
            ),
            'phone_raw': customer.mobile or '',
        })
        total_due += outstanding

    rows.sort(key=lambda r: r['outstanding'], reverse=True)

    context = {
        'rows': rows,
        'total_due': total_due,
        'as_of': today,
    }

    return render(request, 'core/overdue_reminders.html', context)
