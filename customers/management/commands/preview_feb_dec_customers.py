from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from customers.models import Customer


TRIP_FILES = [
    "historical_data/Feb_to_Dec_Hyva_White_Sand_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Hyva_Fly_Ash_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Halfton_White_Sand_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Tractor_White_Sand_CLEAN.xlsx",
]


def clean(value):
    if value is None:
        return ""

    return str(value).strip()


class Command(BaseCommand):

    help = "Preview customers found in Feb-Dec Excel files but missing from DB."

    def handle(self, *args, **options):

        # ---------------------------------------------------------
        # EXISTING DB CUSTOMERS
        # ---------------------------------------------------------

        existing_customers = {
            customer.customer_code: customer.name
            for customer in Customer.objects.all()
        }

        # ---------------------------------------------------------
        # READ UNIQUE CUSTOMERS FROM 4 TRIP FILES
        # ---------------------------------------------------------

        excel_customers = {}

        for file_path in TRIP_FILES:

            path = Path(file_path)

            self.stdout.write(
                f"Reading: {path.name}"
            )

            wb = openpyxl.load_workbook(
                path,
                data_only=True,
            )

            for ws in wb.worksheets:

                headers = {
                    clean(ws.cell(1, col).value): col
                    for col in range(
                        1,
                        ws.max_column + 1
                    )
                }

                if "Customer Code" not in headers:
                    continue

                if "Customer Name" not in headers:
                    continue

                code_col = headers["Customer Code"]
                name_col = headers["Customer Name"]

                for row in range(
                    2,
                    ws.max_row + 1
                ):

                    customer_code = clean(
                        ws.cell(
                            row,
                            code_col
                        ).value
                    )

                    customer_name = clean(
                        ws.cell(
                            row,
                            name_col
                        ).value
                    )

                    if not customer_code:
                        continue

                    # Internal Stock / unknown customer
                    if customer_code == "0":
                        continue

                    if not customer_name:
                        customer_name = "Customer Unknown"

                    if customer_code not in excel_customers:

                        excel_customers[
                            customer_code
                        ] = customer_name

                    elif (
                        excel_customers[customer_code]
                        != customer_name
                    ):

                        self.stdout.write(
                            self.style.WARNING(
                                f"WARNING: {customer_code} "
                                f"has different names: "
                                f"'{excel_customers[customer_code]}' "
                                f"/ '{customer_name}'"
                            )
                        )

        # ---------------------------------------------------------
        # FIND MISSING CUSTOMERS
        # ---------------------------------------------------------

        missing_customers = {
            code: name
            for code, name in excel_customers.items()
            if code not in existing_customers
        }

        # ---------------------------------------------------------
        # SUMMARY
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write(
            "FEBRUARY - DECEMBER CUSTOMER PREVIEW"
        )
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Customers in Excel : {len(excel_customers)}"
        )

        self.stdout.write(
            f"Customers in DB    : {len(existing_customers)}"
        )

        self.stdout.write(
            f"Missing customers  : {len(missing_customers)}"
        )

        # ---------------------------------------------------------
        # MISSING CUSTOMER LIST
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write(
            "MISSING CUSTOMERS"
        )
        self.stdout.write("-" * 70)

        for code, name in sorted(
            missing_customers.items()
        ):

            self.stdout.write(
                f"{code:<12} | {name}"
            )

        # ---------------------------------------------------------
        # SAME CODE / DIFFERENT NAME CHECK
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write(
            "EXISTING CODE / NAME CHECK"
        )
        self.stdout.write("-" * 70)

        conflicts = []

        for code, excel_name in excel_customers.items():

            if code not in existing_customers:
                continue

            db_name = existing_customers[code]

            if db_name.strip().lower() != excel_name.strip().lower():

                conflicts.append(
                    (
                        code,
                        db_name,
                        excel_name,
                    )
                )

        if conflicts:

            for code, db_name, excel_name in conflicts:

                self.stdout.write(
                    self.style.WARNING(
                        f"{code} | DB: '{db_name}' "
                        f"| Excel: '{excel_name}'"
                    )
                )

        else:

            self.stdout.write(
                self.style.SUCCESS(
                    "No existing-code/name conflicts found."
                )
            )

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write(
            "PREVIEW ONLY - NO DATABASE CHANGES WERE MADE."
        )
        self.stdout.write("=" * 70)
