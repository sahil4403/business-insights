from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from customers.models import Customer
from master_data.models import CustomerType


TRIP_FILES = [
    "historical_data/Feb_to_Dec_Hyva_White_Sand_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Hyva_Fly_Ash_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Halfton_White_Sand_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Tractor_White_Sand_CLEAN.xlsx",
]


def clean(value):
    if value is None:
        return ""

    return str(value).strip()


class Command(BaseCommand):

    help = "Import missing February-December customers."

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Preview customer creation without changing database.",
        )

    def handle(self, *args, **options):

        dry_run = options["dry_run"]

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write(
            "FEBRUARY - DECEMBER CUSTOMER IMPORT"
        )
        self.stdout.write("=" * 70)

        if dry_run:
            self.stdout.write(
                self.style.WARNING(
                    "DRY RUN - NO DATABASE CHANGES WILL BE MADE"
                )
            )
        else:
            self.stdout.write(
                self.style.ERROR(
                    "LIVE IMPORT - DATABASE WILL BE CHANGED"
                )
            )

        # ---------------------------------------------------------
        # CUSTOMER TYPE
        # ---------------------------------------------------------

        customer_type = CustomerType.objects.filter(
            code="CUSTOMER"
        ).first()

        if not customer_type:
            raise RuntimeError(
                "CustomerType with code='CUSTOMER' does not exist."
            )

        # ---------------------------------------------------------
        # EXISTING CUSTOMERS
        # ---------------------------------------------------------

        existing_customers = {
            customer.customer_code: customer
            for customer in Customer.objects.all()
        }

        # ---------------------------------------------------------
        # READ CUSTOMERS FROM EXCEL
        # ---------------------------------------------------------

        excel_customers = {}

        for file_path in TRIP_FILES:

            path = Path(file_path)

            self.stdout.write(
                f"Reading: {path.name}"
            )

            wb = openpyxl.load_workbook(
                path,
                data_only=True,
            )

            for ws in wb.worksheets:

                headers = {
                    clean(ws.cell(1, col).value): col
                    for col in range(
                        1,
                        ws.max_column + 1
                    )
                }

                if "Customer Code" not in headers:
                    continue

                if "Customer Name" not in headers:
                    continue

                code_col = headers["Customer Code"]
                name_col = headers["Customer Name"]

                for row in range(
                    2,
                    ws.max_row + 1
                ):

                    customer_code = clean(
                        ws.cell(
                            row,
                            code_col
                        ).value
                    )

                    customer_name = clean(
                        ws.cell(
                            row,
                            name_col
                        ).value
                    )

                    # Ignore blank/customer-zero rows.
                    if not customer_code:
                        continue

                    if customer_code == "0":
                        continue

                    if not customer_name:
                        customer_name = "Customer Unknown"

                    # First occurrence becomes source value.
                    if customer_code not in excel_customers:

                        excel_customers[
                            customer_code
                        ] = customer_name

                    # Same code but different spelling/capitalization.
                    elif (
                        excel_customers[customer_code].strip().lower()
                        != customer_name.strip().lower()
                    ):

                        raise RuntimeError(
                            f"Customer code {customer_code} "
                            f"has conflicting names: "
                            f"'{excel_customers[customer_code]}' "
                            f"and '{customer_name}'."
                        )

        # ---------------------------------------------------------
        # FIND MISSING CUSTOMERS
        # ---------------------------------------------------------

        missing_customers = {
            code: name
            for code, name in excel_customers.items()
            if code not in existing_customers
        }

        # ---------------------------------------------------------
        # SUMMARY
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("CUSTOMER IMPORT SUMMARY")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Customers in Excel : {len(excel_customers)}"
        )

        self.stdout.write(
            f"Customers in DB    : {len(existing_customers)}"
        )

        self.stdout.write(
            f"New customers      : {len(missing_customers)}"
        )

        # ---------------------------------------------------------
        # SHOW NEW CUSTOMERS
        # ---------------------------------------------------------

        self.stdout.write("")
        self.stdout.write("NEW CUSTOMERS")
        self.stdout.write("-" * 70)

        for code, name in sorted(
            missing_customers.items()
        ):

            self.stdout.write(
                f"{code:<12} | {name}"
            )

        # ---------------------------------------------------------
        # NOTHING TO CREATE
        # ---------------------------------------------------------

        if not missing_customers:

            self.stdout.write("")
            self.stdout.write(
                self.style.SUCCESS(
                    "No new customers need to be created."
                )
            )
            return

        # ---------------------------------------------------------
        # DRY RUN
        # ---------------------------------------------------------

        if dry_run:

            self.stdout.write("")
            self.stdout.write("=" * 70)
            self.stdout.write(
                self.style.SUCCESS(
                    f"DRY RUN PASSED - "
                    f"{len(missing_customers)} customers "
                    f"would be created."
                )
            )
            self.stdout.write(
                "NO DATABASE CHANGES WERE MADE."
            )
            self.stdout.write("=" * 70)

            return

        # ---------------------------------------------------------
        # LIVE IMPORT
        # ---------------------------------------------------------

        new_customers = [
            Customer(
                customer_code=code,
                name=name,
                customer_type=customer_type,
                is_active=True,
            )
            for code, name in sorted(
                missing_customers.items()
            )
        ]

        with transaction.atomic():

            Customer.objects.bulk_create(
                new_customers,
                batch_size=100,
            )

        # ---------------------------------------------------------
        # VERIFY
        # ---------------------------------------------------------

        imported_codes = set(
            Customer.objects.filter(
                customer_code__in=missing_customers.keys()
            ).values_list(
                "customer_code",
                flat=True,
            )
        )

        if imported_codes != set(
            missing_customers.keys()
        ):

            raise RuntimeError(
                "Customer post-import verification failed."
            )

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("CUSTOMER IMPORT COMPLETE")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Customers created : {len(new_customers)}"
        )

        self.stdout.write(
            f"Customers verified: {len(imported_codes)}"
        )

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                "CUSTOMER POST-IMPORT VERIFICATION PASSED."
            )
        )
