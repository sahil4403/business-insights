"""
Labour views — flexible per-trip groups, extras, advances, driver payments,
old balance tracking & settlement.

All views require login. Mobile-first styling. Date filters use the same
"from / to" pattern as the rest of the app.
"""

from collections import defaultdict
from datetime import date
from decimal import Decimal
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

from .forms import (
    LabourForm,
    LabourTripGroupForm,
    LabourHyvaTripForm,
    LabourExtraPaymentForm,
    LabourAdvanceForm,
    LabourAdvanceMultiForm,
    LabourDriverPaymentForm,
    LabourSettlementForm,
    LabourRoziForm,
)
from .models import (
    Labour,
    LabourTripGroup,
    LabourExtraPayment,
    LabourAdvance,
    LabourDriverPayment,
    LabourOldBalance,
    LabourSettlement,
    LabourRozi,
)


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def _parse_date(s, default=None):
    if not s:
        return default
    try:
        return date.fromisoformat(s)
    except (TypeError, ValueError):
        return default


def _ensure_old_balance(labour):
    ob, _ = LabourOldBalance.objects.get_or_create(labour=labour, defaults={'amount': Decimal('0')})
    return ob


def _per_labour_trip_share(labour, period_start, period_end):
    """
    Return total trip-share for this labour between period_start..period_end.
    Each trip group splits total_amount across its labourers.
    """
    total = Decimal('0')
    for grp in LabourTripGroup.objects.filter(
        date__gte=period_start, date__lte=period_end
    ).prefetch_related('labourers'):
        n = grp.labourers.count()
        if n and grp.labourers.filter(pk=labour.pk).exists():
            total += grp.total_amount / n
    return total


def _day_aggregates_for_labour(labour, period_start, period_end):
    """
    Build per-day rows: Date | Trips amount | Extra | Total | Advance.
    Returns list of dicts.
    """
    # Build a per-date map of all values
    days = {}

    def _ensure(d):
        if d not in days:
            days[d] = {
                'date': d,
                'trips_amount': Decimal('0'),
                'extra_amount': Decimal('0'),
                'advance_amount': Decimal('0'),
            }
        return days[d]

    # Trips: walk groups in range
    for grp in LabourTripGroup.objects.filter(
        date__gte=period_start, date__lte=period_end
    ).prefetch_related('labourers'):
        n = grp.labourers.count()
        if n and grp.labourers.filter(pk=labour.pk).exists():
            _ensure(grp.date)['trips_amount'] += grp.total_amount / n

    # Extras
    for ep in LabourExtraPayment.objects.filter(
        labour=labour, date__gte=period_start, date__lte=period_end
    ):
        _ensure(ep.date)['extra_amount'] += ep.amount

    # Rozi (daily wages) — counted like extra into the day total
    for rz in LabourRozi.objects.filter(
        labour=labour, date__gte=period_start, date__lte=period_end
    ):
        _ensure(rz.date)['extra_amount'] += rz.amount

    # Advances
    for ad in LabourAdvance.objects.filter(
        labour=labour, date__gte=period_start, date__lte=period_end
    ):
        _ensure(ad.date)['advance_amount'] += ad.amount

    # Sort by date desc
    rows = list(days.values())
    rows.sort(key=lambda r: r['date'], reverse=True)
    return rows


# ----------------------------------------------------------------------------
# List
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
def labour_list(request):
    category_filter = request.GET.get('category', '')

    labours_qs = Labour.objects.filter(is_active=True).exclude(is_vendor=True).order_by('name')

    # Pre-compute outstanding per labour
    ob_map = {ob.labour_id: ob.amount for ob in LabourOldBalance.objects.all()}

    # Count trip groups this month for context
    today = timezone.localdate()
    month_start = today.replace(day=1)

    def _card_for(l):
        return {
            'obj': l,
            'old_balance': ob_map.get(l.id, Decimal('0')),
            'month_trip_groups': l.trip_groups.filter(date__gte=month_start).count(),
        }

    total_outstanding = Decimal('0')
    driver_count_total = 0

    # Build per-category tiles (main labour page shows category cards only)
    category_meta_map = {
        'TRACTOR': {'icon': '🚜', 'bg': '#ecfdf5'},
        'HYVA_DRIVER': {'icon': '🚛', 'bg': '#eff6ff'},
        'JCB_OPERATOR': {'icon': '🏗️', 'bg': '#fef3c7'},
        'MISTRI': {'icon': '👷', 'bg': '#fdf2f8'},
    }
    categories = []
    for code, label in Labour.CATEGORY_CHOICES:
        cat_labours = labours_qs.filter(category=code)
        cat_count = cat_labours.count()
        driver_count_total += sum(1 for l in cat_labours if l.is_driver)
        meta = category_meta_map.get(code, {'icon': '👷', 'bg': '#f1f5f9'})
        add_url = None
        add_label = ''
        if code == 'TRACTOR':
            add_url = 'labour:trip_add'
            add_label = 'Add Trip'
        elif code == 'HYVA_DRIVER':
            add_url = 'labour:hyva_trip_add'
            add_label = 'Add Hyva Trip'
        elif code == 'JCB_OPERATOR':
            add_url = ''
            add_label = 'Coming Soon'
        elif code == 'MISTRI':
            add_url = 'labour:rozi_multi'
            add_label = 'Add Mistri Rozi'
        categories.append({
            'code': code,
            'label': label,
            'count': cat_count,
            'icon': meta['icon'],
            'bg': meta['bg'],
            'add_url': add_url,
            'add_label': add_label,
            'add_query': 'category=MISTRI' if code == 'MISTRI' else '',
            'coming_soon': code == 'JCB_OPERATOR',
        })

    cards = []
    for l in labours_qs:
        cards.append(_card_for(l))
    total_outstanding = sum((c['old_balance'] for c in cards), Decimal('0'))

    categories_meta = [
        {'code': c['code'], 'label': c['label'], 'count': c['count']}
        for c in categories
    ]
    active_category = category_filter or ''
    active_category_label = ''
    if active_category:
        for code, label in Labour.CATEGORY_CHOICES:
            if code == active_category:
                active_category_label = label
                break

    # -----------------------------------
    # EXPORTS — labour summary (list)
    # -----------------------------------

    export = request.GET.get('export')
    if export in ('pdf', 'pdf_preview'):
        response = HttpResponse(content_type='application/pdf')

        from core.pdf_utils import (
            get_registered_font,
            build_pdf_header_elements,
            get_indian_current_time_str,
            apply_data_table_style,
            finish_document,
        )
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph

        font_name = get_registered_font()

        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=25,
            leftMargin=25,
            topMargin=25,
            bottomMargin=25,
        )

        styles = getSampleStyleSheet()
        header_style = ParagraphStyle(
            'HeaderStyle', parent=styles['Normal'],
            fontName=font_name, fontSize=9, leading=11, textColor=colors.white,
        )
        header_right = ParagraphStyle('HeaderRight', parent=header_style, alignment=2)
        header_center = ParagraphStyle('HeaderCenter', parent=header_style, alignment=1)
        body_style = ParagraphStyle(
            'BodyStyle', parent=styles['Normal'],
            fontName=font_name, fontSize=8.5, leading=11,
            textColor=colors.HexColor('#0F172A'),
        )
        right_body = ParagraphStyle('RightBody', parent=body_style, alignment=2)
        center_body = ParagraphStyle('CenterBody', parent=body_style, alignment=1)

        elements = build_pdf_header_elements(
            font_name=font_name,
            report_title="Labour Wise Summary Report",
            report_subtitle="All labour — driver status, monthly trip entries and outstanding",
            extra_meta=f"Generated on: {get_indian_current_time_str()}",
        )

        data = [
            [
                Paragraph('<b>Labour</b>', header_style),
                Paragraph('<b>Driver</b>', header_center),
                Paragraph('<b>Mobile</b>', header_center),
                Paragraph('<b>Trip Entries (Month)</b>', header_center),
                Paragraph('<b>Outstanding (₹)</b>', header_right),
            ]
        ]

        total_trips = 0
        for c in cards:
            l = c['obj']
            data.append([
                Paragraph(str(l.name or '—'), body_style),
                Paragraph('Yes' if l.is_driver else 'No', center_body),
                Paragraph(str(l.mobile or '—'), center_body),
                Paragraph(str(c.get('month_trip_groups', 0)), center_body),
                Paragraph(f"₹{c['old_balance']:,.2f}", right_body),
            ])
            total_trips += c.get('month_trip_groups', 0)

        data.append([
            Paragraph('<b>TOTAL</b>', body_style),
            Paragraph('', center_body),
            Paragraph('', center_body),
            Paragraph(f'<b>{total_trips}</b>', center_body),
            Paragraph(f'<b>₹{total_outstanding:,.2f}</b>', right_body),
        ])

        table = Table(data, repeatRows=1, colWidths=[180, 70, 130, 150, 150])
        apply_data_table_style(table, total_row=True)
        elements.append(table)
        finish_document(document, elements, font_name=font_name)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        disposition = 'inline' if request.GET.get('preview') == 'true' or export == 'pdf_preview' else 'attachment'
        response['Content-Disposition'] = f'{disposition}; filename="labour_report.pdf"'
        return response

    if export == 'excel':
        from openpyxl import Workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Labour Report'
        worksheet.append(['Labour', 'Driver', 'Mobile', 'Trip Entries (Month)', 'Outstanding (₹)'])
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for c in cards:
            l = c['obj']
            worksheet.append([
                l.name,
                'Yes' if l.is_driver else 'No',
                l.mobile or '',
                c.get('month_trip_groups', 0),
                float(c['old_balance']),
            ])
        worksheet.append(['TOTAL', '', '', sum(c.get('month_trip_groups', 0) for c in cards), float(total_outstanding)])
        for row in worksheet.iter_rows(min_row=2, max_col=5):
            for cell in row:
                if cell.column == 5:
                    cell.number_format = '₹#,##0.00'
        worksheet.column_dimensions['A'].width = 24
        worksheet.column_dimensions['C'].width = 16
        worksheet.column_dimensions['D'].width = 22
        worksheet.column_dimensions['E'].width = 18
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="labour_report.xlsx"'
        workbook.save(response)
        return response

    if export == 'csv':
        import csv as _csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="labour_report.csv"'
        writer = _csv.writer(response)
        writer.writerow(['Labour', 'Driver', 'Mobile', 'Trip Entries (Month)', 'Outstanding (₹)'])
        for c in cards:
            l = c['obj']
            writer.writerow([
                l.name,
                'Yes' if l.is_driver else 'No',
                l.mobile or '',
                c.get('month_trip_groups', 0),
                c['old_balance'],
            ])
        return response

    context = {
        'categories': categories,
        'categories_meta': categories_meta,
        'labour_count': sum(c['count'] for c in categories),
        'driver_count': driver_count_total,
    }
    return render(request, 'labour/labour_list.html', context)


@login_required(login_url='/login/')
def labour_category_detail(request, category_code):
    """Show all labour belonging to a single category."""
    valid = dict(Labour.CATEGORY_CHOICES)
    if category_code not in valid:
        messages.info(request, 'Invalid category.')
        return redirect('labour:list')

    labours_qs = Labour.objects.filter(is_active=True, category=category_code).exclude(is_vendor=True).order_by('name')
    ob_map = {ob.labour_id: ob.amount for ob in LabourOldBalance.objects.all()}
    today = timezone.localdate()
    month_start = today.replace(day=1)

    cards = []
    for l in labours_qs:
        cards.append({
            'obj': l,
            'old_balance': ob_map.get(l.id, Decimal('0')),
            'month_trip_groups': l.trip_groups.filter(date__gte=month_start).count(),
        })

    # Quick-switch tabs metadata
    all_labours = Labour.objects.filter(is_active=True).exclude(is_vendor=True)
    categories_meta = [
        {
            'code': code,
            'label': label,
            'count': all_labours.filter(category=code).count(),
        }
        for code, label in Labour.CATEGORY_CHOICES
    ]

    add_url = None
    add_label = ''
    add_query = ''
    if category_code == 'TRACTOR':
        add_url = 'labour:trip_add'
        add_label = 'Add Trip'
        add_query = 'category=TRACTOR'
    elif category_code == 'HYVA_DRIVER':
        add_url = 'labour:hyva_trip_add'
        add_label = 'Add Hyva Trip'
    elif category_code == 'JCB_OPERATOR':
        add_url = ''
        add_label = 'Add Trip'
        add_query = ''
    elif category_code == 'MISTRI':
        add_url = 'labour:rozi_multi'
        add_label = 'Add Mistri Rozi'
        add_query = 'category=MISTRI'

    category_meta_map = {
        'TRACTOR': {'icon': '🚜', 'bg': '#ecfdf5'},
        'HYVA_DRIVER': {'icon': '🚛', 'bg': '#eff6ff'},
        'JCB_OPERATOR': {'icon': '🏗️', 'bg': '#fef3c7'},
        'MISTRI': {'icon': '👷', 'bg': '#fdf2f8'},
    }
    meta = category_meta_map.get(category_code, {'icon': '👷', 'bg': '#eef7f5'})

    return render(request, 'labour/labour_category_detail.html', {
        'cards': cards,
        'category_label': valid[category_code],
        'category_code': category_code,
        'labour_count': len(cards),
        'categories_meta': categories_meta,
        'add_url': add_url,
        'add_label': add_label,
        'add_query': add_query,
        'cat_bg': meta['bg'],
        'cat_icon': meta['icon'],
    })


@login_required(login_url='/login/')
def labour_create(request):
    preselect = request.GET.get('category', '')
    if preselect not in dict(Labour.CATEGORY_CHOICES):
        preselect = ''

    if request.method == 'POST':
        form = LabourForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Labour "{obj.name}" added.')
            return redirect('labour:list')
    else:
        form = LabourForm(initial={'category': preselect} if preselect else None)

    button_label = dict(Labour.CATEGORY_CHOICES).get(preselect, 'Labour') if preselect else 'Labour'
    return render(request, 'labour/labour_form.html', {
        'form': form,
        'page_title': f'Add {button_label}',
        'preselect_category': preselect,
    })


@login_required(login_url='/login/')
def labour_edit(request, labour_id):
    labour = get_object_or_404(Labour, pk=labour_id)
    if request.method == 'POST':
        form = LabourForm(request.POST, instance=labour)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Labour "{obj.name}" updated.')
            return redirect('labour:detail', labour_id=obj.id)
    else:
        form = LabourForm(instance=labour)

    return render(request, 'labour/labour_form.html', {
        'form': form,
        'page_title': f'Edit {labour.name}',
    })


# ----------------------------------------------------------------------------
# Detail
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
@require_POST
def labour_deactivate(request, labour_id):
    """Soft-delete: hide labour from list/details by setting is_active=False.
    Historical records (trips, advances, settlements) stay safe in the DB."""
    labour = get_object_or_404(Labour, pk=labour_id)
    labour.is_active = False
    labour.status = 'INACTIVE'
    labour.save(update_fields=['is_active', 'status', 'updated_at'])
    messages.success(request, f'"{labour.name}" has been removed. Historical data is preserved.')
    return redirect('labour:list')


@login_required(login_url='/login/')
def labour_detail(request, labour_id):
    labour = get_object_or_404(Labour, pk=labour_id)
    ob = _ensure_old_balance(labour)

    # Date range filter (default: current month)
    today = timezone.localdate()
    default_start = today.replace(day=1)
    period_start = _parse_date(request.GET.get('from_date'), default_start)
    period_end = _parse_date(request.GET.get('to_date'), today)

    rows = _day_aggregates_for_labour(labour, period_start, period_end)

    # Summary block
    trip_total = sum((r['trips_amount'] for r in rows), Decimal('0'))
    extra_total = sum((r['extra_amount'] for r in rows), Decimal('0'))
    advance_total = sum((r['advance_amount'] for r in rows), Decimal('0'))

    # Driver payment for the period (overlapping ranges)
    driver_qs = LabourDriverPayment.objects.filter(labour=labour).filter(
        period_start__lte=period_end, period_end__gte=period_start
    )
    driver_total = driver_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    # This labour's individual trip entries in the period (per-trip breakdown)
    trip_groups = list(
        LabourTripGroup.objects.filter(
            labourers=labour, date__gte=period_start, date__lte=period_end
        ).prefetch_related('labourers').order_by('-date', '-id')
    )

    # Extras in the period
    extras_list = list(LabourExtraPayment.objects.filter(
        labour=labour, date__gte=period_start, date__lte=period_end
    ).order_by('-date', '-id'))

    # Rozi (daily wages) also acts as an extra payment, must be visible to edit/delete
    rozi_list = list(LabourRozi.objects.filter(
        labour=labour, date__gte=period_start, date__lte=period_end
    ).order_by('-date', '-id'))

    for rz in rozi_list:
        rz.is_rozi = True
        if not hasattr(rz, 'note'):
            rz.note = rz.get_day_type_display()
        elif rz.note:
            rz.note = f"{rz.get_day_type_display()} - {rz.note}"
        else:
            rz.note = rz.get_day_type_display()
        extras_list.append(rz)
    
    # Re-sort extras_list by date descending after appending rozis
    extras_list.sort(key=lambda e: getattr(e, 'date'), reverse=True)

    # Advances in the period
    advances_list = list(LabourAdvance.objects.filter(
        labour=labour, date__gte=period_start, date__lte=period_end
    ).order_by('-date', '-id'))

    # Group trips, extras, and advances by date (each date shows all its load lines + extras + advances + day total)
    trip_groups_by_date = []
    _day_order = {}

    def _get_bucket(d):
        if d not in _day_order:
            _day_order[d] = len(trip_groups_by_date)
            trip_groups_by_date.append({
                'date': d,
                'groups': [],
                'extras': [],
                'advances': [],
                'trip_total': Decimal('0'),
                'trip_count': 0,
            })
        return trip_groups_by_date[_day_order[d]]

    for g in trip_groups:
        bucket = _get_bucket(g.date)
        bucket['groups'].append(g)
        bucket['trip_total'] += g.total_amount
        bucket['trip_count'] += g.trip_count

    for e in extras_list:
        bucket = _get_bucket(e.date)
        bucket['extras'].append(e)

    for a in advances_list:
        bucket = _get_bucket(a.date)
        bucket['advances'].append(a)

    trip_groups_by_date.sort(key=lambda x: x['date'], reverse=True)

    total_salary = trip_total + extra_total + driver_total
    payment = total_salary - advance_total
    # Excel formula: Final Amount = Payment − Old Balance
    final_amount = payment - ob.amount

    # Settlements within range
    settlements = LabourSettlement.objects.filter(
        labour=labour, settlement_date__gte=period_start, settlement_date__lte=period_end
    )

    # Pending settlements before this period
    pending_settlements = LabourSettlement.objects.filter(
        labour=labour
    ).order_by('-settlement_date')[:5]

    latest_settlement = LabourSettlement.objects.filter(labour=labour).order_by(
        '-settlement_date', '-id'
    ).first()

    # Category-aware quick action (trip entry link differs per category)
    cat_icon = {
        'TRACTOR': '🚜', 'HYVA_DRIVER': '🚛',
        'JCB_OPERATOR': '🏗️', 'MISTRI': '👷',
    }.get(labour.category, '🚜')
    if labour.category == 'HYVA_DRIVER':
        trip_action = {'url': 'labour:hyva_trip_add', 'label': 'Add Hyva Trip', 'icon': '🚛',
                       'query': f'labour_id={labour.id}'}
    elif labour.category == 'JCB_OPERATOR':
        trip_action = {'url': '', 'label': 'Add Trip', 'icon': cat_icon, 'coming_soon': True}
    elif labour.category == 'MISTRI':
        trip_action = {'url': 'labour:rozi_multi', 'label': 'Add Mistri Rozi', 'icon': '👷', 'rozi': True}
    else:
        trip_action = {'url': 'labour:trip_add', 'label': 'Quick Add Trip', 'icon': cat_icon,
                       'query': f'category={labour.category}'}

    context = {
        'labour': labour,
        'old_balance': ob,
        'rows': rows,
        'period_start': period_start,
        'period_end': period_end,
        'trip_total': trip_total,
        'extra_total': extra_total,
        'driver_total': driver_total,
        'advance_total': advance_total,
        'total_salary': total_salary,
        'payment': payment,
        'latest_settlement_id': latest_settlement.id if latest_settlement else None,
        'final_amount': final_amount,
        'settlements': settlements,
        'pending_settlements': pending_settlements,
        'driver_payments': driver_qs,
        'cat_icon': cat_icon,
        'trip_action': trip_action,
        'trip_groups': trip_groups,
        'trip_groups_by_date': trip_groups_by_date,
    }
    return render(request, 'labour/labour_detail.html', context)


# ----------------------------------------------------------------------------
# Trip Group
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
def trip_group_create(request):
    """Create a new trip-group entry. Supports 'add_another' flow."""
    add_another = request.GET.get('next') == 'add_another' or \
        request.POST.get('add_another') == '1'
    category = request.GET.get('category') or request.POST.get('category') or ''
    if category not in dict(Labour.CATEGORY_CHOICES):
        category = 'TRACTOR'

    form_kwargs = {'category': category}

    if request.method == 'POST':
        form = LabourTripGroupForm(request.POST, **form_kwargs)
        if form.is_valid():
            obj = form.save()
            messages.success(
                request,
                f'Trip entry saved · {obj.trip_count} trips · ₹{obj.total_amount}.',
            )
            if request.POST.get('add_another') == '1':
                return redirect(f"{request.path}?next=add_another&date={obj.date.isoformat()}{('&category=' + category) if category else ''}")
            return redirect('labour:list')
    else:
        initial_date = _parse_date(request.GET.get('date'), timezone.localdate())
        form = LabourTripGroupForm(initial={'date': initial_date}, **form_kwargs)

    context = {
        'form': form,
        'page_title': 'Add Trip Entry',
        'show_add_another': True,
    }
    return render(request, 'labour/trip_group_form.html', context)


# ----------------------------------------------------------------------------
# Hyva Driver trip entry — load type auto-sets the per-trip rate, optional
# daily bhatta (₹200) recorded per selected labourer as an extra payment.
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
def hyva_trip_create(request):
    """Hyva Driver trip entry — supports multiple load types on one day.

    The form lets the user add several rows; each row is a (load type, trip
    count) line. On save we create ONE LabourTripGroup per line (all sharing
    the same date and workers). An optional bhatta (₹200/day) is recorded as
    a single extra payment per selected labourer.
    """
    load_choices = [c for c, _ in LabourTripGroup.HYVA_LOAD_CHOICES if c]
    load_rates = LabourTripGroup.HYVA_LOAD_RATES
    rates = {code: load_rates[code] for code in load_choices}

    if request.method == 'POST':
        form = LabourHyvaTripForm(request.POST)
        rows = []
        i = 0
        while True:
            lt_code = request.POST.get(f'load_type_{i}')
            trip_raw = request.POST.get(f'trip_count_{i}')
            if lt_code is None:
                break
            try:
                trips = int(trip_raw or 0)
            except (TypeError, ValueError):
                trips = 0
            if lt_code in rates and trips and trips > 0:
                rows.append({'load_type': lt_code, 'trip_count': trips})
            i += 1

        if form.is_valid() and rows:
            date = form.cleaned_data['date']
            labourers = list(form.cleaned_data['labourers'])
            bhatta = form.cleaned_data.get('bhatta')
            note = form.cleaned_data.get('note', '')

            groups = []
            total = 0
            for row in rows:
                grp = LabourTripGroup.objects.create(
                    date=date,
                    trip_count=row['trip_count'],
                    rate_per_trip=Decimal(load_rates[row['load_type']]),
                    load_type=row['load_type'],
                    fill_type='HAND',
                    total_amount=Decimal(row['trip_count']) * Decimal(load_rates[row['load_type']]),
                    note=note,
                )
                grp.labourers.set(labourers)
                groups.append(grp)
                total += grp.total_amount

            if bhatta:
                for lab in labourers:
                    LabourExtraPayment.objects.create(
                        labour=lab,
                        date=date,
                        amount=Decimal(LabourHyvaTripForm.BHATTA_AMOUNT),
                        note='Bhatta',
                    )

            total_trips = sum(g.trip_count for g in groups)
            messages.success(
                request,
                f'Hyva trip saved · {len(groups)} load line(s) · '
                f'{total_trips} trips · ₹{total}'
                + (f' (+ bhatta ₹{LabourHyvaTripForm.BHATTA_AMOUNT}/labour)' if bhatta else '')
                + '.',
            )
            redirect_labour = None
            pre_lab_id = request.GET.get('labour_id')
            if pre_lab_id:
                redirect_labour = Labour.objects.filter(
                    pk=pre_lab_id, category='HYVA_DRIVER', is_active=True
                ).first()
            if not redirect_labour and labourers:
                redirect_labour = labourers[0]
            if redirect_labour:
                return redirect('labour:detail', labour_id=redirect_labour.id)
            return redirect('labour:list')

        context = {
            'form': form,
            'page_title': 'Add Hyva Trip',
            'load_options': [
                {'code': code, 'label': label, 'rate': rates[code]}
                for code, label in LabourTripGroup.HYVA_LOAD_CHOICES
                if code
            ],
            'rate_map': rates,
        }
        pre_lab_id = request.GET.get('labour_id')
        if pre_lab_id:
            context['preselected_ids'] = {int(pre_lab_id)}
        if not rows:
            context['row_error'] = 'Kam se kam ek load line mein trips count dalo (load type select karke).'
        return render(request, 'labour/hyva_trip_form.html', context)

    form = LabourHyvaTripForm(initial={'date': timezone.localdate()})
    labour_id = request.GET.get('labour_id')
    preselected_ids = set()
    if labour_id:
        pre = Labour.objects.filter(pk=labour_id, category='HYVA_DRIVER', is_active=True).first()
        if pre:
            form.fields['labourers'].initial = [pre.id]
            preselected_ids = {pre.id}
    return render(request, 'labour/hyva_trip_form.html', {
        'form': form,
        'page_title': 'Add Hyva Trip',
        'load_options': [
            {'code': code, 'label': label, 'rate': rates[code]}
            for code, label in LabourTripGroup.HYVA_LOAD_CHOICES
            if code
        ],
        'rate_map': rates,
        'preselected_ids': preselected_ids,
    })


@login_required(login_url='/login/')
def hyva_trip_edit(request, group_id):
    """Edit a Hyva Driver's full day entry — load lines AND bhatta together.

    Clicking "Edit" on one trip group opens this page which shows all the
    load lines saved for that day (same workers + date), lets the user fix
    trips / load types / labourers, and toggle the daily bhatta (₹200/worker).
    Saves by updating/creating/deleting the day's LabourTripGroup rows and
    creating/removing the matching bhatta LabourExtraPayment(s).
    """
    try:
        group = LabourTripGroup.objects.get(pk=group_id)
    except LabourTripGroup.DoesNotExist:
        messages.info(request, 'Ye entry pehle se change/delete ho chuki hai — page fresh karo.')
        return redirect('labour:list')

    date = group.date
    lab_ids = set(group.labourers.values_list('id', flat=True))
    siblings = [
        g for g in LabourTripGroup.objects.filter(date=date).order_by('id')
        if set(g.labourers.values_list('id', flat=True)) == lab_ids
    ]
    if not siblings:
        siblings = [group]

    load_choices = [c for c, _ in LabourTripGroup.HYVA_LOAD_CHOICES if c]
    load_rates = LabourTripGroup.HYVA_LOAD_RATES
    rates = {code: load_rates[code] for code in load_choices}

    bhatta_count = LabourExtraPayment.objects.filter(
        labour__in=lab_ids, date=date, note='Bhatta'
    ).count()

    if request.method == 'POST':
        form = LabourHyvaTripForm(request.POST)
        rows = []
        i = 0
        while True:
            lt_code = request.POST.get(f'load_type_{i}')
            trip_raw = request.POST.get(f'trip_count_{i}')
            if lt_code is None:
                break
            try:
                trips = int(trip_raw or 0)
            except (TypeError, ValueError):
                trips = 0
            if lt_code in rates and trips and trips > 0:
                rows.append({'load_type': lt_code, 'trip_count': trips})
            i += 1

        if form.is_valid() and rows:
            labourers = list(form.cleaned_data['labourers'])
            bhatta = form.cleaned_data.get('bhatta')
            note = form.cleaned_data.get('note', '')

            # Update / extend the day's groups to match the submitted rows
            for idx, row in enumerate(rows):
                if idx < len(siblings):
                    g = siblings[idx]
                    g.date = date
                    g.trip_count = row['trip_count']
                    g.load_type = row['load_type']
                    g.fill_type = 'HAND'
                    g.note = note
                    g.save()
                    g.labourers.set(labourers)
                else:
                    g = LabourTripGroup.objects.create(
                        date=date,
                        trip_count=row['trip_count'],
                        rate_per_trip=Decimal(load_rates[row['load_type']]),
                        load_type=row['load_type'],
                        fill_type='HAND',
                        total_amount=Decimal(row['trip_count']) * Decimal(load_rates[row['load_type']]),
                        note=note,
                    )
                    g.labourers.set(labourers)

            # Remove extra leftover groups if fewer load lines now
            for g in siblings[len(rows):]:
                g.delete()

            # Toggle bhatta for the day
            labourer_ids = [l.id for l in labourers]
            bhatta_scope = LabourExtraPayment.objects.filter(
                labour__in=lab_ids | set(labourer_ids),
                date=date, note='Bhatta',
            )
            if bhatta:
                existing = set(bhatta_scope.values_list('labour_id', flat=True))
                for l in labourers:
                    if l.id not in existing:
                        LabourExtraPayment.objects.create(
                            labour=l, date=date,
                            amount=Decimal(LabourHyvaTripForm.BHATTA_AMOUNT),
                            note='Bhatta',
                        )
                bhatta_scope.exclude(labour_id__in=labourer_ids).delete()
            else:
                bhatta_scope.delete()

            total_trips = sum(r['trip_count'] for r in rows)
            total = sum(
                Decimal(r['trip_count']) * Decimal(load_rates[r['load_type']])
                for r in rows
            )
            messages.success(
                request,
                f'Hyva entry updated · {len(rows)} load line(s) · '
                f'{total_trips} trips · ₹{total}'
                + (f' (+ bhatta ₹{LabourHyvaTripForm.BHATTA_AMOUNT}/labour)' if bhatta else '')
                + '.',
            )
            target_labour = group.labourers.first() or (labourers[0] if labourers else None)
            return redirect('labour:detail', labour_id=target_labour.id if target_labour else 1)

        context = {
            'form': form,
            'page_title': 'Edit Hyva Trip',
            'editing': True,
            'editing_date': date,
            'existing_rows': rows,
            'lab_ids': lab_ids,
            'bhatta_on': bool(request.POST.get('bhatta')),
            'load_options': [
                {'code': code, 'label': label, 'rate': rates[code]}
                for code, label in LabourTripGroup.HYVA_LOAD_CHOICES
                if code
            ],
            'rate_map': rates,
        }
        if not rows:
            context['row_error'] = 'Kam se kam ek load line mein trips count dalo (load type select karke).'
        return render(request, 'labour/hyva_trip_edit.html', context)

    form = LabourHyvaTripForm(initial={
        'date': date,
        'bhatta': bool(bhatta_count),
    })
    form.fields['labourers'].initial = list(lab_ids)
    return render(request, 'labour/hyva_trip_edit.html', {
        'form': form,
        'page_title': 'Edit Hyva Trip',
        'editing': True,
        'editing_date': date,
        'existing_rows': [
            {'load_type': g.load_type, 'trip_count': g.trip_count, 'load_label': g.load_label}
            for g in siblings
        ],
        'lab_ids': lab_ids,
        'bhatta_on': bool(bhatta_count),
        'load_options': [
            {'code': code, 'label': label, 'rate': rates[code]}
            for code, label in LabourTripGroup.HYVA_LOAD_CHOICES
            if code
        ],
        'rate_map': rates,
    })


# ----------------------------------------------------------------------------
# Trip entry edit / delete / list — fix a wrong entry, shares recalculate
# automatically (total = trips × rate, split across the group).
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
def trip_group_edit(request, group_id):
    try:
        group = LabourTripGroup.objects.get(pk=group_id)
    except LabourTripGroup.DoesNotExist:
        messages.info(request, 'Entry already changed/deleted — page fresh kar ke dekho.')
        return redirect('labour:trips')

    existing_ids = set(group.labourers.values_list('id', flat=True))
    existing_labourers = list(group.labourers.all()[:1])
    category = existing_labourers[0].category if existing_labourers else ''

    if request.method == 'POST':
        form = LabourTripGroupForm(request.POST, instance=group,
                                   category=category, keep_labourers=existing_ids)
        if form.is_valid():
            obj = form.save()
            messages.success(
                request,
                f'Trip entry updated · {obj.trip_count} trips · ₹{obj.total_amount}.',
            )
            return redirect('labour:list')
    else:
        form = LabourTripGroupForm(instance=group,
                                   category=category, keep_labourers=existing_ids)

    return render(request, 'labour/trip_group_form.html', {
        'form': form,
        'page_title': 'Edit Trip Entry',
        'show_add_another': False,
        'editing_group': group,
        'selected_labourer_ids': existing_ids,
    })


@login_required(login_url='/login/')
@require_POST
def trip_group_delete(request, group_id):
    try:
        group = LabourTripGroup.objects.get(pk=group_id)
    except LabourTripGroup.DoesNotExist:
        messages.info(request, 'Ye entry pehle se delete ho chuki hai.')
        return redirect('labour:list')
    first_labour = group.labourers.order_by('id').first()
    messages.warning(request, f"Trip entry deleted · {group.date:%d-%b-%Y} · ₹{group.total_amount} (entry paise ab split nahi honge).")
    group.delete()
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('labour:trips')


@login_required(login_url='/login/')
def trip_group_list(request):
    category = request.GET.get('category', '')
    qs = LabourTripGroup.objects.prefetch_related('labourers').order_by('-date', '-id')
    
    if category:
        qs = qs.filter(labourers__category=category).distinct()
        
    groups = qs[:100]
    years = []
    current_year = None
    current_month = None
    current_day = None
    
    for g in groups:
        labourers = list(g.labourers.order_by('name'))
        row = {
            'group': g,
            'labourers': labourers,
            'per_head': g.per_labour_share if labourers else Decimal('0'),
        }
        
        year_str = str(g.date.year)
        month_str = g.date.strftime('%B')
        
        if not current_year or current_year['name'] != year_str:
            if current_day:
                current_month['days'].append(current_day)
                current_day = None
            if current_month:
                current_year['months'].append(current_month)
                current_month = None
            if current_year:
                years.append(current_year)
                
            current_year = {
                'name': year_str,
                'months': [],
                'total_trips': 0,
                'total_amount': Decimal('0')
            }
            
        if not current_month or current_month['name'] != month_str:
            if current_day:
                current_month['days'].append(current_day)
                current_day = None
            if current_month:
                current_year['months'].append(current_month)
                
            current_month = {
                'name': month_str,
                'days': [],
                'total_trips': 0,
                'total_amount': Decimal('0'),
                'load_breakdown': {},
            }
        
        if not current_day or current_day['date'] != g.date:
            if current_day:
                current_month['days'].append(current_day)
            current_day = {
                'date': g.date,
                'entries': [],
                'total_trips': 0,
                'total_amount': Decimal('0'),
                'load_breakdown': {},
            }
            
        current_day['entries'].append(row)
        current_day['total_trips'] += g.trip_count
        current_day['total_amount'] += g.total_amount
        
        current_month['total_trips'] += g.trip_count
        current_month['total_amount'] += g.total_amount
        
        current_year['total_trips'] += g.trip_count
        current_year['total_amount'] += g.total_amount
        
        if g.load_type:
            label = g.load_label
            current_day['load_breakdown'][label] = current_day['load_breakdown'].get(label, 0) + g.trip_count
            current_month['load_breakdown'][label] = current_month['load_breakdown'].get(label, 0) + g.trip_count
        elif g.fill_type:
            label = g.fill_type
            current_day['load_breakdown'][label] = current_day['load_breakdown'].get(label, 0) + g.trip_count
            current_month['load_breakdown'][label] = current_month['load_breakdown'].get(label, 0) + g.trip_count
        
    if current_day:
        current_month['days'].append(current_day)
    if current_month:
        current_year['months'].append(current_month)
    if current_year:
        years.append(current_year)

    add_url = 'labour:hyva_trip_add' if category == 'HYVA_DRIVER' else 'labour:trip_add'
    add_label = '＋ Add Hyva Trip' if category == 'HYVA_DRIVER' else '＋ Add Trip'
    page_title = 'Hyva Trip Entries' if category == 'HYVA_DRIVER' else ('Tractor Trip Entries' if category == 'TRACTOR' else 'Trip Entries')

    return render(request, 'labour/trip_group_list.html', {
        'years': years,
        'page_title': page_title,
        'category': category,
        'add_url': add_url,
        'add_label': add_label,
    })


# ----------------------------------------------------------------------------
# Extra payment
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
def extra_create(request, labour_id=None):
    labour = None
    if labour_id:
        labour = get_object_or_404(Labour, pk=labour_id)

    if request.method == 'POST':
        form = LabourExtraPaymentForm(request.POST, labour=labour)
        if form.is_valid():
            obj = form.save(commit=False)
            if labour is not None:
                obj.labour = labour
            obj.save()
            messages.success(request, f'Extra payment of ₹{obj.amount} saved.')
            if labour is not None:
                return redirect('labour:detail', labour_id=labour.id)
            return redirect('labour:list')
    else:
        form = LabourExtraPaymentForm(labour=labour)

    return render(request, 'labour/extra_form.html', {
        'form': form,
        'page_title': 'Add Extra Payment',
        'labour': labour,
    })

@login_required(login_url='/login/')
@require_POST
def extra_delete(request, extra_id):
    extra = get_object_or_404(LabourExtraPayment, pk=extra_id)
    labour_id = extra.labour.id
    amt = extra.amount
    extra.delete()
    messages.success(request, f'Extra payment of ₹{amt} deleted.')
    return redirect('labour:detail', labour_id=labour_id)

@login_required(login_url='/login/')
def extra_edit(request, extra_id):
    extra = get_object_or_404(LabourExtraPayment, pk=extra_id)
    if request.method == 'POST':
        form = LabourExtraPaymentForm(request.POST, instance=extra, labour=extra.labour)
        if form.is_valid():
            form.save()
            messages.success(request, f'Extra payment updated to ₹{extra.amount}.')
            return redirect('labour:detail', labour_id=extra.labour.id)
    else:
        form = LabourExtraPaymentForm(instance=extra, labour=extra.labour)

    return render(request, 'labour/extra_form.html', {
        'form': form,
        'page_title': 'Edit Extra Payment',
        'labour': extra.labour,
    })


# ----------------------------------------------------------------------------
# Advance
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
def advance_create(request, labour_id=None):
    """Single-labour advance form."""
    labour = None
    if labour_id:
        labour = get_object_or_404(Labour, pk=labour_id)

    if request.method == 'POST':
        form = LabourAdvanceForm(request.POST, labour=labour)
        if form.is_valid():
            obj = form.save(commit=False)
            if labour is not None:
                obj.labour = labour
            try:
                obj.save()
            except Exception as e:  # unique constraint
                messages.error(request, f'Already an advance for this labour on that date.')
                return render(request, 'labour/advance_form.html', {
                    'form': form, 'page_title': 'Add Advance', 'labour': labour,
                })
            messages.success(request, f'Advance ₹{obj.amount} saved.')
            if labour is not None:
                return redirect('labour:detail', labour_id=labour.id)
            return redirect('labour:list')
    else:
        form = LabourAdvanceForm(labour=labour)

    return render(request, 'labour/advance_form.html', {
        'form': form,
        'page_title': 'Add Advance',
        'labour': labour,
    })

@login_required(login_url='/login/')
@require_POST
def advance_delete(request, advance_id):
    advance = get_object_or_404(LabourAdvance, pk=advance_id)
    labour_id = advance.labour.id
    amt = advance.amount
    advance.delete()
    messages.success(request, f'Advance of ₹{amt} deleted.')
    return redirect('labour:detail', labour_id=labour_id)

@login_required(login_url='/login/')
def advance_edit(request, advance_id):
    adv = get_object_or_404(LabourAdvance, pk=advance_id)
    if request.method == 'POST':
        form = LabourAdvanceForm(request.POST, instance=adv, labour=adv.labour)
        if form.is_valid():
            form.save()
            messages.success(request, f'Advance updated to ₹{adv.amount}.')
            return redirect('labour:detail', labour_id=adv.labour.id)
    else:
        form = LabourAdvanceForm(instance=adv, labour=adv.labour)

    return render(request, 'labour/advance_form.html', {
        'form': form,
        'page_title': 'Edit Advance',
        'labour': adv.labour,
    })


@login_required(login_url='/login/')
def rozi_create(request, labour_id=None):
    """
    Add daily rozi (wages) for a labourer. Amount is auto-computed from the
    labourer's base_daily_rate based on Full / One & Half / Half day.
    """
    labour = None
    if labour_id is not None:
        labour = get_object_or_404(Labour, pk=labour_id)

    if request.method == 'POST':
        form = LabourRoziForm(request.POST, labour=labour)
        if form.is_valid():
            obj = form.save(commit=False)
            if labour is not None:
                obj.labour = labour
            obj.save()
            messages.success(
                request,
                f"Rozi ₹{obj.amount} saved for {obj.labour.name} "
                f"({obj.get_day_type_display()}).",
            )
            if labour is not None:
                return redirect('labour:detail', labour_id=labour.id)
            return redirect('labour:list')
    else:
        form = LabourRoziForm(labour=labour)

    return render(request, 'labour/rozi_form.html', {
        'form': form,
        'page_title': 'Add Rozi',
        'labour': labour,
    })

@login_required(login_url='/login/')
@require_POST
def rozi_delete(request, rozi_id):
    rozi = get_object_or_404(LabourRozi, pk=rozi_id)
    labour_id = rozi.labour.id
    amt = rozi.amount
    rozi.delete()
    messages.success(request, f'Rozi entry of ₹{amt} deleted.')
    return redirect('labour:detail', labour_id=labour_id)

@login_required(login_url='/login/')
def rozi_edit(request, rozi_id):
    rozi = get_object_or_404(LabourRozi, pk=rozi_id)
    if request.method == 'POST':
        form = LabourRoziForm(request.POST, instance=rozi, labour=rozi.labour)
        if form.is_valid():
            form.save()
            messages.success(request, f'Rozi entry updated to ₹{rozi.amount}.')
            return redirect('labour:detail', labour_id=rozi.labour.id)
    else:
        form = LabourRoziForm(instance=rozi, labour=rozi.labour)

    return render(request, 'labour/rozi_form.html', {
        'form': form,
        'page_title': 'Edit Rozi',
        'labour': rozi.labour,
    })


@login_required(login_url='/login/')
def rozi_multi(request):
    """
    Quick rozi entry: pick a date, then choose Full / One&Half / Half for
    each labourer. Only those with a day_type selected are saved.

    Rozi only applies to Mistri labour (Mistri + Mistri Helper). Default
    shows Mistri category unless an explicit ?category= is passed.
    """
    today = timezone.localdate()
    selected_date = _parse_date(request.POST.get('date') or request.GET.get('date'), today)

    category_filter = request.POST.get('category') or request.GET.get('category') or 'MISTRI'
    if category_filter:
        labours = list(Labour.objects.filter(
            is_active=True, category=category_filter
        ).exclude(is_vendor=True).order_by('name'))
    else:
        labours = list(Labour.objects.filter(is_active=True).exclude(is_vendor=True).order_by('name'))

    existing = {
        rz.labour_id: rz
        for rz in LabourRozi.objects.filter(date=selected_date)
    }

    reset = request.GET.get('reset') == '1'
    labour_rows = [
        (l,
         None if reset else (existing[l.id].day_type if l.id in existing else None),
         l.id in existing)
        for l in labours
    ]

    if request.method == 'POST':
        saved = 0
        for l in labours:
            day_type = request.POST.get(f'day_type_{l.id}', '').strip()
            if not day_type or day_type not in LabourRozi.DAY_MULTIPLIER:
                continue
            if l.category == 'MISTRI' and l.sub_category == 'MISTRI':
                amount = LabourRozi.MISTRI_RATES[day_type]
            else:
                amount = l.base_daily_rate * LabourRozi.DAY_MULTIPLIER[day_type]
            note = request.POST.get(f'note_{l.id}', '').strip()
            if l.id in existing:
                rz = existing[l.id]
                rz.day_type = day_type
                rz.amount = amount
                rz.note = note
                rz.save()
            else:
                LabourRozi.objects.create(
                    labour=l, date=selected_date, day_type=day_type, note=note,
                )
            saved += 1
        if saved:
            messages.success(request, f'{saved} rozi entry(ies) saved for {selected_date}.')
        else:
            messages.info(request, 'No day type selected.')
        base = reverse('labour:list')
        qs = []
        if category_filter:
            qs.append(f'category={category_filter}')
        url = f"{base}?{'&'.join(qs)}" if qs else base
        return redirect(url)

    context = {
        'page_title': 'Quick Rozi',
        'selected_date': selected_date,
        'labours': labours,
        'existing': existing,
        'labour_rows': labour_rows,
        'DAY_TYPE_CHOICES': LabourRozi.DAY_TYPE_CHOICES,
        'DAY_MULTIPLIER': LabourRozi.DAY_MULTIPLIER,
        'MISTRI_RATES': LabourRozi.MISTRI_RATES,
        'category_filter': category_filter,
    }
    return render(request, 'labour/rozi_multi.html', context)


@login_required(login_url='/login/')
def advance_multi(request):
    """
    Day-view quick entry: pick a date, then enter amount for whichever
    labourers took advance. Only those with a value > 0 are saved.
    """
    today = timezone.localdate()
    selected_date = _parse_date(request.POST.get('date') or request.GET.get('date'), today)

    category_filter = request.POST.get('category') or request.GET.get('category') or ''
    if category_filter:
        labours = list(Labour.objects.filter(
            is_active=True, category=category_filter
        ).exclude(is_vendor=True).order_by('name'))
    else:
        labours = list(Labour.objects.filter(is_active=True).exclude(is_vendor=True).order_by('name'))
    # Existing advances for selected_date, to pre-fill
    existing = {
        a.labour_id: a
        for a in LabourAdvance.objects.filter(date=selected_date)
    }
    # After a save, clear the amount inputs (reset) so it doesn't feel stuck,
    # but keep the ✓ marker for labourers who already have an advance that day.
    reset = request.GET.get('reset') == '1'
    # Flatten to (labour, existing_amount, has_existing) tuples for easy template iteration
    labour_rows = [
        (l,
         None if reset else (existing[l.id].amount if l.id in existing else None),
         l.id in existing)
        for l in labours
    ]

    if request.method == 'POST':
        saved = 0
        for l in labours:
            raw = request.POST.get(f'amount_{l.id}', '').strip()
            if not raw:
                continue
            try:
                amount = Decimal(raw)
            except Exception:
                continue
            if amount <= 0:
                continue
            # Upsert: only one advance per labour per day
            obj, created = LabourAdvance.objects.update_or_create(
                labour=l, date=selected_date,
                defaults={'amount': amount, 'note': request.POST.get(f'note_{l.id}', '').strip()},
            )
            saved += 1
        if not saved:
            messages.info(request, 'No advance amounts entered.')
        # After saving, jump back to the current date so the old-date screen
        # doesn't linger blank while entries already exist for it.
        qs = [f'date={today.isoformat()}']
        if category_filter:
            qs.append(f'category={category_filter}')
        if saved:
            qs.append(f'saved={saved}')
        return redirect(f"{request.path}?{'&'.join(qs)}")

    context = {
        'page_title': 'Quick Advances',
        'selected_date': selected_date,
        'labours': labours,
        'existing': existing,
        'labour_rows': labour_rows,
        'category_filter': category_filter,
    }
    return render(request, 'labour/advance_multi.html', context)


# ----------------------------------------------------------------------------
# Driver payment
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
def driver_payment_create(request, labour_id=None):
    labour = None
    if labour_id:
        labour = get_object_or_404(Labour, pk=labour_id)
        if not labour.is_driver:
            messages.warning(request, f'{labour.name} is not marked as driver.')

    if request.method == 'POST':
        form = LabourDriverPaymentForm(request.POST, labour=labour)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Driver payment ₹{obj.amount} saved for {obj.labour.name}.')
            if labour is not None:
                return redirect('labour:detail', labour_id=labour.id)
            return redirect('labour:list')
    else:
        form = LabourDriverPaymentForm(labour=labour)

    return render(request, 'labour/driver_payment_form.html', {
        'form': form,
        'page_title': 'Add Driver Payment',
        'labour': labour,
    })


# ----------------------------------------------------------------------------
# Settlement
# ----------------------------------------------------------------------------

@login_required(login_url='/login/')
@require_POST
def labour_set_outstanding(request, labour_id):
    """Directly set the running old balance (outstanding) for a labour.
    Positive amount = labour owes owner. Use for prior dues / manual adjust."""
    labour = get_object_or_404(Labour, pk=labour_id)
    ob = _ensure_old_balance(labour)
    raw = request.POST.get('outstanding_amount', '').strip()
    try:
        new_amount = Decimal(raw or '0')
    except Exception:
        messages.error(request, 'Please enter a valid amount.')
        return redirect('labour:detail', labour_id=labour.id)
    if new_amount < 0:
        messages.error(request, 'Amount cannot be negative. Use 0 to clear dues.')
        return redirect('labour:detail', labour_id=labour.id)
    ob.amount = new_amount
    ob.save()
    messages.success(request, f'Outstanding for {labour.name} set to ₹{new_amount:.2f}.')
    return redirect('labour:detail', labour_id=labour.id)


@login_required(login_url='/login/')
@transaction.atomic
def settlement_create(request, labour_id):
    labour = get_object_or_404(Labour, pk=labour_id)
    ob = _ensure_old_balance(labour)
    today = timezone.localdate()

    if request.method == 'POST':
        form = LabourSettlementForm(request.POST)
        if form.is_valid():
            settlement = form.save(commit=False)
            settlement.labour = labour
            settlement.old_balance_before = ob.amount
            settlement.recalculate()
            settlement.save()
            # Update labour's old balance
            ob.amount = settlement.final_old_balance
            ob.save()
            messages.success(
                request,
                f'Settlement saved for {labour.name}. New old balance: ₹{ob.amount}.',
            )
            return redirect('labour:detail', labour_id=labour.id)
    else:
        form = LabourSettlementForm()

    # GET: compute live preview based on GET params or default period
    period_start = _parse_date(
        request.GET.get('period_start') or request.POST.get('period_start'),
        today.replace(day=1),
    )
    period_end = _parse_date(
        request.GET.get('period_end') or request.POST.get('period_end'),
        today,
    )

    # Recompute a preview object (not saved)
    preview = LabourSettlement(
        labour=labour,
        period_start=period_start,
        period_end=period_end,
        old_balance_before=ob.amount,
    )
    preview.recalculate()

    context = {
        'form': form,
        'page_title': 'Settle Payment',
        'labour': labour,
        'old_balance': ob,
        'preview': preview,
    }
    return render(request, 'labour/settlement_form.html', context)


# ----------------------------------------------------------------------------
# Settlement edit / revert — sirf sabse recent settlement pe, taaki old-balance
# chain sahi rahe. Revert: record delete + old balance (before) restore.
# ----------------------------------------------------------------------------

def _latest_settlement(labour):
    return LabourSettlement.objects.filter(labour=labour).order_by(
        '-settlement_date', '-id'
    ).first()


@login_required(login_url='/login/')
@transaction.atomic
def settlement_edit(request, settlement_id):
    settlement = get_object_or_404(LabourSettlement, pk=settlement_id)
    latest = _latest_settlement(settlement.labour)
    if latest is None or settlement.pk != latest.pk:
        messages.info(request, 'Sirf sabse recent settlement edit ho sakti hai — pahle baad waali revert karo.')
        return redirect('labour:detail', labour_id=settlement.labour_id)

    if request.method == 'POST':
        form = LabourSettlementForm(request.POST, instance=settlement)
        if form.is_valid():
            s = form.save(commit=False)
            # Period + old_balance_before locked — sirf amount/date/note badlo
            s.period_start = settlement.period_start
            s.period_end = settlement.period_end
            s.old_balance_before = settlement.old_balance_before
            s.recalculate()
            s.save()
            ob = _ensure_old_balance(settlement.labour)
            ob.amount = s.final_old_balance
            ob.save()
            messages.success(
                request,
                f'Settlement updated for {settlement.labour.name}. New old balance: ₹{ob.amount}.',
            )
            return redirect('labour:detail', labour_id=settlement.labour_id)
    else:
        form = LabourSettlementForm(instance=settlement)

    return render(request, 'labour/settlement_form.html', {
        'form': form,
        'page_title': 'Edit Settlement',
        'labour': settlement.labour,
        'old_balance': {'amount': settlement.old_balance_before},
        'preview': settlement,
        'editing': True,
        'settlement': settlement,
    })


@login_required(login_url='/login/')
@require_POST
def settlement_revert(request, settlement_id):
    settlement = get_object_or_404(LabourSettlement, pk=settlement_id)
    latest = _latest_settlement(settlement.labour)
    if latest is None or settlement.pk != latest.pk:
        messages.info(request, 'Sirf sabse recent settlement revert ho sakti hai.')
        return redirect('labour:detail', labour_id=settlement.labour_id)

    ob = _ensure_old_balance(settlement.labour)
    ob.amount = settlement.old_balance_before
    ob.save()
    messages.warning(
        request,
        f'Settlement {settlement.settlement_date:%d-%b-%Y} revert kar diya. '
        f'Old balance restore: ₹{ob.amount}.',
    )
    settlement.delete()
    return redirect('labour:detail', labour_id=settlement.labour_id)


# ----------------------------------------------------------------------------
# Labour Book — saare labour ke period statements ek hi PDF/Excel me
# ----------------------------------------------------------------------------

def _labour_statement_for_period(labour, period_start, period_end):
    """Full statement data for one labour in a period (mirrors labour_detail)."""
    ob, _ = LabourOldBalance.objects.get_or_create(
        labour=labour, defaults={'amount': Decimal('0')}
    )

    rows = _day_aggregates_for_labour(labour, period_start, period_end)

    trip_total = sum((r['trips_amount'] for r in rows), Decimal('0'))
    extra_total = sum((r['extra_amount'] for r in rows), Decimal('0'))
    advance_total = sum((r['advance_amount'] for r in rows), Decimal('0'))

    driver_qs = LabourDriverPayment.objects.filter(labour=labour).filter(
        period_start__lte=period_end, period_end__gte=period_start
    )
    driver_total = driver_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    trip_groups = list(
        LabourTripGroup.objects.filter(
            labourers=labour, date__gte=period_start, date__lte=period_end
        ).prefetch_related('labourers').order_by('-date', '-id')
    )

    total_salary = trip_total + extra_total + driver_total
    payment = total_salary - advance_total
    final_amount = payment - ob.amount

    settlements = list(LabourSettlement.objects.filter(
        labour=labour,
        settlement_date__gte=period_start,
        settlement_date__lte=period_end,
    ))

    return {
        'labour': labour,
        'old_balance': ob.amount,
        'rows': rows,
        'trip_total': trip_total,
        'extra_total': extra_total,
        'driver_total': driver_total,
        'advance_total': advance_total,
        'total_salary': total_salary,
        'payment': payment,
        'final_amount': final_amount,
        'settlements': settlements,
        'driver_payments': list(driver_qs),
        'trip_groups': trip_groups,
    }


@login_required(login_url='/login/')
def labour_book(request):
    today = timezone.localdate()
    period_start = _parse_date(request.GET.get('from_date'), today.replace(day=1))
    period_end = _parse_date(request.GET.get('to_date'), today)
    if period_start and period_end and period_start > period_end:
        period_start, period_end = period_end, period_start

    export = request.GET.get('export')

    labours = list(Labour.objects.filter(is_active=True).exclude(is_vendor=True).order_by('name'))
    statements = [
        _labour_statement_for_period(labour, period_start, period_end)
        for labour in labours
    ]

    if export == 'excel':
        return _labour_book_excel(statements, period_start, period_end)

    return _labour_book_pdf(statements, period_start, period_end)


# ----------------------------------------------------------------------------
# Labour Summary — category-wise (Tractor / Hyva / Mistri / JCB) earnings,
# advances aur over-view for a date range. Excel + PDF + web page.
# ----------------------------------------------------------------------------

def _labour_summary_data(period_start, period_end):
    """Aggregate summary data per category and per labour in the period."""
    period_start = period_start or timezone.localdate().replace(day=1)
    period_end = period_end or timezone.localdate()

    labours = list(Labour.objects.filter(is_active=True).exclude(is_vendor=True).order_by('name'))

    # Aggregate per labour using the same statement builder as detail/book
    per_labour = []
    cat_totals = {}
    for l in labours:
        st = _labour_statement_for_period(l, period_start, period_end)
        item = {
            'labour': l,
            'category': l.category,
            'trip_total': st['trip_total'],
            'extra_total': st['extra_total'],
            'driver_total': st['driver_total'],
            'advance_total': st['advance_total'],
            'total_salary': st['total_salary'],
            'final_amount': st['final_amount'],
        }
        per_labour.append(item)
        cat = cat_totals.setdefault(l.category, {
            'workers': 0,
            'trip_total': Decimal('0'),
            'extra_total': Decimal('0'),
            'driver_total': Decimal('0'),
            'advance_total': Decimal('0'),
            'total_salary': Decimal('0'),
            'final_amount': Decimal('0'),
        })
        cat['workers'] += 1
        cat['trip_total'] += st['trip_total']
        cat['extra_total'] += st['extra_total']
        cat['driver_total'] += st['driver_total']
        cat['advance_total'] += st['advance_total']
        cat['total_salary'] += st['total_salary']
        cat['final_amount'] += st['final_amount']

    # Order categories: Tractor, Hyva Driver, JCB Operator, Mistri, then others
    order = ['TRACTOR', 'HYVA_DRIVER', 'JCB_OPERATOR', 'MISTRI']
    cat_rows = []
    for code in order:
        if code in cat_totals:
            cat_rows.append({'code': code, 'label': dict(Labour.CATEGORY_CHOICES).get(code, code), **cat_totals[code]})
    for code in sorted(k for k in cat_totals if k not in order):
        cat_rows.append({'code': code, 'label': code, **cat_totals[code]})

    grand = {
        'workers': sum(c['workers'] for c in cat_rows),
        'trip_total': sum(c['trip_total'] for c in cat_rows),
        'extra_total': sum(c['extra_total'] for c in cat_rows),
        'driver_total': sum(c['driver_total'] for c in cat_rows),
        'advance_total': sum(c['advance_total'] for c in cat_rows),
        'total_salary': sum(c['total_salary'] for c in cat_rows),
        'final_amount': sum(c['final_amount'] for c in cat_rows),
    }

    return {
        'period_start': period_start,
        'period_end': period_end,
        'per_labour': per_labour,
        'cat_rows': cat_rows,
        'grand': grand,
        'period_label': _period_label(period_start, period_end),
    }


@login_required(login_url='/login/')
def labour_summary(request):
    today = timezone.localdate()
    period_start, period_end = _resolve_period(request, today)

    data = _labour_summary_data(period_start, period_end)

    export = request.GET.get('export')
    if export == 'excel':
        return _labour_summary_excel(data)
    if export == 'pdf':
        return _labour_summary_pdf(data)

    category_label = request.GET.get('category', '')

    return render(request, 'labour/labour_summary.html', {
        'data': data,
        'from_date': period_start,
        'to_date': period_end,
        'page_title': 'Labour Summary',
        'months': ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                   'August', 'September', 'October', 'November', 'December'],
    })


def _labour_summary_excel(data):
    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _A

    workbook = Workbook()
    ws = workbook.active
    ws.title = 'Summary'

    header_fill = _PF(start_color='16665A', end_color='16665A', fill_type='solid')
    header_font = _F(color='FFFFFF', bold=True)
    total_fill = _PF(start_color='EEF7F5', end_color='EEF7F5', fill_type='solid')

    period_label = data['period_label']

    # Title
    ws.cell(row=1, column=1, value='Labour Summary').font = _F(bold=True, size=14)
    ws.cell(row=2, column=1, value=period_label).font = _F(size=10, color='64748B')
    ws.cell(row=3, column=1, value='Generated via Business Insights System').font = _F(size=9, color='94A3B8')

    # ---- Category summary table ----
    headers = ['Category', 'Workers', 'Trip ₹', 'Extra ₹', 'Driver ₹', 'Advance ₹', 'Total Earned ₹', 'Final ₹']
    hr = 5
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = _A(horizontal='center', vertical='center')

    r = hr + 1
    for cat in data['cat_rows']:
        ws.cell(row=r, column=1, value=cat['label'])
        ws.cell(row=r, column=2, value=cat['workers'])
        for col, key in [(3, 'trip_total'), (4, 'extra_total'), (5, 'driver_total'),
                         (6, 'advance_total'), (7, 'total_salary'), (8, 'final_amount')]:
            cell = ws.cell(row=r, column=col, value=float(cat[key]))
            cell.number_format = '#,##0.00'
        r += 1

    # Grand total
    g = data['grand']
    ws.cell(row=r, column=1, value='TOTAL')
    ws.cell(row=r, column=2, value=g['workers'])
    for col, key in [(3, 'trip_total'), (4, 'extra_total'), (5, 'driver_total'),
                     (6, 'advance_total'), (7, 'total_salary'), (8, 'final_amount')]:
        cell = ws.cell(row=r, column=col, value=float(g[key]))
        cell.number_format = '#,##0.00'
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = total_fill
        ws.cell(row=r, column=col).font = _F(bold=True)
    r += 2

    # ---- Per-labour detail per category ----
    for cat in data['cat_rows']:
        cat_people = [p for p in data['per_labour'] if p['category'] == cat['code']]
        if not cat_people:
            continue
        ws.cell(row=r, column=1, value=f"{cat['label']} — Labour Detail").font = _F(bold=True, size=12)
        r += 1
        for col, h in enumerate(['Labour', 'Trip ₹', 'Extra ₹', 'Driver ₹', 'Advance ₹', 'Total Earned ₹', 'Final ₹'], start=1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        r += 1
        for p in cat_people:
            ws.cell(row=r, column=1, value=p['labour'].name)
            for col, key in [(2, 'trip_total'), (3, 'extra_total'), (4, 'driver_total'),
                             (5, 'advance_total'), (6, 'total_salary'), (7, 'final_amount')]:
                cell = ws.cell(row=r, column=col, value=float(p[key]))
                cell.number_format = '#,##0.00'
            r += 1
        r += 1

    for col, width in [(1, 26), (2, 10), (3, 12), (4, 12), (5, 12), (6, 12), (7, 15), (8, 12)]:
        ws.column_dimensions[chr(64 + col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="labour_summary.xlsx"'
    workbook.save(response)
    return response


def _labour_summary_pdf(data):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    from core.pdf_utils import (
        get_registered_font,
        build_pdf_header_elements,
        get_indian_current_time_str,
        build_summary_cards,
        apply_data_table_style,
        finish_document,
        build_thankyou_note,
        get_pdf_styles,
        BRAND,
        BRAND_DARK,
    )
    font_name = get_registered_font()

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=10 * mm, leftMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = get_pdf_styles(font_name)

    def _wrapped_style(name, **kw):
        return ParagraphStyle(name, parent=getSampleStyleSheet()['Normal'],
                              fontName=font_name, **kw)

    elements = build_pdf_header_elements(
        font_name=font_name,
        report_title="Labour Summary",
        report_subtitle=(
            f"{data['period_label']} · "
            f"Generated on {get_indian_current_time_str()}"
        ),
    )

    # ---- KPI cards ----
    cards = [
        {'label': 'Workers', 'value': data['grand']['workers']},
        {'label': 'Total Earned', 'value': f"₹{data['grand']['total_salary']:,.0f}"},
        {'label': 'Advance', 'value': f"₹{data['grand']['advance_total']:,.0f}"},
        {'label': 'Final', 'value': f"₹{data['grand']['final_amount']:,.0f}"},
    ]
    elements.append(build_summary_cards(cards, font_name=font_name))
    elements.append(Spacer(1, 6 * mm))

    # ---- Category table ----
    cell_lbl = _wrapped_style('CellLbl', fontSize=8, leading=10, textColor=colors.HexColor('#0f172a'))
    cell_r = _wrapped_style('CellR', fontSize=8, leading=10, textColor=colors.HexColor('#0f172a'), alignment=2)
    head_cell = _wrapped_style('HeadCell', fontSize=8, leading=10, textColor=colors.white)

    cat_headers = ['Category', 'Workers', 'Trip ₹', 'Extra ₹', 'Driver ₹', 'Advance ₹', 'Earned ₹', 'Final ₹']
    tdata = [[Paragraph(h, head_cell) for h in cat_headers]]
    for cat in data['cat_rows']:
        tdata.append([
            Paragraph(f"<b>{cat['label']}</b>", cell_lbl),
            Paragraph(str(cat['workers']), cell_r),
            Paragraph(f"{cat['trip_total']:,.0f}", cell_r),
            Paragraph(f"{cat['extra_total']:,.0f}", cell_r),
            Paragraph(f"{cat['driver_total']:,.0f}", cell_r),
            Paragraph(f"{cat['advance_total']:,.0f}", cell_r),
            Paragraph(f"{cat['total_salary']:,.0f}", cell_r),
            Paragraph(f"{cat['final_amount']:,.0f}", cell_r),
        ])
    g = data['grand']
    tdata.append([
        Paragraph('<b>TOTAL</b>', cell_lbl),
        Paragraph(str(g['workers']), cell_r),
        Paragraph(f"<b>{g['trip_total']:,.0f}</b>", cell_r),
        Paragraph(f"<b>{g['extra_total']:,.0f}</b>", cell_r),
        Paragraph(f"<b>{g['driver_total']:,.0f}</b>", cell_r),
        Paragraph(f"<b>{g['advance_total']:,.0f}</b>", cell_r),
        Paragraph(f"<b>{g['total_salary']:,.0f}</b>", cell_r),
        Paragraph(f"<b>{g['final_amount']:,.0f}</b>", cell_r),
    ])
    cat_table = Table(tdata, colWidths=[32 * mm, 20 * mm, 23 * mm, 23 * mm, 23 * mm, 23 * mm, 25 * mm, 23 * mm])
    apply_data_table_style(cat_table, total_row=True)
    elements.append(cat_table)
    elements.append(Spacer(1, 8 * mm))

    # ---- Per-labour detail per category ----
    for cat in data['cat_rows']:
        cat_people = [p for p in data['per_labour'] if p['category'] == cat['code']]
        if not cat_people:
            continue
        sec_title = _wrapped_style('SecTitle', fontSize=10.5, leading=14, textColor=BRAND_DARK)
        elements.append(Paragraph(f"{cat['label']} — Labour Detail", sec_title))
        elements.append(Spacer(1, 2 * mm))
        p_headers = ['Labour', 'Trip ₹', 'Extra ₹', 'Driver ₹', 'Advance ₹', 'Earned ₹', 'Final ₹']
        pdata = [[Paragraph(h, head_cell) for h in p_headers]]
        for p in cat_people:
            pdata.append([
                Paragraph(p['labour'].name, cell_lbl),
                Paragraph(f"{p['trip_total']:,.0f}", cell_r),
                Paragraph(f"{p['extra_total']:,.0f}", cell_r),
                Paragraph(f"{p['driver_total']:,.0f}", cell_r),
                Paragraph(f"{p['advance_total']:,.0f}", cell_r),
                Paragraph(f"{p['total_salary']:,.0f}", cell_r),
                Paragraph(f"{p['final_amount']:,.0f}", cell_r),
            ])
        p_table = Table(pdata, colWidths=[42 * mm, 24 * mm, 24 * mm, 24 * mm, 25 * mm, 26 * mm, 24 * mm])
        apply_data_table_style(p_table)
        elements.append(KeepTogether([p_table]))
        elements.append(Spacer(1, 5 * mm))

    elements += build_thankyou_note(font_name=font_name)

    finish_document(document, elements, font_name=font_name)
    response = HttpResponse(content_type='application/pdf')
    name = re.sub(r'[^A-Za-z0-9]+', '_', data['period_label']).strip('_')
    response['Content-Disposition'] = f'attachment; filename="labour_summary_{name}.pdf"'
    response.write(buffer.getvalue())
    return response


# ----------------------------------------------------------------------------
# Daily Activity — ek date range (ya month) ki saari daily transactions ek
# jagah: trips (per load type), extras, advances, driver payments aur
# settlements. Page pe summary + Excel/CSV export.
# ----------------------------------------------------------------------------

def _daily_activity_data(period_start, period_end):
    """Aggregate all daily transactions in [period_start, period_end]."""
    if period_start is None and period_end is None:
        trips = list(LabourTripGroup.objects.all())
        extras = list(LabourExtraPayment.objects.all())
        rozis = list(LabourRozi.objects.all())
        advances = list(LabourAdvance.objects.all())
        driver_pmts = list(LabourDriverPayment.objects.all())
        settlements = list(LabourSettlement.objects.all())
    else:
        trips = list(LabourTripGroup.objects.filter(date__range=(period_start, period_end)))
        extras = list(LabourExtraPayment.objects.filter(date__range=(period_start, period_end)))
        rozis = list(LabourRozi.objects.filter(date__range=(period_start, period_end)))
        advances = list(LabourAdvance.objects.filter(date__range=(period_start, period_end)))
        driver_pmts = list(
            LabourDriverPayment.objects.filter(
                period_start__lte=period_end, period_end__gte=period_start
            )
        )
        settlements = list(
            LabourSettlement.objects.filter(settlement_date__range=(period_start, period_end))
        )
    
    # Merge rozis into extras for UI/Export since they act exactly like extra payments (Bhatta/Rozi)
    for rz in rozis:
        # Duck-typing to match LabourExtraPayment expectations in UI
        if not hasattr(rz, 'note'):
            rz.note = rz.get_day_type_display()
        elif rz.note:
            rz.note = f"{rz.get_day_type_display()} - {rz.note}"
        else:
            rz.note = rz.get_day_type_display()
        extras.append(rz)

    # Trips grouped by load type (with totals)
    trip_by_type = defaultdict(lambda: {'count': 0, 'amount': Decimal('0')})
    trip_total = Decimal('0')
    for g in trips:
        key = g.load_label or 'Tractor Trip'
        trip_by_type[key]['count'] += g.trip_count
        trip_by_type[key]['amount'] += g.total_amount
        trip_total += g.total_amount

    extra_total = sum((e.amount for e in extras), Decimal('0'))
    advance_total = sum((a.amount for a in advances), Decimal('0'))
    driver_total = sum((d.amount for d in driver_pmts), Decimal('0'))
    settle_cash = sum((s.cash_paid for s in settlements), Decimal('0'))
    settle_deduct = sum((s.old_balance_deducted for s in settlements), Decimal('0'))

    # Per-day rollup across all types (except driver payments which span periods)
    day_rollup = defaultdict(lambda: {'trips': Decimal('0'), 'extra': Decimal('0'),
                                      'advance': Decimal('0')})
    for g in trips:
        day_rollup[g.date]['trips'] += g.total_amount
    for e in extras:
        day_rollup[e.date]['extra'] += e.amount
    for a in advances:
        day_rollup[a.date]['advance'] += a.amount

    days = sorted(day_rollup.keys())
    grand_trips = sum((d['trips'] for d in day_rollup.values()), Decimal('0'))
    grand_extra = sum((d['extra'] for d in day_rollup.values()), Decimal('0'))
    grand_advance = sum((d['advance'] for d in day_rollup.values()), Decimal('0'))

    # Per-day grouped detail entries (for drill-down, avoids dumping everything)
    trips_by_day = defaultdict(lambda: defaultdict(lambda: {'count': 0, 'amount': Decimal('0')}))
    for g in trips:
        lbl = g.load_label or g.fill_type or 'Tractor Trip'
        trips_by_day[g.date][lbl]['count'] += g.trip_count
        trips_by_day[g.date][lbl]['amount'] += g.total_amount

    extras_by_day = defaultdict(list)
    for e in extras:
        extras_by_day[e.date].append(e)
    advances_by_day = defaultdict(list)
    for a in advances:
        advances_by_day[a.date].append(a)
    settles_by_day = defaultdict(list)
    for s in settlements:
        settles_by_day[s.settlement_date].append(s)

    day_list = [
        {
            'date': d,
            'trips': day_rollup[d]['trips'],
            'extra': day_rollup[d]['extra'],
            'advance': day_rollup[d]['advance'],
            'day_trips': [
                {'label': k, 'count': v['count'], 'amount': v['amount']}
                for k, v in trips_by_day.get(d, {}).items()
            ],
            'day_extras': extras_by_day.get(d, []),
            'day_advances': advances_by_day.get(d, []),
            'day_settlements': settles_by_day.get(d, []),
        }
        for d in days
    ]

    return {
        'period_start': period_start,
        'period_end': period_end,
        'trips': trips,
        'extras': extras,
        'advances': advances,
        'driver_pmts': driver_pmts,
        'settlements': settlements,
        'trip_by_type': dict(trip_by_type),
        'trip_count_total': sum(t['count'] for t in trip_by_type.values()),
        'trip_total': trip_total,
        'extra_total': extra_total,
        'advance_total': advance_total,
        'driver_total': driver_total,
        'settle_cash': settle_cash,
        'settle_deduct': settle_deduct,
        'days': days,
        'day_list': day_list,
        'day_rollup': dict(day_rollup),
        'grand_trips': grand_trips,
        'grand_extra': grand_extra,
        'grand_advance': grand_advance,
        'period_label': (
            _period_label(period_start, period_end)
        ),
    }


def _period_label(start, end):
    if start and end:
        if start == end:
            return f"{start:%d-%b-%Y}"
        return f"{start:%d-%b-%Y} to {end:%d-%b-%Y}"
    return 'All Transactions'


def _daily_activity_excel(data, filename='daily_activity.xlsx'):
    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _A
    from openpyxl.utils import get_column_letter as _gcl

    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Activity'

    header_fill = _PF(start_color='16665A', end_color='16665A', fill_type='solid')
    header_font = _F(color='FFFFFF', bold=True)
    bold = _F(bold=True)
    total_fill = _PF(start_color='EEF7F5', end_color='EEF7F5', fill_type='solid')

    ws.cell(row=1, column=1, value='DAILY ACTIVITY').font = _F(bold=True, size=13)
    ws.cell(row=2, column=1, value='Period').font = bold
    ws.cell(row=2, column=2, value=data['period_label'])

    # --- Summary block ---
    ws.cell(row=4, column=1, value='SUMMARY').font = _F(bold=True, color='2563EB')
    summary = [
        ('Trip Trips Count', data['trip_count_total']),
        ('Trip Amount', data['trip_total']),
        ('Extra Amount', data['extra_total']),
        ('Advance Amount', data['advance_total']),
        ('Driver Payment', data['driver_total']),
        ('Settlement Cash Paid', data['settle_cash']),
        ('Settlement Old-Balance Deducted', data['settle_deduct']),
    ]
    for i, (label, value) in enumerate(summary):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = bold
        val = value if isinstance(value, int) else f"₹{value:,.2f}"
        ws.cell(row=r, column=2, value=val)

    # --- Trips by load type ---
    r = 5 + len(summary) + 1
    ws.cell(row=r, column=1, value='TRIPS BY LOAD TYPE').font = _F(bold=True, color='2563EB')
    r += 1
    hdr_r = r
    for col, h in enumerate(['Load Type', 'Trips', 'Amount'], start=1):
        c = ws.cell(row=hdr_r, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
    r += 1
    for label, info in data['trip_by_type'].items():
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=info['count'])
        ws.cell(row=r, column=3, value=float(info['amount']))
        r += 1
    ws.cell(row=r, column=1, value='TOTAL').font = bold
    ws.cell(row=r, column=2, value=data['trip_count_total']).font = bold
    ws.cell(row=r, column=3, value=float(data['trip_total'])).font = bold

    # --- Day-wise rollup ---
    r += 2
    ws.cell(row=r, column=1, value='DAY-WISE ROLLUP').font = _F(bold=True, color='2563EB')
    r += 1
    hdr_r = r
    for col, h in enumerate(['Date', 'Trips', 'Extra', 'Advance'], start=1):
        c = ws.cell(row=hdr_r, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
    r += 1
    for d in data['days']:
        row = data['day_rollup'][d]
        ws.cell(row=r, column=1, value=d.strftime('%d-%b-%Y'))
        ws.cell(row=r, column=2, value=float(row['trips']))
        ws.cell(row=r, column=3, value=float(row['extra']))
        ws.cell(row=r, column=4, value=float(row['advance']))
        r += 1
    ws.cell(row=r, column=1, value='TOTAL').font = bold
    for col, val in [(2, data['grand_trips']), (3, data['grand_extra']), (4, data['grand_advance'])]:
        ws.cell(row=r, column=col, value=float(val)).font = bold

    # --- Transactions detail ---
    def detail(table, headers, rows):
        nonlocal r
        r += 2
        ws.cell(row=r, column=1, value=table).font = _F(bold=True, color='2563EB')
        r += 1
        hdr_r = r
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=hdr_r, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
        r += 1
        for row in rows:
            for col, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
            r += 1

    trip_rows = []
    for g in data['trips']:
        names = ', '.join(g.labourers.values_list('name', flat=True))
        trip_rows.append([
            g.date.strftime('%d-%b-%Y'), g.load_label or 'Tractor Trip',
            g.trip_count, float(g.rate_per_trip), float(g.total_amount),
            names, g.note,
        ])
    detail('TRIPS DETAIL',
           ['Date', 'Load Type', 'Trips', 'Rate', 'Amount', 'Workers', 'Note'],
           trip_rows)

    detail('EXTRAS DETAIL',
           ['Date', 'Worker', 'Amount', 'Note'],
           [[e.date.strftime('%d-%b-%Y'), e.labour.name, float(e.amount), e.note] for e in data['extras']])

    detail('ADVANCES DETAIL',
           ['Date', 'Worker', 'Amount', 'Note'],
           [[a.date.strftime('%d-%b-%Y'), a.labour.name, float(a.amount), a.note] for a in data['advances']])

    detail('DRIVER PAYMENTS DETAIL',
           ['Worker', 'Period', 'Amount', 'Note'],
           [[d.labour.name, f"{d.period_start:%d-%b}→{d.period_end:%d-%b}", float(d.amount), d.note] for d in data['driver_pmts']])

    detail('SETTLEMENTS DETAIL',
           ['Date', 'Worker', 'Period', 'Salary', 'Advance', 'Old-Deduct', 'Cash Paid'],
           [[s.settlement_date.strftime('%d-%b-%Y'), s.labour.name,
             f"{s.period_start:%d-%b}→{s.period_end:%d-%b}",
             float(s.total_salary), float(s.total_advance),
             float(s.old_balance_deducted), float(s.cash_paid)]
            for s in data['settlements']])

    # Column widths
    for col, w in {'A': 16, 'B': 24, 'C': 14, 'D': 14, 'E': 14, 'F': 26, 'G': 20}.items():
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _daily_activity_csv(data, filename='daily_activity.csv'):
    import csv as _csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = _csv.writer(response)
    writer.writerow(['DAILY ACTIVITY', data['period_label']])
    writer.writerow([])
    writer.writerow(['SUMMARY'])
    for label, value in [
        ('Trip Trips Count', data['trip_count_total']),
        ('Trip Amount', float(data['trip_total'])),
        ('Extra Amount', float(data['extra_total'])),
        ('Advance Amount', float(data['advance_total'])),
        ('Driver Payment', float(data['driver_total'])),
        ('Settlement Cash Paid', float(data['settle_cash'])),
        ('Settlement Old-Balance Deducted', float(data['settle_deduct'])),
    ]:
        writer.writerow([label, value])
    writer.writerow([])
    writer.writerow(['TRIPS BY LOAD TYPE'])
    writer.writerow(['Load Type', 'Trips', 'Amount'])
    for label, info in data['trip_by_type'].items():
        writer.writerow([label, info['count'], float(info['amount'])])
    writer.writerow(['TOTAL', data['trip_count_total'], float(data['trip_total'])])
    writer.writerow([])
    writer.writerow(['DAY-WISE ROLLUP'])
    writer.writerow(['Date', 'Trips', 'Extra', 'Advance'])
    for d in data['days']:
        row = data['day_rollup'][d]
        writer.writerow([d.strftime('%d-%b-%Y'), float(row['trips']), float(row['extra']), float(row['advance'])])
    writer.writerow(['TOTAL', float(data['grand_trips']), float(data['grand_extra']), float(data['grand_advance'])])
    writer.writerow([])
    writer.writerow(['TRIPS DETAIL'])
    writer.writerow(['Date', 'Load Type', 'Trips', 'Rate', 'Amount', 'Workers', 'Note'])
    for g in data['trips']:
        names = ', '.join(g.labourers.values_list('name', flat=True))
        writer.writerow([g.date.strftime('%d-%b-%Y'), g.load_label or 'Tractor Trip',
                         g.trip_count, float(g.rate_per_trip), float(g.total_amount), names, g.note])
    writer.writerow([])
    writer.writerow(['EXTRAS DETAIL'])
    writer.writerow(['Date', 'Worker', 'Amount', 'Note'])
    for e in data['extras']:
        writer.writerow([e.date.strftime('%d-%b-%Y'), e.labour.name, float(e.amount), e.note])
    writer.writerow([])
    writer.writerow(['ADVANCES DETAIL'])
    writer.writerow(['Date', 'Worker', 'Amount', 'Note'])
    for a in data['advances']:
        writer.writerow([a.date.strftime('%d-%b-%Y'), a.labour.name, float(a.amount), a.note])
    writer.writerow([])
    writer.writerow(['DRIVER PAYMENTS DETAIL'])
    writer.writerow(['Worker', 'Period', 'Amount', 'Note'])
    for d in data['driver_pmts']:
        writer.writerow([d.labour.name, f"{d.period_start:%d-%b}→{d.period_end:%d-%b}", float(d.amount), d.note])
    writer.writerow([])
    writer.writerow(['SETTLEMENTS DETAIL'])
    writer.writerow(['Date', 'Worker', 'Period', 'Salary', 'Advance', 'Old-Deduct', 'Cash Paid'])
    for s in data['settlements']:
        writer.writerow([s.settlement_date.strftime('%d-%b-%Y'), s.labour.name,
                         f"{s.period_start:%d-%b}→{s.period_end:%d-%b}",
                         float(s.total_salary), float(s.total_advance),
                         float(s.old_balance_deducted), float(s.cash_paid)])
    return response


@login_required(login_url='/login/')
def _resolve_period(request, today):
    """Resolve from_date/to_date from query params, honouring a month_preset shortcut."""
    mp = request.GET.get('month_preset')
    if mp:
        try:
            m = int(mp)
            if 1 <= m <= 12:
                import calendar
                last = calendar.monthrange(today.year, m)[1]
                return today.replace(month=m, day=1), today.replace(month=m, day=last)
        except (ValueError, TypeError):
            pass
    default_start = today.replace(day=1)
    period_start = _parse_date(request.GET.get('from_date'), default_start)
    period_end = _parse_date(request.GET.get('to_date'), today)
    if period_start and period_end and period_start > period_end:
        period_start, period_end = period_end, period_start
    return period_start, period_end


def daily_activity(request):
    """Daily Activity page + Excel/CSV exports (month-wise or date range)."""
    today = timezone.localdate()
    period_start, period_end = _resolve_period(request, today)

    data = _daily_activity_data(period_start, period_end)

    export = request.GET.get('export')
    if export == 'excel':
        return _daily_activity_excel(data)
    if export == 'csv':
        return _daily_activity_csv(data)

    return render(request, 'labour/daily_activity.html', {
        'data': data,
        'from_date': period_start,
        'to_date': period_end,
        'page_title': 'Daily Activity',
        'months': ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                   'August', 'September', 'October', 'November', 'December'],
    })


@login_required(login_url='/login/')
def labour_statement_export(request, labour_id):
    """Ek single labour ka statement PDF/Excel — detail page se."""
    labour = get_object_or_404(Labour, pk=labour_id)
    today = timezone.localdate()
    period_start = _parse_date(request.GET.get('from_date'), today.replace(day=1))
    period_end = _parse_date(request.GET.get('to_date'), today)
    if period_start and period_end and period_start > period_end:
        period_start, period_end = period_end, period_start

    export = request.GET.get('export', 'pdf')
    st = _labour_statement_for_period(labour, period_start, period_end)

    safe_name = re.sub(r'[^A-Za-z0-9_-]+', '_', (labour.name or 'labour')).strip('_')

    from core.pdf_utils import get_indian_current_time_str
    period_label = (
        f"{period_start:%d-%b-%Y} to {period_end:%d-%b-%Y}"
        if period_start and period_end else 'All Transactions'
    )

    if export == 'excel':
        return _labour_book_excel(
            [st], period_start, period_end,
            filename=f'labour_statement_{safe_name}.xlsx',
        )

    return _labour_book_pdf(
        [st], period_start, period_end,
        filename=f'labour_statement_{safe_name}.pdf',
        report_title="Labour Account Statement",
        report_subtitle=(
            f"{labour.name} · {period_label} · "
            f"Generated on {get_indian_current_time_str()}"
        ),
        include_grand_summary=False,
    )


def _labour_book_excel(statements, period_start, period_end, filename='labour_book.xlsx'):
    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _A
    from openpyxl.utils import get_column_letter as gcl

    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = _PF(start_color='16665A', end_color='16665A', fill_type='solid')
    header_font = _F(color='FFFFFF', bold=True)
    total_font = _F(bold=True)
    total_fill = _PF(start_color='EEF7F5', end_color='EEF7F5', fill_type='solid')

    if not statements:
        ws = workbook.create_sheet(title='No Data')
        ws.cell(row=1, column=1, value='No active labour found.').font = _F(bold=True)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    period_label = (
        f"{period_start:%d-%b-%Y} to {period_end:%d-%b-%Y}"
        if period_start and period_end else 'All Transactions'
    )

    for st in statements:
        labour = st['labour']
        sheet_name = ''.join(
            c for c in labour.name if c not in r'[]:*?/\\'
        )[:31] or 'Labour'

        ws = workbook.create_sheet(title=sheet_name)

        # ---- Title bar ----
        ws.cell(row=1, column=1, value=f"{labour.name} — Statement").font = _F(bold=True, size=13)

        # ---- Info (simple A|B) ----
        info = [
            ('Labour', labour.name or '—'),
            ('Mobile', labour.mobile or '—'),
            ('Period', period_label),
        ]
        r = 3
        for label, val in info:
            ws.cell(row=r, column=1, value=label).font = _F(bold=True)
            ws.cell(row=r, column=2, value=val)
            r += 1

        # ---- SUMMARY (vertical, clean) ----
        summary_start = r + 1
        summary_items = [
            ('Trip Wages', st['trip_total'], False),
            ('Extra', st['extra_total'], False),
            ('Driver Payment', st['driver_total'], False),
            ('Total Salary', st['total_salary'], False),
            ('Advance', st['advance_total'], False),
            ('Payment', st['payment'], False),
            ('Old Balance', st['old_balance'], False),
            ('FINAL AMOUNT', st['final_amount'], True),
        ]
        ws.cell(row=summary_start, column=1, value='SUMMARY').font = _F(bold=True, size=11)
        rr = summary_start + 1
        for label, value, bold in summary_items:
            ws.cell(row=rr, column=1, value=label).font = _F(bold=True)
            vcell = ws.cell(row=rr, column=2, value=float(value))
            vcell.number_format = '#,##0.00'
            if bold:
                vcell.font = _F(bold=True, size=11, color='16665A')
                ws.cell(row=rr, column=1).font = _F(bold=True, size=11, color='16665A')
                for cc in (1, 2):
                    ws.cell(row=rr, column=cc).fill = total_fill
            rr += 1

        # ---- ENTRIES (single list: trips + extra + advance) ----
        extra_by_date = {row['date']: row['extra_amount'] for row in st['rows']}
        advance_by_date = {row['date']: row['advance_amount'] for row in st['rows']}

        entries_start = rr + 1
        ws.cell(row=entries_start, column=1, value='ENTRIES').font = _F(bold=True, size=11)
        head_row = entries_start + 1
        headers = ['Date', 'Description', 'Trips', 'Rate', 'Trip ₹', 'Extra ₹', 'Advance ₹', 'Total ₹']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=head_row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = _A(horizontal='center', vertical='center')

        rr = head_row + 1
        trip_total = Decimal('0')
        extras_done = set()
        if st.get('trip_groups'):
            for g in st['trip_groups']:
                d = g.date
                first_of_day = d not in extras_done
                extras_done.add(d)
                extra = extra_by_date.get(d, Decimal('0')) if first_of_day else Decimal('0')
                advance = advance_by_date.get(d, Decimal('0')) if first_of_day else Decimal('0')
                desc = g.load_label or 'Tractor Trip'
                grand = g.total_amount + extra + advance
                ws.cell(row=rr, column=1, value=d.strftime('%d-%b-%Y'))
                ws.cell(row=rr, column=2, value=desc)
                ws.cell(row=rr, column=3, value=g.trip_count)
                ws.cell(row=rr, column=4, value=float(g.rate_per_trip)).number_format = '#,##0'
                ws.cell(row=rr, column=5, value=float(g.total_amount)).number_format = '#,##0'
                ws.cell(row=rr, column=6, value=float(extra)).number_format = '#,##0'
                ws.cell(row=rr, column=7, value=float(advance)).number_format = '#,##0'
                ws.cell(row=rr, column=8, value=float(grand)).font = _F(bold=True)
                ws.cell(row=rr, column=8).number_format = '#,##0'
                trip_total += g.total_amount
                rr += 1

        # Show extra/advance-only days (days with no trip entry)
        for row in st['rows']:
            d = row['date']
            if not any(g.date == d for g in (st.get('trip_groups') or [])):
                if row['extra_amount'] or row['advance_amount']:
                    grand = row['extra_amount'] + row['advance_amount']
                    ws.cell(row=rr, column=1, value=d.strftime('%d-%b-%Y'))
                    ws.cell(row=rr, column=2, value='Extra / Advance')
                    ws.cell(row=rr, column=6, value=float(row['extra_amount'])).number_format = '#,##0'
                    ws.cell(row=rr, column=7, value=float(row['advance_amount'])).number_format = '#,##0'
                    ws.cell(row=rr, column=8, value=float(grand)).font = _F(bold=True)
                    ws.cell(row=rr, column=8).number_format = '#,##0'
                    rr += 1

        ws.cell(row=rr, column=7, value='TOTAL').font = total_font
        ws.cell(row=rr, column=8, value=float(st['total_salary'] - st['driver_total'])).font = total_font
        ws.cell(row=rr, column=8).number_format = '#,##0'
        for col in range(1, 9):
            ws.cell(row=rr, column=col).fill = total_fill

        # ---- SETTLEMENTS (simple, optional) ----
        if st['settlements']:
            ss = rr + 2
            ws.cell(row=ss, column=1, value='SETTLEMENTS').font = _F(bold=True, size=11)
            head = ss + 1
            for col, h in enumerate(['Date', 'Cash Paid'], start=1):
                cell = ws.cell(row=head, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
            rr = head + 1
            for s in st['settlements']:
                ws.cell(row=rr, column=1, value=s.settlement_date.strftime('%d-%b-%Y'))
                ws.cell(row=rr, column=2, value=float(s.cash_paid)).number_format = '#,##0'
                rr += 1

        for col, width in {'A': 14, 'B': 22, 'C': 10, 'D': 10, 'E': 12, 'F': 12, 'G': 12, 'H': 12}.items():
            ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def _labour_book_pdf(statements, period_start, period_end, filename='labour_book.pdf', report_title='Labour Statements Book', report_subtitle=None, include_grand_summary=True):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    from core.pdf_utils import (
        get_registered_font,
        build_pdf_header_elements,
        get_indian_current_time_str,
        build_summary_cards,
        apply_data_table_style,
        finish_document,
        build_thankyou_note,
        get_pdf_styles,
        BRAND,
        BRAND_DARK,
    )
    font_name = get_registered_font()

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=9 * mm,
        leftMargin=9 * mm,
        topMargin=11 * mm,
        bottomMargin=11 * mm,
    )

    styles = get_pdf_styles(font_name)

    def _wrapped_style(name, **kw):
        return ParagraphStyle(
            name, parent=getSampleStyleSheet()['Normal'],
            fontName=font_name, **kw,
        )

    if period_start and period_end:
        book_period = f"{period_start:%d-%b-%Y} to {period_end:%d-%b-%Y}"
    else:
        book_period = 'All Transactions'

    elements = build_pdf_header_elements(
        font_name=font_name,
        report_title=report_title,
        report_subtitle=(
            report_subtitle
            if report_subtitle is not None
            else (
                f"All labour · {book_period} · "
                f"Generated on {get_indian_current_time_str()}"
            )
        ),
    )

    total_salary = Decimal('0')
    total_advance = Decimal('0')
    total_opening = Decimal('0')
    total_final = Decimal('0')

    for i, st in enumerate(statements):
        labour = st['labour']

        # ---------- 1. IDENTITY BAND (clean, boxed, professional) ----------
        name_style = _wrapped_style('LName', fontSize=13, leading=17, textColor=BRAND_DARK)
        field_lbl = _wrapped_style('LFieldLbl', fontSize=6.5, leading=8.5, textColor=colors.HexColor('#94a3b8'))
        field_val = _wrapped_style('LFieldVal', fontSize=9, leading=12, textColor=colors.HexColor('#0f172a'))

        field_rows = []
        for lbl, val in [
            ('Type', 'Driver' if labour.is_driver else 'Labour'),
            ('Mobile', labour.mobile or '—'),
            ('Period', book_period),
        ]:
            field_rows.append([
                Paragraph(lbl.upper(), field_lbl),
                Paragraph(f'<b>{val}</b>', field_val),
            ])
        fields = Table(field_rows, colWidths=[22 * mm, 48 * mm])
        fields.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ]))

        identity = Table(
            [[Paragraph(f'<b>{labour.name}</b>', name_style), fields]],
            colWidths=['58%', '42%'],
        )
        identity.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fbfcfd')),
            ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#e2e8f0')),
            ('LINEAFTER', (0, 0), (0, 0), 0.8, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (0, 0), 16),
            ('RIGHTPADDING', (1, 0), (1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))

        head_block = [identity, Spacer(1, 10)]

        # ---------- 2. KPI CARDS (4 + 4, readable) ----------
        earnings_cards = build_summary_cards(
            [
                {'label': 'Total Trip Wages', 'value': f"₹{st['trip_total']:,.2f}", 'color': '#16665a'},
                {'label': 'Total Extra', 'value': f"₹{st['extra_total']:,.2f}", 'color': '#2563eb'},
                {'label': 'Driver Payment', 'value': f"₹{st['driver_total']:,.2f}", 'color': '#7c3aed'},
                {'label': 'Total Salary', 'value': f"₹{st['total_salary']:,.2f}", 'color': '#0f766e'},
            ],
            font_name=font_name,
        )

        settlement_cards = build_summary_cards(
            [
                {'label': 'Total Advance', 'value': f"₹{st['advance_total']:,.2f}", 'color': '#dc2626'},
                {'label': 'Payment (Salary - Adv)', 'value': f"₹{st['payment']:,.2f}", 'color': '#075985'},
                {'label': 'Old Balance', 'value': f"₹{st['old_balance']:,.2f}", 'color': '#b45309'},
                {
                    'label': 'Final Amount',
                    'value': (
                        f"₹{st['final_amount']:,.2f}"
                        if st['final_amount'] >= 0
                        else f"Outstanding ₹{abs(st['final_amount']):,.2f}"
                    ),
                    'color': '#dc2626' if st['final_amount'] < 0 else '#059669',
                    'sub': (
                        'Labour owes owner' if st['final_amount'] < 0
                        else 'To pay labour'
                    ),
                },
            ],
            font_name=font_name,
        )

        head_block.append(earnings_cards)
        head_block.append(Spacer(1, 6))
        head_block.append(settlement_cards)
        head_block.append(Spacer(1, 8))

        elements.append(KeepTogether(head_block))

        # ---------- 3. ENTRIES (single table: trips + extra/bhatta + advance) ----------
        extra_col_label = 'Bhatta (₹)' if labour.category == 'HYVA_DRIVER' else 'Extra (₹)'
        extra_by_date = {row['date']: row['extra_amount'] for row in st['rows']}
        advance_by_date = {row['date']: row['advance_amount'] for row in st['rows']}

        entries_data = [
            [
                Paragraph('<b>Date</b>', styles['header']),
                Paragraph('<b>Work Done</b>', styles['header']),
                Paragraph('<b>Trips</b>', styles['header_r']),
                Paragraph('<b>Rate (₹)</b>', styles['header_r']),
                Paragraph('<b>Trip ₹</b>', styles['header_r']),
                Paragraph(f'<b>{extra_col_label}</b>', styles['header_r']),
                Paragraph('<b>Advance ₹</b>', styles['header_r']),
                Paragraph('<b>Total ₹</b>', styles['header_r']),
            ]
        ]

        extras_done = set()
        if st.get('trip_groups'):
            for g in st['trip_groups']:
                d = g.date
                first_of_day = d not in extras_done
                extras_done.add(d)
                extra = extra_by_date.get(d, Decimal('0')) if first_of_day else Decimal('0')
                advance = advance_by_date.get(d, Decimal('0')) if first_of_day else Decimal('0')
                desc = (g.load_label or 'Tractor Trip')
                if labour.category == 'HYVA_DRIVER':
                    desc = 'Hyva' + (f' ({g.load_label})' if g.load_label else '')
                grand = g.total_amount + extra + advance
                entries_data.append([
                    Paragraph(d.strftime('%d-%b-%Y'), styles['body']),
                    Paragraph(desc, styles['body']),
                    Paragraph(str(g.trip_count), styles['body_r']),
                    Paragraph(f"₹{g.rate_per_trip:,.2f}", styles['body_r']),
                    Paragraph(f"₹{g.total_amount:,.2f}", styles['body_r']),
                    Paragraph(f"₹{extra:,.2f}", styles['body_r']),
                    Paragraph(f"₹{advance:,.2f}", styles['body_r']),
                    Paragraph(f"<b>₹{grand:,.2f}</b>", styles['body_r']),
                ])

        # Extra/advance-only days (no trip entry)
        for row in st['rows']:
            d = row['date']
            if not any(g.date == d for g in (st.get('trip_groups') or [])):
                if row['extra_amount'] or row['advance_amount']:
                    desc = 'Bhatta' if labour.category == 'HYVA_DRIVER' else 'Extra'
                    grand = row['extra_amount'] + row['advance_amount']
                    entries_data.append([
                        Paragraph(d.strftime('%d-%b-%Y'), styles['body']),
                        Paragraph(desc, styles['body']),
                        Paragraph('', styles['body_r']),
                        Paragraph('', styles['body_r']),
                        Paragraph('', styles['body_r']),
                        Paragraph(f"₹{row['extra_amount']:,.2f}", styles['body_r']),
                        Paragraph(f"₹{row['advance_amount']:,.2f}", styles['body_r']),
                        Paragraph(f"<b>₹{grand:,.2f}</b>", styles['body_r']),
                    ])

        entries_data.append([
            Paragraph('<b>TOTAL</b>', styles['body']),
            Paragraph('', styles['body']),
            Paragraph('', styles['body']),
            Paragraph('', styles['body']),
            Paragraph('', styles['body']),
            Paragraph(f"<b>₹{st['extra_total']:,.2f}</b>", styles['body_r']),
            Paragraph(f"<b>₹{st['advance_total']:,.2f}</b>", styles['body_r']),
            Paragraph(f"<b>₹{(st['total_salary'] - st['driver_total']):,.2f}</b>", styles['body_r']),
        ])
        entries_table = Table(
            entries_data,
            repeatRows=1,
            colWidths=[24 * mm, 28 * mm, 12 * mm, 22 * mm, 24 * mm, 24 * mm, 24 * mm, 24 * mm],
        )
        apply_data_table_style(entries_table, total_row=True)
        elements.append(entries_table)

        # ---------- 4. SETTLEMENTS IN RANGE ----------
        if st['settlements']:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph('<b>Settlements in this period</b>', _wrapped_style('SettleHead', fontSize=9.5, leading=12, textColor=BRAND_DARK)))
            s_data = [
                [
                    Paragraph('<b>Date</b>', styles['header']),
                    Paragraph('<b>Period</b>', styles['header']),
                    Paragraph('<b>Cash Paid (₹)</b>', styles['header_r']),
                    Paragraph('<b>New Old Balance (₹)</b>', styles['header_r']),
                ]
            ]
            for s in st['settlements']:
                s_data.append([
                    Paragraph(s.settlement_date.strftime('%d-%b-%Y'), styles['body']),
                    Paragraph(f"{s.period_start:%d-%b} to {s.period_end:%d-%b}", styles['body']),
                    Paragraph(f"₹{s.cash_paid:,.2f}", styles['body_r']),
                    Paragraph(f"₹{s.final_old_balance:,.2f}", styles['body_r']),
                ])
            s_table = Table(s_data, repeatRows=1, colWidths=[34 * mm, 44 * mm, 56 * mm, 58 * mm])
            apply_data_table_style(s_table, total_row=False)
            elements.append(s_table)

        total_salary += st['total_salary']
        total_advance += st['advance_total']
        total_opening += st['old_balance']
        total_final += st['final_amount']

        if i < len(statements) - 1:
            elements.append(Spacer(1, 22))
            elements.append(PageBreak())

    if include_grand_summary:
        grand_cards = build_summary_cards(
            [
                {'label': 'Total Salary', 'value': f"₹{total_salary:,.2f}", 'color': '#0f766e'},
                {'label': 'Total Advance', 'value': f"₹{total_advance:,.2f}", 'color': '#dc2626'},
                {'label': 'Total Old Balance', 'value': f"₹{total_opening:,.2f}", 'color': '#b45309'},
                {
                    'label': 'Total Final Amount',
                    'value': (
                        f"₹{total_final:,.2f}"
                        if total_final >= 0 else f"Outstanding ₹{abs(total_final):,.2f}"
                    ),
                    'color': '#dc2626' if total_final < 0 else '#059669',
                },
            ],
            font_name=font_name,
        )
        elements.append(Spacer(1, 10))
        elements.append(Paragraph('<b>GRAND SUMMARY — All Labour</b>', _wrapped_style('GrandHead', fontSize=9.5, leading=12, textColor=BRAND_DARK)))
        elements.append(grand_cards)

    elements.extend(build_thankyou_note(
        "Books prepared from the labour ledger. For any query, contact us.",
        font_name=font_name,
    ))

    finish_document(document, elements, font_name=font_name)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
