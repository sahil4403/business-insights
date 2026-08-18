from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from labour.models import Labour


TRIP_FILES = [
    "historical_data/Feb_to_Dec_Hyva_White_Sand_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Hyva_Fly_Ash_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Halfton_White_Sand_CLEAN.xlsx",
    "historical_data/Feb_to_Dec_Tractor_White_Sand_CLEAN.xlsx",
]


def clean(value):
    if value is None:
        return ""

    return str(value).strip()


class Command(BaseCommand):

    help = "Preview labour names found in Feb-Dec trip files but missing from DB."

    def handle(self, *args, **options):

        existing_labour = {
            labour.name.strip().lower(): labour.name
            for labour in Labour.objects.all()
        }

        excel_labour = {}

        for file_path in TRIP_FILES:

            path = Path(file_path)

            self.stdout.write(
                f"Reading: {path.name}"
            )

            wb = openpyxl.load_workbook(
                path,
                data_only=True,
            )

            for ws in wb.worksheets:

                headers = {
                    clean(ws.cell(1, col).value): col
                    for col in range(
                        1,
                        ws.max_column + 1
                    )
                }

                driver_columns = [
                    header
                    for header in (
                        "Driver 1",
                        "Driver 2",
                        "Driver 3",
                    )
                    if header in headers
                ]

                if not driver_columns:
                    continue

                for row in range(
                    2,
                    ws.max_row + 1
                ):

                    for driver_column in driver_columns:

                        driver_name = clean(
                            ws.cell(
                                row,
                                headers[driver_column]
                            ).value
                        )

                        if not driver_name:
                            continue

                        key = driver_name.lower()

                        if key not in excel_labour:

                            excel_labour[key] = driver_name

        missing_labour = {
            key: name
            for key, name in excel_labour.items()
            if key not in existing_labour
        }

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("FEBRUARY - DECEMBER LABOUR PREVIEW")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Labour names in Excel : {len(excel_labour)}"
        )

        self.stdout.write(
            f"Labour in DB          : {len(existing_labour)}"
        )

        self.stdout.write(
            f"Missing labour        : {len(missing_labour)}"
        )

        self.stdout.write("")
        self.stdout.write("MISSING LABOUR")
        self.stdout.write("-" * 70)

        for name in sorted(
            missing_labour.values(),
            key=str.lower
        ):
            self.stdout.write(name)

        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write(
            "PREVIEW ONLY - NO DATABASE CHANGES WERE MADE."
        )
        self.stdout.write("=" * 70)